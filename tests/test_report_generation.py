import os
import tempfile
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from reporting import data as data_module
from reporting import xlreport as xlreport_module
from SCPI_Library.simulation import reset_simulation
from execution.run_storage import create_run_storage


class ReportGenerationTests(unittest.TestCase):
    def setUp(self):
        reset_simulation()

    def _create_storage(self, base_directory, mode):
        storage = create_run_storage(base_directory, f"{mode}_SIMULATION")
        data_module.configure_run_storage(storage.raw, storage.charts)
        (storage.root / "SIMULATION_RUN.txt").write_text(
            "Simulated report integration test.\n", encoding="utf-8"
        )
        pd.DataFrame(
            {
                "Parameter": ["simulation_mode", "unit"],
                "Value": [True, mode],
            }
        ).to_csv(storage.raw / "config.csv", index=False)
        return storage

    def _assert_report_artifacts(self, storage, report_prefix):
        expected_raw_files = {
            "config.csv",
            "data.csv",
            "error.csv",
            "instrumentData.csv",
        }
        self.assertTrue(expected_raw_files.issubset(
            {path.name for path in storage.raw.iterdir()}
        ))
        self.assertTrue((storage.charts / "Chart.png").is_file())
        self.assertTrue((storage.charts / "Chart2.png").is_file())
        self.assertTrue((storage.root / "SIMULATION_RUN.txt").is_file())

        reports = list(storage.reports.glob(f"{report_prefix}_*.xlsx"))
        self.assertEqual(len(reports), 1)
        workbook = load_workbook(reports[0])
        worksheet = workbook["Data"]
        self.assertEqual(len(worksheet._images), 2)
        values = {
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        self.assertIn("simulation_mode", values)

    def test_generates_simulated_voltage_report(self):
        with tempfile.TemporaryDirectory() as temporary_directory, patch.dict(
            os.environ, {"AUTOMATION_SIMULATION": "1"}, clear=False
        ):
            storage = self._create_storage(temporary_directory, "VOLTAGE")
            info = [[5.0, 1.0, 0]]
            measured = [[5.0, 0.0]]
            readback = [[5.0, 1.0]]

            data_module.instrumentData(
                "USB0::SIM::PSU::INSTR",
                "USB0::SIM::DMM::INSTR",
                "USB0::SIM::ELOAD::INSTR",
            )
            graph = data_module.datatoGraph(info, measured, readback)
            graph.scatterCompareVoltage(
                0.001, 0.001, 0.001, 0.001, "VOLTAGE", 5.0
            )
            report = xlreport_module.xlreport(
                storage.reports, "SIMULATED_VOLTAGE"
            )
            report.run()

            self._assert_report_artifacts(storage, "SIMULATED_VOLTAGE")

    def test_generates_simulated_current_report(self):
        with tempfile.TemporaryDirectory() as temporary_directory, patch.dict(
            os.environ, {"AUTOMATION_SIMULATION": "1"}, clear=False
        ):
            storage = self._create_storage(temporary_directory, "CURRENT")
            info = [[5.0, 2.0, 0]]
            measured = [[0.0, 2.0]]
            readback = [[0.0, 2.0]]

            data_module.instrumentData(
                "USB0::SIM::PSU::INSTR",
                "USB0::SIM::DMM2::INSTR",
                "USB0::SIM::ELOAD::INSTR",
            )
            graph = data_module.datatoGraph2(info, measured, readback)
            graph.scatterCompareCurrent2(
                0.001, 0.001, 0.001, 0.001, "CURRENT", 2.0
            )
            report = xlreport_module.xlreport(
                storage.reports, "SIMULATED_CURRENT"
            )
            report.run()

            self._assert_report_artifacts(storage, "SIMULATED_CURRENT")

    def test_current_percentage_graph_sweeps_current_horizontally(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            storage = self._create_storage(temporary_directory, "CURRENT")
            info = [[1.0, current, 0] for current in (1.0, 2.0, 3.0)]
            measured = [[0.0, current] for current in (1.0, 2.0, 3.0)]
            readback = [[1.0, current] for current in (1.0, 2.0, 3.0)]
            graph = data_module.datatoGraph2(info, measured, readback)

            with patch(
                "matplotlib.axes.Axes.plot",
                autospec=True,
                return_value=[],
            ) as plot:
                graph.scatterCompareCurrent2(
                    0.001,
                    0.001,
                    0.001,
                    0.001,
                    "CURRENT",
                    10.0,
                )

            voltage_series = [
                call
                for call in plot.call_args_list
                if call.kwargs.get("label") == "Voltage = 1.0"
            ]

        self.assertEqual(len(voltage_series), 4)
        for call in voltage_series:
            self.assertEqual(list(call.args[1]), [1.0, 2.0, 3.0])


if __name__ == "__main__":
    unittest.main()
