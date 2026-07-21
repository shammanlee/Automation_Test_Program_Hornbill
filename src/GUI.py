"""Main Module that runs the Graphical User Interface (GUI) that is the main point of interaction between user and the program"""

import sys
import  shutil
import traceback
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pyvisa
from SCPI_Library.simulation import (
    create_resource_manager,
    initialize_main_thread_visa,
    is_simulation_mode,
)

if __name__ == "__main__":
    main_thread_resource_manager = initialize_main_thread_visa()
    if "--visa-diagnostics" in sys.argv:
        print(f"VISA library: {main_thread_resource_manager.visalib}")
        print("VISA resources:")
        visa_resources = main_thread_resource_manager.list_resources()
        for visa_resource in visa_resources:
            print(f"  {visa_resource}")
        for visa_resource in visa_resources:
            if not visa_resource.upper().startswith("GPIB"):
                continue
            gpib_instrument = None
            try:
                gpib_instrument = main_thread_resource_manager.open_resource(
                    visa_resource
                )
                gpib_instrument.timeout = 10000
                gpib_instrument.write_termination = "\n"
                gpib_instrument.read_termination = "\n"
                gpib_instrument.clear()
                gpib_instrument.write("ID?")
                print(f"GPIB identity ({visa_resource}): {gpib_instrument.read()}")
            except Exception as exception:
                print(f"GPIB probe failed ({visa_resource}): {exception}")
                raise SystemExit(1) from exception
            finally:
                if gpib_instrument is not None:
                    gpib_instrument.close()
        raise SystemExit(0)

import pandas as pd

import time #Shamman changes
import csv  #Shamman changes

import datetime

import pyqtgraph as pg
from PIL import Image, ImageDraw, ImageFont
from pathlib import Path
from openpyxl import load_workbook

from tkinter import Tk, filedialog
import tkinter as tk
from tkinter import ttk
from time import sleep

from common.output_capture import my_result

from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore  import *
from common.path import (
    config_folder,
    csv_folder,
    DATA_CSV_PATH,
    ERROR_CSV_PATH,
    IMAGE_DIR,
    IMAGE_PATH,
    IMAGE_PATH_2,
    POWER_DATA_CSV_PATH,
    POWER_ERROR_CSV_PATH,
    POWER_IMAGE_PATH,
    setup_img_folder,
)
from ui.error_dialogs import show_error_dialog
from ui.documentation_tab import ProgramDocumentationTab, TestPatternsTab
from ui.keysight_command_tab import KeysightCommandTab
from common.output_logging import print_console_safe
from instruments.instrument_discovery import (
    get_all_visa_resources as NewGetVisaSCPIResources,
    get_visa_scpi_resources as GetVisaSCPIResources,
)

from DUT_Test_Scripts.Dolphin.DUT_Test import NewCurrentMeasurement, NewVoltageMeasurement
from instruments.TestVolt_HB_3458 import (
    VoltageCalibrationDialog as VoltageCalibration3458Dialog,
)

from DUT_Test_Scripts.DUT_screenshot import ScreenShotDialog


from DUT_Test_Scripts.Hornbill.Hornbill_DUT_Test_With_ELoad import (
    HornbillCurrentMeasurementwithELoad_IMON_200uA,
    HornbillCurrentMeasurementwithELoad_IMON_2mA,
    HornbillCurrentMeasurementwithELoad_IMON_FULL,
    HornbillVoltageCalibration,
    HornbillVoltageMeasurementwithELoad,
)

from DUT_Test_Scripts.Hornbill.Hornbill_DUT_Test_No_ELoad import (
    HornbillVoltageMeasurementNoELoad,)

from DUT_Test_Scripts.EDU36311A_DUT_Test_No_Load import (
    EDU36311AVoltageMeasurementNoELoad,)
#########################################################################################
from DUT_Test_Scripts.Dolphin.Dolphin_DUT_Test_With_ELoad_With_DMM import (
    DolphinLoadRegulationwithELoad,
    DolphinNewCurrentMeasurementwithELoad,
    DolphinNewVoltageMeasurementwithELoad,
    DolphinProgrammingResponse,
    DolphinRiseFallTimewithELoad,
)

from DUT_Test_Scripts.Dolphin.Dolphin_DUT_Test_No_ELoad_With_DMM import (
    DolphinNewCurrentMeasurementNoELoadWithDMM,
    DolphinNewVoltageMeasurementNoELoadWithDMM,
)

from DUT_Test_Scripts.Dolphin.Dolphin_DUT_Test_With_ELoad_No_DMM import (
    DolphinNewCurrentMeasurementwithELoadNoDMM,
    DolphinNewVoltageMeasurementwithELoadNoDMM,
)

from DUT_Test_Scripts.Dolphin.Dolphin_DUT_Test_No_ELoad_No_DMM import (
    ActivateAC,
    DolphinNewCurrentMeasurementNoELoadNoDMM,
    DolphinNewVoltageMeasurementNoELoadNoDMM,
    LoadRegulation,
    PowerMeasurement,
    RiseFallTime,
    VisaResourceManager,
    dictGenerator,
)
##########################################################################################

from reporting.data import (
    datatoCSV_Accuracy,
    datatoCSV_Accuracy2,
    datatoCSV_PowerAccuracy,
    datatoGraph,
    datatoGraph2,
    datatoGraph3,
    instrumentData,
    powerinstrumentData,
)
from reporting.xlreport import Graph_Plotting, xlreport
from reporting.xlreportpower import xlreportpower

from External_Auxiliary_Equipment.Relay_Control import (
    RelayController_Current,
    RelayController_Voltage,
)


x=0
desp_font = QFont("Times New Roman", 14, QFont.Bold) 
AdvancedSettingsList = []


def run_tabulated_excel_file(output_file, state, wb=None):
 
    if state == "open":
        wb = load_workbook(output_file)
        ws = wb["Data"]
   
        return wb,ws
    elif state == "close":
        wb.save(output_file)

#############################################################################
def resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller exe."""
    if hasattr(sys, "_MEIPASS"):  # Running in exe
        return str(Path(sys._MEIPASS) / relative_path)
    return str(Path(__file__).resolve().parent.parent / relative_path)

def classify_visa_resource(visa_id):   #Shamman changes (makes program slower due to scanning)
    if visa_id.startswith("USB"):
        return "USB"

    if visa_id.startswith("TCPIP"):
        address = visa_id.split("::")[1]
        try:
            import ipaddress
            ipaddress.ip_address(address)
            return "TCPIP_IP"
        except ValueError:
            return "TCPIP_HOSTNAME"

    return "OTHER"

# Scroll Area
def scroll_area(self,layout):
    # Create a QWidget that will be the content inside the QScrollArea
    content_widget = QWidget()
    content_widget.setLayout(layout)

    # Create the QScrollArea
    scroll_area = QScrollArea()
    scroll_area.setWidget(content_widget)
    scroll_area.setWidgetResizable(True)

    self.setLayout(QVBoxLayout())
    self.layout().addWidget(scroll_area)

    return scroll_area

###########################################################################
#------------------------- Main GUI window and application tabs--------------
class MainWindow(QMainWindow):
    """Main window with controls for displaying DUT tests."""
    
    def __init__(self):
        super().__init__()
        self.params = Parameters() 
        
        # Initialization/Font Set-Up
        window_title = "DUT TEST GUI"
        if is_simulation_mode():
            window_title += " - SIMULATION MODE"
        self.setWindowTitle(window_title)
        self.resize(1000, 700) 
        self.image_window = None
        self.setWindowFlags(Qt.Window)
        window_font = QFont("Arial", 12)
        self.setFont(window_font)

        # Track current tab index
        self.CurrentTab = 0

        self.dialog_registry = self._create_dialog_registry()
        self.test_options = self.dialog_registry.selection_options
        
        #Show the tab on the Main Window
        self.initTabs()
        if is_simulation_mode():
            simulation_message = "SIMULATION MODE - no real instruments are controlled"
            self.statusBar().showMessage(simulation_message)
            self.statusBar().setStyleSheet(
                "background-color: #b45309; color: white; font-weight: bold;"
            )

        # Main layout components
        QLabel_Widget = QLabel("Choose Test:")
        QLabel_Widget.setFont(QFont("Arial", 14, QFont.Bold))
        QLabel_Widget.setStyleSheet("color: #2c3e50; margin: 10px;")
        
        self.QButton_Widget = QPushButton("Confirm Selection")
        self.QButton_Widget.setFont(QFont("Arial", 14, QFont.Bold))
        self.QButton_Widget.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #229954;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)

        Main_Window_Layout = QVBoxLayout()
        Main_Window_Layout.addWidget(QLabel_Widget)
        Main_Window_Layout.addWidget(self.tab_widget)
        Main_Window_Layout.addWidget(self.QButton_Widget)
        Main_Window_Layout.setSpacing(10)
        Main_Window_Layout.setContentsMargins(20, 20, 20, 20)

        # Main page
        self.page_home = QWidget()
        self.page_home.setLayout(Main_Window_Layout)
        self.page_home.setStyleSheet("background-color: #f8f9fa;")
        self.setCentralWidget(self.page_home)
        
        self.tab_widget.currentChanged.connect(self.currentTabChanged)
        self.QButton_Widget.clicked.connect(self.PushBtnClicked)

    def initTabs(self):
        """Initialize tab widgets - merged from tab class"""
        # Create the main tab widget
        self.tab_widget = QTabWidget()
        
        # Initialize tab widgets
        self.tab_NewBundle = QWidget()
        self.tab_ScreenShot = QWidget()
        self.tab_TestList = QWidget()
        self.tab_Documentation = ProgramDocumentationTab()
        self.tab_TestPatterns = TestPatternsTab()
        self.tab_KeysightCommands = KeysightCommandTab()

        # Add tabs to the tab widget
        self.tab_widget.addTab(self.tab_NewBundle, "Bundle Test")
        self.tab_widget.addTab(self.tab_ScreenShot, "Screenshot")
        self.tab_widget.addTab(self.tab_TestList, "Test Selection")
        self.tab_widget.addTab(self.tab_Documentation, "Documentation")
        self.tab_widget.addTab(self.tab_TestPatterns, "Test Patterns")
        self.tab_widget.addTab(self.tab_KeysightCommands, "Keysight Commands")

        # Set larger font for tab labels
        tab_font = QFont("Arial", 16)
        self.tab_widget.setFont(tab_font)

        # Initialize UI for each tab
        self.NewBundleUI()
        self.ScreenShotUI()
        self.TestListUI()
  
    #################UI Settings for application tabs#########################
    def setupTabUI(self, tab, image_path, description, tab_index, tab_text, label_size=(800, 600)):
        """Setup UI for image-based tabs"""
        # Create a placeholder for image
        try:
            pixmap = QPixmap(image_path).scaled(*label_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            pixmap = pixmap.scaled(1280, 720, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        except:
            # Create a placeholder pixmap if image doesn't exist
            pixmap = QPixmap(800, 600)
            pixmap.fill(Qt.lightGray)

        label = QLabel()
        label.setPixmap(pixmap)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("border: 2px solid gray; background-color: #f0f0f0;")
        label.setText(f"Image Placeholder\n{image_path}" if pixmap.isNull() else "")

        # Set font for the description label
        description_font = QFont("Arial", 16)
        description_label = QLabel(description)
        description_label.setFont(description_font)
        description_label.setAlignment(Qt.AlignCenter)
        description_label.setWordWrap(True)
    
        layout = QVBoxLayout()
        layout.addWidget(label)
        layout.addWidget(description_label)
        tab.setLayout(layout)
        
        self.tab_widget.setTabText(tab_index, tab_text)

    def NewBundleUI(self):
        """Setup Bundle Test tab"""
        self.setupTabUI(
            self.tab_NewBundle,
            str(setup_img_folder / "2.png"),
            "Test GUI for Bundle Measurement\n\nThis tab provides voltage and current testing capabilities for DUT bundles.",
            0,
            "Bundle Test (Voltage/Current)"
        )

    def ScreenShotUI(self):
        """Setup Screenshot tab"""
        self.setupTabUI(
            self.tab_ScreenShot,
            str(setup_img_folder / "7.png"),
            "Instrument Display Capture\n\nCapture screenshots of the DUT instrument display for documentation and analysis.",
            1,
            "Instrument ScreenShot"
        )

    def TestListUI(self):
        """Create the third tab with a list of available test dialogs"""
        layout = QVBoxLayout()
        
        # Title label
        title_label = QLabel("Available Test Dialogs")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; margin: 10px;")
        
        # Description label
        desc_label = QLabel("Select from the list below to open different test dialogs:")
        desc_label.setFont(QFont("Arial", 16))
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setStyleSheet("color: #34495e; margin: 5px;")
        
        # Create list widget
        self.test_list = QListWidget()
        self.test_list.setFont(QFont("Arial", 16))
        self.test_list.setStyleSheet("""
            QListWidget {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                background-color: white;
                alternate-background-color: #f8f9fa;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #6145FF;
            }
        """)
        
        for i, (title, description,) in enumerate(self.test_options):
            item = QListWidgetItem()
            item.setText(f"{title}\n{description}")
            item.setData(Qt.UserRole, i + 2)  # Store the tab index
            self.test_list.addItem(item)
        
        # Connect double-click signal
        self.test_list.itemDoubleClicked.connect(self.on_test_selected)
        
        # Add select button
        select_btn = QPushButton("Open Selected Test")
        select_btn.setFont(QFont("Arial", 16, QFont.Bold))
        select_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        select_btn.clicked.connect(self.on_select_button_clicked)
        
        # Add widgets to layout
        layout.addWidget(title_label)
        layout.addWidget(desc_label)
        layout.addWidget(self.test_list)
        layout.addWidget(select_btn)
        
        self.tab_TestList.setLayout(layout)
        self.tab_widget.setTabText(2, "Test Selection")

    ##################Interactive Action #################################
    def on_test_selected(self, item):
        """Handle test selection from list"""
        test_index = item.data(Qt.UserRole)
        self.open_test_dialog(test_index)

    def on_select_button_clicked(self):
        """Handle select button click"""
        current_item = self.test_list.currentItem()
        if current_item:
            self.on_test_selected(current_item)

    def open_test_dialog(self, test_index):
        """Open the selected test dialog"""
        self.open_test_by_index(test_index)
    
    def currentTabChanged(self, index):
        """Slot to update the current tab index."""
        self.CurrentTab = index
        self.QButton_Widget.setVisible(index < 3)
        print_console_safe(f"Current tab changed to: {index}")

    def PushBtnClicked(self):
        """Open the appropriate dialog based on the selected tab."""
        if self.CurrentTab >= 3:
            return
        if self.CurrentTab < 2:  # For first two tabs, use existing logic
            self.open_test_by_index(self.CurrentTab)
        else:  # For third tab, get selection from list
            if hasattr(self, 'test_list'):
                current_item = self.test_list.currentItem()
                if current_item:
                    test_index = current_item.data(Qt.UserRole)
                    self.open_test_by_index(test_index)
                else:
                    print_console_safe("Please select a test from the list first.")

    def open_test_by_index(self, index):
        """Open a registered dialog by its stable launcher index."""
        return self.dialog_registry.open(self, index)

    def _create_dialog_registry(self):
        from ui.dialog_registry import DialogRegistration, DialogRegistry

        def multithread_voltage_dialog():
            from experiments.multithread_voltage import MultiThreadVoltageMeasurementDialog

            return MultiThreadVoltageMeasurementDialog()

        registrations = (
            DialogRegistration("Bundle Test", "Production queued test dialog", "bundle_dialog", AllTestMeasurement),
            DialogRegistration("Screenshot", "Instrument screenshot dialog", "screenshot_dialog", lambda: ScreenShotDialog(self)),
            DialogRegistration("Voltage Measurement", "Precise voltage measurement dialog - Support only - Dolphin, Excavator", "voltage_dialog", VoltageMeasurementDialog),
            DialogRegistration("Current Measurement", "Precise current measurement dialog - Support only - Dolphin, Excavator", "current_dialog", CurrentMeasurementDialog),
            DialogRegistration("CV Load Regulation", "Constant Voltage load regulation testing - Support only - Dolphin, Excavator", "CV_load_dialog", CV_LoadRegulationDialog),
            DialogRegistration("CC Load Regulation", "Constant Current load regulation testing - Support only - Dolphin, Excavator", "CC_load_dialog", CC_LoadRegulationDialog),
            DialogRegistration("Transient Recovery Time", "Measure transient response recovery time - Support only - Dolphin, Excavator", "transient_load_dialog", TransientRecoveryTime),
            DialogRegistration("Transient Recovery Time Using Current Probe", "Measure transient response recovery time using current probe - Support only - Dolphin, Excavator", "transient_current_probe_dialog", TransientRecoveryTimeWithCurrentSensor),
            DialogRegistration("Programming Speed", "Test programming speed capabilities - Support only - Dolphin, Excavator", "programming_speed_dialog", ProgrammingSpeed),
            DialogRegistration("Power Measurement", "Comprehensive power measurement dialog - Support only - Dolphin, Excavator", "power_dialog", PowerMeasurementDialog),
            DialogRegistration("Bundle Measurement - Voltage", "Specialized bundle voltage measurement - Support only - Dolphin, Excavator", "bundle_voltage_dialog", BundleMeasurementVoltageDialog),
            DialogRegistration("Bundle Measurement - Current/Power", "Bundle current and power measurement - Support only - Dolphin, Excavator", "bundle_current_dialog", BundleMeasurementCurrentandPowerDialog),
            DialogRegistration("AC Source Settings", "Simple Configuration of AC Source - Support only - Dolphin, Excavator", "ac_source_dialog", lambda: ACSourceSetting(self.params)),
            DialogRegistration("Voltage Calibration - 3458A", "Hornbill voltage calibration using a 3458A DMM", "voltage_calibration_dialog", VoltageCalibration3458Dialog),
            DialogRegistration("Peak Power Test - Passive Mode", "Peak Current Pulse - Support only - Hornbill", "peak_power_test_dialog", PeakPowerTestDialog),
            DialogRegistration("MultiThreading - Voltage Measurement", "Precise voltage measurement dialog", "multithread_voltage_dialog", multithread_voltage_dialog),
        )
        return DialogRegistry(registrations, print_console_safe)

#######------------------------Standalone Test Scripts in Tab 3-----------------------------#####################
class VoltageMeasurementDialog(QDialog):
    """Class for configuring the voltage measurement DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.


    """
    def __init__(self):
        """Method where Widgets, Signals and Slots are defined in the GUI for Voltage Measurement"""
        super().__init__()
        self.DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/data.csv"  # or use a dynamic path if needed
        self.IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/Chart.png"
        self.ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/error.csv"
        self.image_dialog = None
    
        self.Power_Rating = "6000"
        self.Current_Rating = "24"
        self.Voltage_Rating = "800"
        self.Power = "2000"
        self.currloop = "1"
        self.voltloop = "1"
        self.estimatetime = "0"
        self.readbackvoltage = "0"
        self.readbackcurrent = "0"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"
        self.noofloop = "1"
        self.updatedelay = "3.0"
        self.unit = "VOLTAGE"
        self.Programming_Error_Offset = "0.0003"
        self.Programming_Error_Gain = "0.0003"
        self.Readback_Error_Offset = "0.0003"
        self.Readback_Error_Gain = "0.0003"
        self.minCurrent = "0"
        self.maxCurrent = "24"
        self.current_step_size = "1"
        self.minVoltage = "0"
        self.maxVoltage = "800"
        self.voltage_step_size = "80"
        self.PSU = "USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"
        self.OSC = "USB0::0x0957::0x17B0::MY52060151::0::INSTR"
        self.DMM = "USB0::0x2A8D::0x1601::MY60094787::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"
        self.OSC_Channel = "1"
        self.DMM_Instrument = "Keysight"
        self.DUT = "Others"

        self.PSU_Channel = "1"
        self.ELoad_Channel = "1"
        self.setFunction = "Current" #Set Eload in CC Mode
        self.VoltageRes = "DEF"
        self.VoltageSense = "INT"
        self.SPOperationMode = "Independent"

        self.relay_voltage = "None"

        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.checkbox_test_VoltageAccuracy = 2
        self.checkbox_test_VoltageLoadRegulation = 2
        self.checkbox_test_TransientRecovery = 2

        self.Range = "Auto"
        self.Aperture = "MAX"
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.UpTime = "50"
        self.DownTime = "50"

        self.Channel_CouplingMode = "AC"
        self.Trigger_CouplingMode = "DC"
        self.Trigger_Mode = "EDGE"
        self.Trigger_SweepMode = "NORMAL"
        self.Trigger_SlopeMode = "EITHer"
        self.TimeScale = "0.01"
        self.VerticalScale = "0.00001"
        self.I_Step = ""
        self.V_Settling_Band = "0.8"
        self.T_Settling_Band = "0.001"
        self.Probe_Setting = "X10"
        self.Acq_Type = "AVERage"

        self.simulation_mode = 0
        self.output_file = r"C:\Users\shamlee3\OneDrive - Keysight Technologies\Wei Jing See's files - Shamman Xian Jun Lee\Test Data\Monk3.xlsx"  
        self.selected_text  = "Hornbill"
        self.stop_requested = False

        self.checkbox_SpecialCase = 2
        self.checkbox_NormalCase = 2

        self.checkbox_Voltage_Accuracy_Voltage_Mode = 2
        self.checkbox_Voltage_Accuracy_Current_Mode = 0
        
        #Create Voltage Window
        self.setWindowTitle("Voltage Measurement")
        self.image_window = None
        self.setWindowFlags(Qt.Window)
        font = QFont()
        font.setPointSize(16)   # Adjust size here
        self.setFont(font)

        #Create find button 
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget3 = QPushButton()
        QPushButton_Widget3.setText("Estimate Data Collection Time")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
        QPushButton_Widget5 = QPushButton()
        QPushButton_Widget5.setText("Stop Test")
        self.QCheckBox_Report_Widget = QCheckBox()
        self.QCheckBox_Report_Widget.setText("Generate Excel Report")
        self.QCheckBox_Report_Widget.setCheckState(Qt.Checked)
        self.QCheckBox_Image_Widget = QCheckBox()
        self.QCheckBox_Image_Widget.setText("Show Graph")
        self.QCheckBox_Image_Widget.setCheckState(Qt.Checked)
        QCheckBox_Lock_Widget = QCheckBox()
        QCheckBox_Lock_Widget.setText("Lock Widget")

        self.QCheckBox_Voltage_Accuracy_Voltage_Mode_Widget = QCheckBox()
        self.QCheckBox_Voltage_Accuracy_Voltage_Mode_Widget.setText("Current Static (Voltage Change)")
        self.QCheckBox_Voltage_Accuracy_Voltage_Mode_Widget.setCheckState(Qt.Checked)
        
        self.QCheckBox_Voltage_Accuracy_Current_Mode_Widget = QCheckBox()
        self.QCheckBox_Voltage_Accuracy_Current_Mode_Widget.setText("Current Change(Load Change)")
        self.QCheckBox_Voltage_Accuracy_Current_Mode_Widget.setCheckState(Qt.Unchecked)

        self.QCheckbox_Simulation_Mode = QCheckBox()
        self.QCheckbox_Simulation_Mode.setText("Simulation Mode")
        self.QCheckbox_Simulation_Mode.setCheckState(Qt.Unchecked)
        

        #Output Display
        self.OutputBox = QTextBrowser()
        self.OutputBox.append(f"{my_result.getvalue()}")
        self.OutputBox.append("")  # Empty line after each append

        #Description 1-7
        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()
        Desp6 = QLabel()
        Desp7 = QLabel()
        Desp8 = QLabel()

        Desp0.setFont(desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)
        Desp6.setFont(desp_font)
        Desp7.setFont(desp_font)
        Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Voltage Sweep:")
        Desp4.setText("Current Sweep:")
        Desp5.setText("No. of Collection:")
        Desp6.setText("Rated Power [W]")
        Desp7.setText("Maximum Current")
        Desp8.setText("External Auxiliary Equipment:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        #Voltage Accuracy Mode
        QLabel_Voltage_Accuracy_Mode = QLabel()
        QLabel_Voltage_Accuracy_Mode.setFont(desp_font)
        QLabel_Voltage_Accuracy_Mode.setText("Voltage Accuracy Test Mode Selection:")

        #Simulation Mode
        QLabel_Simulation_Mode =QLabel()    
        QLabel_Simulation_Mode.setFont(desp_font)
        QLabel_Simulation_Mode.setText("Enable Simulation Mode (No Instruments Connected):")

        # Connections section
        QLabel_PSU_VisaAddress = QLabel()
        QLabel_DMM_VisaAddress = QLabel()
        QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DMM_Instrument = QLabel()
        QLabel_DUT = QLabel()
        QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        QLabel_DMM_VisaAddress.setText("Visa Address (DMM):")
        QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        QLabel_DUT.setText("DUT Test Profile:")

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        self.QComboBox_DMM_Instrument = QComboBox()
        self.QComboBox_DUT = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Programming_Error_Gain.setText("Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Readback Desired Specification (Offset):")

        # External Auxiliary Equipment section
        QLabel_External_Auxiliary_Equipment = QLabel()
        QLabel_External_Auxiliary_Equipment.setText("Relay")
        self.QComboBox_External_Auxiliary_Equipment = QComboBox()
        self.QComboBox_External_Auxiliary_Equipment.addItems(["None", "RELAY"])

        self.QComboBox_Voltage_Res = QComboBox()
        self.QComboBox_set_PSU_Channel = QComboBox()
        self.QComboBox_set_ELoad_Channel = QComboBox()
        self.QComboBox_set_Function = QComboBox()
        self.QComboBox_Voltage_Sense = QComboBox()
        self.QLineEdit_Programming_Error_Gain = QLineEdit()
        self.QLineEdit_Programming_Error_Offset = QLineEdit()
        self.QLineEdit_Readback_Error_Gain = QLineEdit()
        self.QLineEdit_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_DMM_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QComboBox_DUT.addItems(["Others", "Excavator", "EDU36311A", "Dolphin", "Hornbill", "SMU"])

        self.QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        self.QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        self.QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        self.QComboBox_set_Function.setEnabled(False)
        self.QComboBox_set_PSU_Channel.addItems(["None", "1", "2", "3", "4"])
        self.QComboBox_set_PSU_Channel.setEnabled(True)
        self.QComboBox_set_ELoad_Channel.addItems(["None", "1", "2"])
        self.QComboBox_set_ELoad_Channel.setEnabled(True)
        self.QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])
        self.QComboBox_Voltage_Sense.setEnabled(True)

        #Power
        QLabel_Power = QLabel()
        QLabel_Power.setText("Rated Power (W):")
        self.QLineEdit_Power = QLineEdit()

        # Current Sweep
        QLabel_minCurrent = QLabel()
        QLabel_maxCurrent = QLabel()
        QLabel_current_step_size = QLabel()
        QLabel_minCurrent.setText("Initial Current (A):")
        QLabel_maxCurrent.setText("Final Current (A):")
        QLabel_current_step_size.setText("Step Size:")
        self.QLineEdit_minCurrent = QLineEdit()
        self.QLineEdit_maxCurrent = QLineEdit()
        self.QLineEdit_current_stepsize = QLineEdit()

        # Voltage Sweep
        QLabel_minVoltage = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_step_size = QLabel()
        QLabel_minVoltage.setText("Initial Voltage (V):")
        QLabel_maxVoltage.setText("Final Voltage (V):")
        QLabel_voltage_step_size.setText("Step Size:")

        self.QLineEdit_minVoltage = QLineEdit()
        self.QLineEdit_maxVoltage = QLineEdit()
        self.QLineEdit_voltage_stepsize = QLineEdit()

        #Loop & Delay
        QLabel_noofloop = QLabel()
        QLabel_noofloop.setText("No. of Data Collection:")
        self.QComboBox_noofloop = QComboBox()
        self.QComboBox_noofloop.addItems(["1","2","3","4","5","6","7","8","9","10"])

        QLabel_updatedelay = QLabel()
        QLabel_updatedelay.setText("Delay Time (second) :(Default=50ms)")
        self.QComboBox_updatedelay = QComboBox()
        self.QComboBox_updatedelay.addItems(["0.0","0.8","1.0","2.0","3.0", "4.0"])
        
        #Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QVBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
        #save_path_layout.addWidget(QLineEdit_Save_Path)  # QLineEdit for the path
        save_path_layout.addWidget(self.QCheckBox_Report_Widget)  # Checkbox for "Generate Excel Report"
        save_path_layout.addWidget(self.QCheckBox_Image_Widget)  # Checkbox for "Show Graph"
        save_path_layout.addWidget(QCheckBox_Lock_Widget)  # Checkbox for "Show Graph"

        #Execute Layout + Outputbox in Right Container
        Right_container = QVBoxLayout()
        exec_layout_box = QHBoxLayout()
        exec_layout = QFormLayout()

        #exec_layout.addRow(Desp0)
        exec_layout.addWidget(self.OutputBox)
        exec_layout.addRow(QPushButton_Widget0)

        exec_layout.addRow(QPushButton_Widget3)
        exec_layout.addRow(QPushButton_Widget2)
        exec_layout.addRow(QPushButton_Widget1)
        exec_layout.addRow(QPushButton_Widget5)   

        exec_layout_box.addLayout(exec_layout)
 
        Right_container.addLayout(save_path_layout)
        Right_container.addLayout(exec_layout_box)

        #Setting Form Layout with Left Container
        Left_container = QHBoxLayout()
        setting_layout = QFormLayout()

        setting_layout.addRow(Desp1)
        setting_layout.addRow(self.QCheckbox_Simulation_Mode)
        setting_layout.addRow(QLabel_DUT, self.QComboBox_DUT)
        setting_layout.addRow(QLabel_Voltage_Accuracy_Mode)
        setting_layout.addRow(self.QCheckBox_Voltage_Accuracy_Voltage_Mode_Widget)
        setting_layout.addRow(self.QCheckBox_Voltage_Accuracy_Current_Mode_Widget)
        setting_layout.addRow(QPushButton_Widget4)
        setting_layout.addRow(QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        setting_layout.addRow(QLabel_DMM_VisaAddress, self.QLineEdit_DMM_VisaAddress)
        setting_layout.addRow(QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        setting_layout.addRow(QLabel_DMM_Instrument, self.QComboBox_DMM_Instrument)
        setting_layout.addRow(Desp2)
        setting_layout.addRow(QLabel_set_PSU_Channel, self.QComboBox_set_PSU_Channel)
        setting_layout.addRow(QLabel_set_ELoad_Channel, self.QComboBox_set_ELoad_Channel)
        setting_layout.addRow(QLabel_set_Function, self.QComboBox_set_Function)
        setting_layout.addRow(QLabel_Voltage_Sense, self.QComboBox_Voltage_Sense)
        setting_layout.addRow(QLabel_Programming_Error_Gain, self.QLineEdit_Programming_Error_Gain)
        setting_layout.addRow(QLabel_Programming_Error_Offset, self.QLineEdit_Programming_Error_Offset)
        setting_layout.addRow(QLabel_Readback_Error_Gain, self.QLineEdit_Readback_Error_Gain)
        setting_layout.addRow(QLabel_Readback_Error_Offset, self.QLineEdit_Readback_Error_Offset)
        setting_layout.addRow(Desp8)
        setting_layout.addRow(QLabel_External_Auxiliary_Equipment, self.QComboBox_External_Auxiliary_Equipment)
        setting_layout.addRow(Desp6)
        setting_layout.addRow(QLabel_Power, self.QLineEdit_Power)
        setting_layout.addRow(Desp3)
        setting_layout.addRow(QLabel_minVoltage, self.QLineEdit_minVoltage)
        setting_layout.addRow(QLabel_maxVoltage, self.QLineEdit_maxVoltage)
        setting_layout.addRow(QLabel_voltage_step_size, self.QLineEdit_voltage_stepsize)
        setting_layout.addRow(Desp4)
        setting_layout.addRow(QLabel_minCurrent, self.QLineEdit_minCurrent)
        setting_layout.addRow(QLabel_maxCurrent, self.QLineEdit_maxCurrent)
        setting_layout.addRow(QLabel_current_step_size, self.QLineEdit_current_stepsize)
        setting_layout.addRow(Desp5)
        setting_layout.addRow(QLabel_noofloop, self.QComboBox_noofloop)
        setting_layout.addRow(QLabel_updatedelay, self.QComboBox_updatedelay)

        Left_container.addLayout(setting_layout)

        #Main Layout
        Main_Layout = QHBoxLayout()
        Main_Layout.addLayout(Left_container,stretch= 2)
        Main_Layout.addLayout(Right_container,stretch = 1)
        self.setLayout(Main_Layout)
        scroll_area(self,Main_Layout)

        self.QCheckbox_Simulation_Mode.stateChanged.connect(self.simulation_mode_changed)   

        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddress.currentTextChanged.connect(self.DMM_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)

        self.QCheckBox_Voltage_Accuracy_Voltage_Mode_Widget.stateChanged.connect(self.checkbox_state_Voltage_Accuracy_Voltage_Mode)
        self.QCheckBox_Voltage_Accuracy_Current_Mode_Widget.stateChanged.connect(self.checkbox_state_Voltage_Accuracy_Current_Mode)
        
        self.QComboBox_DUT.currentTextChanged.connect(self.DUT_changed)
        self.QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        self.QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        self.QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        self.QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)

        self.QComboBox_External_Auxiliary_Equipment.currentTextChanged.connect(self.External_Auxiliary_Equipment_changed)

        self.QLineEdit_Power.textEdited.connect(self.Power_changed)
        self.QComboBox_DUT.currentIndexChanged.connect(self.on_current_index_changed)

        self.QLineEdit_minVoltage.textEdited.connect(self.minVoltage_changed)
        self.QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)
        self.QLineEdit_minCurrent.textEdited.connect(self.minCurrent_changed)
        self.QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        self.QLineEdit_voltage_stepsize.textEdited.connect(self.voltage_step_size_changed)
        self.QLineEdit_current_stepsize.textEdited.connect(self.current_step_size_changed)
        self.QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        self.QComboBox_set_PSU_Channel.currentTextChanged.connect(self.PSU_Channel_changed)
        self.QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        self.QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        self.QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)

        self.QComboBox_noofloop.currentTextChanged.connect(self.noofloop_changed)
        self.QComboBox_updatedelay.currentTextChanged.connect(self.updatedelay_changed)

        self.QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)
        self.QCheckBox_Report_Widget.stateChanged.connect(self.checkbox_state_Report)
        self.QCheckBox_Image_Widget.stateChanged.connect(self.checkbox_state_Image)
        QCheckBox_Lock_Widget.stateChanged.connect(self.checkbox_state_lock)

        QPushButton_Widget0.clicked.connect(self.savepath)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget3.clicked.connect(self.estimateTime)
        QPushButton_Widget4.clicked.connect(self.doFind)

        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)
    
    def simulation_mode_changed(self, state):
        if state == Qt.Checked:
            self.simulation_mode = 2
        else:
            self.simulation_mode = 0
    
    def checkbox_state_Voltage_Accuracy_Voltage_Mode(self, state):
        if state == Qt.Checked:
            self.checkbox_Voltage_Accuracy_Voltage_Mode = 2
        else:
            self.checkbox_Voltage_Accuracy_Voltage_Mode = 0

    def checkbox_state_Voltage_Accuracy_Current_Mode(self, state):
        if state == Qt.Checked:
            self.checkbox_Voltage_Accuracy_Current_Mode = 2
        else:
            self.checkbox_Voltage_Accuracy_Current_Mode = 0

    def External_Auxiliary_Equipment_changed(self, s):
        if s == "None":
            self.relay_voltage = "None"
        else:
            self.relay_voltage = "RELAY"

    def on_current_index_changed(self):
        selected_text = self.QComboBox_DUT.currentText()
        self.update_selection(selected_text)
 
        self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
        self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
        self.QLineEdit_Programming_Error_Offset.setText(self.Programming_Error_Offset)
        self.QLineEdit_Readback_Error_Gain.setText(self.Readback_Error_Gain)
        self.QLineEdit_Readback_Error_Offset.setText(self.Readback_Error_Offset)
        self.QLineEdit_Power.setText(self.Power)
        """self.QLineEdit_rshunt.setText(self.rshunt)"""
        self.QLineEdit_minVoltage.setText(self.minVoltage)
        self.QLineEdit_maxVoltage.setText(self.maxVoltage)
        self.QLineEdit_voltage_stepsize.setText(self.voltage_step_size)
        self.QLineEdit_minCurrent.setText(self.minCurrent)
        self.QLineEdit_maxCurrent.setText(self.maxCurrent)
        self.QLineEdit_current_stepsize.setText(self.current_step_size)

        channel_str = str(self.PSU_Channel)
        index = self.QComboBox_set_PSU_Channel.findText(channel_str)
        self.QComboBox_set_PSU_Channel.setCurrentIndex(index)
        self.QComboBox_set_PSU_Channel.setCurrentIndex(int(self.PSU_Channel))
        self.QComboBox_set_ELoad_Channel.setCurrentIndex(int(self.ELoad_Channel))
        self.QComboBox_Voltage_Sense.setCurrentText("4 Wire" if self.VoltageSense == "EXT" else "2 Wire")
        self.QComboBox_noofloop.setCurrentText(self.noofloop)
        self.QComboBox_updatedelay.setCurrentText(self.updatedelay)

    def update_selection(self, selected_text):
        """Update selected text and reload config file"""
        self.selected_text = selected_text
        self.load_data()

    """def load_data(self):
        Reads configuration file and returns a dictionary of key-value pairs.
        config_data = {}
        if self.selected_text =="Excavator":
            self.config_file = os.path.join(config_folder,"config_Excavator.txt")
            
        elif self.selected_text =="Dolphin":
            self.config_file = os.path.join(config_folder,"config_Dolphin.txt")
            
        elif self.selected_text =="SMU":
            self.config_file = os.path.join(config_folder,"config_SMU.txt")
        
        elif self.selected_text =="Hornbill":
            self.config_file = os.path.join(config_folder,"config_Hornbill.txt")
        else:
            self.config_file = os.path.join(config_folder,"config_Others.txt")

        try:
            with open(self.config_file, "r") as file: 
                for line in file:

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue 

                    if "=" in line:
                        key, value = line.strip().split("=")
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    # If savelocation has a valid value, do not overwrite it
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue 
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe("Config file not found. Using default values.")

        return config_data"""
    

    """def load_data(self):
        Reads configuration file and returns a dictionary of key-value pairs.
        config_data = {}

        # Determine base folder for configs
        if getattr(sys, "frozen", False):
            # Running as PyInstaller exe
            exe_folder = Path(sys.executable).parent
            # Configs should be in Executable/Instrument_Config_Files
            config_base = exe_folder / "Instrument_Config_Files"
        else:
            # Running as normal Python script
            project_root = Path(__file__).resolve().parent.parent  # <-- go up one level from src
            config_base = project_root / "Instrument_Config_Files"

        # Map selected_text to config filename
        file_map = {
            "Excavator": "config_Excavator.txt",
            "Dolphin": "config_Dolphin.txt",
            "SMU": "config_SMU.txt",
            "Hornbill": "config_Hornbill.txt"
        }
        config_filename = file_map.get(self.selected_text, "config_Others.txt")
        self.config_file = config_base / config_filename

        try:
            with open(self.config_file, "r") as file:
                for line in file:
                    line = line.strip()
                    if not line or line.startswith("#") or line.startswith("//"):
                        continue
                    if "=" in line:
                        key, value = line.split("=", 1)
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe(f"⚠️ Config file not found: {self.config_file}. Using default values.")

        return config_data"""
    
    def load_data(self):
        """Reads configuration file and returns a dictionary of key-value pairs."""
        config_data = {}

        config_map = {
            "Excavator": "config_Excavator.txt",
            "Dolphin": "config_Dolphin.txt",
            "SMU": "config_SMU.txt",
            "Hornbill": "config_Hornbill.txt",
        }

        filename = config_map.get(self.selected_text, "config_Others.txt")

        # Ensure path is a string, avoids crashes
        self.config_file = str(config_folder / filename)

        if not os.path.exists(self.config_file):
            print_console_safe(f"⚠ Config file not found: {self.config_file}")
            return {}

        try:
            with open(self.config_file, "r", encoding="utf-8-sig") as file:
                for line in file:
                    line = line.strip()

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue

                    if "=" not in line:
                        print_console_safe("⚠ Skipping invalid line:", repr(line))
                        continue

                    key, value = line.split("=", 1)
                    config_data[key.strip()] = value.strip()


            for key, value in config_data.items():
                if key == "savelocation":
                    if getattr(self, "savelocation", None) and \
                    self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue
                    setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except Exception as e:
            print_console_safe(f"ERROR reading config: {e}")

        return config_data



    def Power_changed(self, value):
        self.Power = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = NewGetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            instrument_roles = discovery.roles
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])

            if 'PSU' in instrument_roles:
                psu_index = self.visaIdList.index(instrument_roles['PSU'])
                self.QLineEdit_PSU_VisaAddress.setCurrentIndex(psu_index)

            if 'ELOAD' in instrument_roles:
                eload_index = self.visaIdList.index(instrument_roles['ELOAD'])
                self.QLineEdit_ELoad_VisaAddress.setCurrentIndex(eload_index)

            if 'DMM' in instrument_roles:
                dmm_index = self.visaIdList.index(instrument_roles['DMM'])
                self.QLineEdit_DMM_VisaAddress.setCurrentIndex(dmm_index)
            
            # Add "None" option at the end
            self.QLineEdit_PSU_VisaAddress.addItem("TCPIP0::141.183.188.184::5025::SOCKET")
            self.QLineEdit_DMM_VisaAddress.addItem("None")
            self.QLineEdit_ELoad_VisaAddress.addItem("None")
            self.QLineEdit_PSU_VisaAddress.addItem("TCPIP0::141.183.188.184::inst0::INSTR")
                    
        except:
            self.OutputBox.append("No Devices Found!!!")
        return   


    def updatedelay_changed(self, value):
        self.updatedelay = value
        #self.OutputBox.append(str(self.updatedelay))

    def noofloop_changed(self, value):
        self.noofloop = value
        #self.OutputBox.append(str(self.noofloop))

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s    

    def DMM_VisaAddress_changed(self, s):
        self.DMM = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s

    def DUT_changed(self, s):
        self.DUT = s

    def ELoad_Channel_changed(self, s):
       self.ELoad_Channel = s

    def PSU_Channel_changed(self, s):
       self.PSU_Channel = s

    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s

    def minVoltage_changed(self, s):
        self.minVoltage = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def minCurrent_changed(self, s):
        self.minCurrent = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def voltage_step_size_changed(self, s):
        self.voltage_step_size = s

    def current_step_size_changed(self, s):
        self.current_step_size = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.VoltageRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setInputZ(self, value):
        AdvancedSettingsList[3] = value

    def checkbox_state_Report(self, s):
        self.checkbox_data_Report = s

    def checkbox_state_Image(self, s):
        self.checkbox_data_Image = s

    def checkbox_state_lock(self, state):
        lockable_widgets = (QPushButton, QLineEdit, QTextEdit, QComboBox)

        for widget in self.findChildren(lockable_widgets):
            widget.setDisabled(state == 2)  # Disable if checkbox is checked

    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value

    def openDialog(self):
        dlg = AdvancedSetting_Voltage()
        dlg.exec()

    def estimateTime(self):
        
        #self.OutputBox.append(str(self.updatedelay))
        try:
            self.currloop = ((float(self.maxCurrent) - float(self.minCurrent))/ float(self.current_step_size)) + 1
            self.voltloop = ((float(self.maxVoltage) - float(self.minVoltage))/ float(self.voltage_step_size)) + 1

            if self.updatedelay == 0.0:
                constant = 0
                    
            else:
                constant = 1

            self.estimatetime = (self.currloop * self.voltloop *(0.2 + 0.8 + (float(self.updatedelay) * constant) )) * float(self.noofloop)
            self.OutputBox.append(f"{self.estimatetime} seconds")
            self.OutputBox.append("") 
        except Exception as e:
            QMessageBox.warning(self, "Warning", "Input cannot be empty!")
            return

    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

    def check_missing_params(self, param_dict):
        """Check which parameters are None or empty and return them."""
        missing = []
        for key, value in param_dict.items():
            # Check for None or empty string
            if value is None or (isinstance(value, str) and value.strip() == ""):
                missing.append(key)
        return missing

    def executeTest(self):
        global globalvv
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(
                self,
                "Test Already Running",
                "Wait for the active test to finish or stop it first.",
            )
            return

        try:
        
            """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
            then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
            will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
            begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
            begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
            are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
            optionally export all the details into a CSV file or display a graph after the test is completed.

            """
            
            self.infoList = []
            self.dataList = []
            self.dataList2 = []
            self.ProgrammingV_percent_error_list = []
            self.ReadbackV_percent_error_list = []
            self.setEnabled(False)

            dict = dictGenerator.input(
                V_Rating=self.Voltage_Rating,
                power=self.Power,
                estimatetime=self.estimatetime,
                updatedelay=self.updatedelay,
                readbackvoltage=self.readbackvoltage,
                readbackcurrent=self.readbackcurrent,
                noofloop=self.noofloop,
                Instrument=self.DMM_Instrument,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,
                relay_voltage=self.relay_voltage,
                unit=self.unit,
                minCurrent=self.minCurrent,
                maxCurrent=self.maxCurrent,
                current_step_size=self.current_step_size,
                minVoltage=self.minVoltage,
                maxVoltage=self.maxVoltage,
                voltage_step_size=self.voltage_step_size,
                selected_DUT=self.selected_text,
                PSU=self.PSU,
                OperationMode=self.SPOperationMode,
                OSC=self.OSC,
                DMM=self.DMM,
                ELoad=self.ELoad,
                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],
            )

            # Check for missing parameters
            missing = self.check_missing_params(dict)
            if missing:
                print_console_safe(f"The following parameters are missing or empty: {missing}")
                return

            if self.simulation_mode == 2:
                QMessageBox.information(
                    self,
                    "Simulation Mode",
                    "Simulation Mode is enabled. No instruments are connected.",
                )
                
                wb = load_workbook(self.output_file)
                ws = wb["Data"]  # get the worksheet

                graph_plot = Graph_Plotting(
                    wb,        # workbook
                    ws,        # sheet
                    self.output_file,
                    "Voltage",
                    1,
                    [8],       # graph_start_rows
                )
                graph_plot.execute_plot()

                wb.save(self.output_file)  # save the charts back
                sleep(10)
                
            else:
                QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
                                    )
                
                """if self.relay_voltage == "RELAY":
                    relay_voltage = RelayController_Voltage()
                    relay_voltage.relay_on()"""
                for x in range (int(self.noofloop)):
                    if self.DMM_Instrument == "Keysight":

                        if self.selected_text == "Dolphin":
                            if self.ELoad != "None" and self.DMM != "None":
                                print_console_safe("ELoad connected and DMM connected") #All connected
                                try:(
                                    self.infoList,
                                    self.dataList,
                                    self.dataList2
                                    ) = DolphinNewVoltageMeasurementwithELoad.Execute_Voltage_Accuracy(self, dict, self.PSU_Channel)
                            
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return

                            elif self.ELoad == "None" and self.DMM != "None":
                                print_console_safe("No ELoad connected and DMM connected") #No Eload connected but DMM connected
                                try:(
                                    self.infoList,
                                    self.dataList,
                                    self.dataList2
                                    ) = DolphinNewVoltageMeasurementNoELoadWithDMM.Execute_Voltage_Accuracy(self, dict, self.PSU_Channel)
                                    
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return
                            
                            elif self.ELoad != "None" and self.DMM == "None":
                                print_console_safe("ELoad connected and No DMM connected") #Eload connected but no DMM connected
                                try:(
                                    self.infoList,
                                    self.dataList,
                                    self.dataList2
                                    ) = DolphinNewVoltageMeasurementwithELoadNoDMM.Execute_Voltage_Accuracy(self, dict, self.PSU_Channel)
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return
                            else :
                                print_console_safe("No ELoad connected and No DMM connected") #No Eload connected and no DMM connected
                                try:(
                                    self.infoList,
                                    self.dataList,
                                    self.dataList2
                                    ) = DolphinNewVoltageMeasurementNoELoadNoDMM.Execute_Voltage_Accuracy(self, dict, self.PSU_Channel)
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return
                                                    
                        if self.selected_text == "EDU36311A":
                            try:(
                                infoList,
                                dataList,
                                dataList2  
                            ) = EDU36311AVoltageMeasurementNoELoad.Execute_Voltage_Accuracy(self, dict)
                            
                            except Exception as e:
                                QMessageBox.warning(self, "Error", str(e))
                                return
                            
                                """   elif self.selected_text == "Excavator":
                                        try:(
                                            infoList,
                                            dataList,
                                            dataList2
                                        ) = ExcavatorVoltageMeasurement.Execute_Voltage_Accuracy(self, dict)

                                        except Exception as e:
                                            QMessageBox.warning(self, "Error", str(e))
                                            return
                                    elif self.selected_text == "SMU":
                                        try:(
                                            infoList, 
                                            dataList,
                                            dataList2
                                        ) = SMUVoltageMeasurement.Execute_Voltage_Accuracy(self, dict)
                                        except Exception as e:
                                            QMessageBox.warning(self, "Error", str(e))  
                                            return"""
                        
                        elif self.selected_text == "Hornbill":                                          #referring
                            if self.ELoad != "None" and self.DMM != "None":
                                if self.checkbox_Voltage_Accuracy_Voltage_Mode == 2:
                                    print_console_safe("ELoad connected and DMM connected") #All connected
                                    try:(
                                        self.infoList,   
                                        self.dataList,
                                        self.dataList2
                                    ) = HornbillVoltageMeasurementwithELoad.Execute_Voltage_Accuracy_Current_Static(self, dict, self.PSU_Channel)
                                    except Exception as e:
                                        QMessageBox.warning(self, "Error", str(e))
                                        return
                                elif self.checkbox_Voltage_Accuracy_Current_Mode == 2:
                                    print_console_safe("ELoad connected and DMM connected") #All connected
                                    try:(
                                        self.infoList,
                                        self.dataList,
                                        self.dataList2
                                    ) = HornbillVoltageMeasurementwithELoad.Execute_Voltage_Accuracy_Current_Change(self, dict, self.PSU_Channel)
                                    except Exception as e:
                                        QMessageBox.warning(self, "Error", str(e))
                                        return
                            else:
                                print_console_safe("No ELoad connected and DMM connected") #No Eload connected but DMM connected
                                
                    """if self.relay_voltage == "RELAY":
                        relay_voltage.relay_off()
                    """
                
                    #Export Data to CSV
                    if self.QCheckBox_Report_Widget.isChecked():
                        if self.QCheckBox_Voltage_Accuracy_Voltage_Mode_Widget.isChecked():
                            instrumentData(self.PSU, self.DMM, self.ELoad)
                            datatoCSV_Accuracy(self.infoList, self.dataList, self.dataList2)
                            datatoGraph(self.infoList, self.dataList, self.dataList2)
                            datatoGraph.scatterCompareVoltage(self, float(self.Programming_Error_Gain), float(self.Programming_Error_Offset), float(self.Readback_Error_Gain), float(self.Readback_Error_Offset), str(self.unit), float(self.Voltage_Rating))

                            
                            df = pd.DataFrame.from_dict(dict, orient="index")
                            df.to_csv(os.path.join(csv_folder,"config.csv"))

                            A = xlreport(save_directory=self.savelocation, file_name=str(self.unit))
                            self.output_file = A.run()

                        elif self.QCheckBox_Voltage_Accuracy_Current_Mode_Widget.isChecked():
                            instrumentData(self.PSU, self.DMM, self.ELoad)
                            datatoCSV_Accuracy(self.infoList, self.dataList, self.dataList2)
                            datatoGraph(self.infoList, self.dataList, self.dataList2)
                            datatoGraph.scatterCompareVoltage_Current_Change(self, float(self.Programming_Error_Gain), float(self.Programming_Error_Offset), float(self.Readback_Error_Gain), float(self.Readback_Error_Offset), str(self.unit), float(self.Voltage_Rating))

                            df = pd.DataFrame.from_dict(dict, orient="index")
                            df.to_csv(os.path.join(csv_folder,"config.csv"))

                            A = xlreport(save_directory=self.savelocation, file_name=str(self.unit))
                            A.run()
                    
                    #Measurement Completion
                    if x == (int(self.noofloop) - 1):   
                        self.OutputBox.append("Measurement is complete !")
                        #Show Graph Image
                        if self.QCheckBox_Image_Widget.isChecked():
                            self.image_dialog = image_Window()
                            self.image_dialog.setModal(True)
                            self.image_dialog.show()


        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

            return
        finally:
            self.setEnabled(True)

    def stop_execution(self):
        self.stop_requested = True

class CurrentMeasurementDialog(QDialog):
    """Class for configuring the current measurement DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.
    """

    def __init__(self):
        """Method where Widgets, Signals and Slots are defined in the GUI for Current Measurement"""
        super().__init__()
         # Initialize default CSV path
        self.DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/data.csv"  # or use a dynamic path if needed
        self.IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/Chart.png"
        self.ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/error.csv"
        self.image_dialog = None
        self.POWER_DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerdata.csv"  # or use a dynamic path if needed
        self.POWER_IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/powerChart.png"
        self.POWER_ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powererror.csv"

        # Default Values
        #self.PSU_VisaAddress = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        #self.DMM_VisaAddress = "USB0::0x2A8D::0x1601::MY60094787::0::INSTR"
        #self.ELoad_VisaAddress = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        self.rshunt = "0.05"
        self.Power_Rating = "6000"
        self.Current_Rating = "24"
        self.Voltage_Rating = "800"
        self.Power = "2000"
        self.powerfin = self.Power
        self.powerini = "0"
        self.power_step_size = "60"
        self.currloop = "1"
        self.voltloop = "1"
        self.estimatetime = "0"
        self.readbackvoltage = "0"
        self.readbackcurrent = "0"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"
        self.noofloop = "1"
        self.updatedelay = "0.0"
        self.unit = "CURRENT"
        self.Programming_Error_Offset = "0.001"
        self.Programming_Error_Gain = "0.001"
        self.Readback_Error_Offset = "0.001"
        self.Readback_Error_Gain = "0.001"

        self.Power_Programming_Error_Offset = "0.005"
        self.Power_Programming_Error_Gain = "0.005"
        self.Power_Readback_Error_Offset = "0.005"
        self.Power_Readback_Error_Gain = "0.005"

        self.relay_current = "None"

        self.minCurrent = "1"
        self.maxCurrent = "1"
        self.current_step_size = "1"
        self.minVoltage = "1"
        self.maxVoltage = "10"
        self.voltage_step_size = "1"

        self.PSU = "USB0::0x2A8D::0xDA04::CN24350083::0::INSTR"
        self.DMM = "USB0::0x2A8D::0x0201::MY57702128::0::INSTR"
        self.DMM2 = "USB0::0x2A8D::0x0201::MY54701197::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"
        self.ELoad_Channel = "1"
        self.PSU_Channel = "1"
        self.DMM_Instrument = "Keysight"

        self.setFunction = "Voltage" #Set Eload in CV Mode
        self.VoltageRes = "DEF"
        self.VoltageSense = "INT"
        self.SPOperationMode = "Independent"

        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.checkbox_test_CurrentAccuracy = 2
        self.checkbox_test_CurrentLoadRegulation = 2
        self.checkbox_test_PowerAccuracy = 2

        self.Range = "Auto"
        self.Aperture = "MAX"  #Im thinking it need 100 NPLC
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.UpTime = "50"
        self.DownTime = "50"

        self.simulation_mode = 0
        self.output_file = r"C:\Users\shamlee3\OneDrive - Keysight Technologies\Wei Jing See's files - Shamman Xian Jun Lee\Test Data\Monk2.xlsx"  
        self.selected_text  = "Hornbill"
        self.stop_requested = False

        self.checkbox_Current_Accuracy_FULL = 2
        self.checkbox_Current_Accuracy_200uA = 0
        self.checkbox_Current_Accuracy_2mA = 0
        
        
        # Ensure DATA_CSV_PATH is defined before use
        if not hasattr(self, 'DATA_CSV_PATH') or not self.DATA_CSV_PATH:
            QMessageBox.warning(self, "Error", "CSV path is not set.")

        if not hasattr(self, 'POWER_DATA_CSV_PATH') or not self.POWER_DATA_CSV_PATH:
            QMessageBox.warning(self, "Error", "POWER CSV path is not set.")

        #Create Voltage Window
        self.setWindowTitle("Current Measurement")
        self.image_window = None
        self.setWindowFlags(Qt.Window)
        font = QFont()
        font.setPointSize(16)   # Adjust size here
        self.setFont(font)

        #Create find button 
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget3 = QPushButton()
        QPushButton_Widget3.setText("Estimate Data Collection Time")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
        self.QCheckBox_Report_Widget = QCheckBox()
        self.QCheckBox_Report_Widget.setText("Generate Excel Report")
        self.QCheckBox_Report_Widget.setCheckState(Qt.Checked)
        self.QCheckBox_Image_Widget = QCheckBox()
        self.QCheckBox_Image_Widget.setText("Show Graph")
        self.QCheckBox_Image_Widget.setCheckState(Qt.Checked)
        QCheckBox_Lock_Widget = QCheckBox()
        QCheckBox_Lock_Widget.setText("Lock Widget")

        #Shamman changes
        self.QCheckBox_Current_Accuracy_FULL_Widget = QCheckBox()
        self.QCheckBox_Current_Accuracy_FULL_Widget.setText("Current at full")
        self.QCheckBox_Current_Accuracy_FULL_Widget.setCheckState(Qt.Checked)
        
        self.QCheckBox_Current_Accuracy_200uA_Widget = QCheckBox()
        self.QCheckBox_Current_Accuracy_200uA_Widget.setText("Current at 200uA")
        self.QCheckBox_Current_Accuracy_200uA_Widget.setCheckState(Qt.Unchecked)

        self.QCheckBox_Current_Accuracy_2mA_Widget = QCheckBox()
        self.QCheckBox_Current_Accuracy_2mA_Widget.setText("Current at 200uA")
        self.QCheckBox_Current_Accuracy_2mA_Widget.setCheckState(Qt.Unchecked)

        self.QCheckbox_Simulation_Mode = QCheckBox()
        self.QCheckbox_Simulation_Mode.setText("Simulation Mode")
        self.QCheckbox_Simulation_Mode.setCheckState(Qt.Unchecked)

        #Output Display
        self.OutputBox = QTextBrowser()
        self.OutputBox.append(f"{my_result.getvalue()}")
        self.OutputBox.append("")  # Empty line after each append

        #Description 1-7
        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()
        Desp6 = QLabel()
        Desp7 = QLabel()
        Desp8 = QLabel()

        Desp0.setFont(desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)
        Desp6.setFont(desp_font)
        Desp7.setFont(desp_font)
        Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Voltage Sweep:")
        Desp4.setText("Current Sweep:")
        Desp5.setText("No. of Collection:")
        Desp6.setText("Rated Power [W]")
        Desp7.setText("Maximum Current")
        Desp8.setText("External Auxiliary Equipment:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        #Voltage Accuracy Mode
        QLabel_Voltage_Accuracy_Mode = QLabel()
        QLabel_Voltage_Accuracy_Mode.setFont(desp_font)
        QLabel_Voltage_Accuracy_Mode.setText("Voltage Accuracy Test Mode Selection:")

        #Simulation Mode
        QLabel_Simulation_Mode =QLabel()    
        QLabel_Simulation_Mode.setFont(desp_font)
        QLabel_Simulation_Mode.setText("Enable Simulation Mode (No Instruments Connected):")

        # Connections section
        QLabel_PSU_VisaAddress = QLabel()
        QLabel_DMM_VisaAddress = QLabel()
        QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DMM_Instrument = QLabel()
        QLabel_DUT = QLabel()
        QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        QLabel_DMM_VisaAddress.setText("Visa Address (DMM):")
        QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        QLabel_DUT.setText("DUT:")

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        self.QComboBox_DMM_Instrument = QComboBox()
        self.QComboBox_DUT = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Programming_Error_Gain.setText("Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Readback Desired Specification (Offset):")

        # External Auxiliary Equipment section
        QLabel_External_Auxiliary_Equipment = QLabel()
        QLabel_External_Auxiliary_Equipment.setText("Relay")
        self.QComboBox_External_Auxiliary_Equipment = QComboBox()
        self.QComboBox_External_Auxiliary_Equipment.addItems(["None", "RELAY"])


        self.QComboBox_Voltage_Res = QComboBox()
        self.QComboBox_set_PSU_Channel = QComboBox()
        self.QComboBox_set_ELoad_Channel = QComboBox()
        self.QComboBox_set_Function = QComboBox()
        self.QComboBox_Voltage_Sense = QComboBox()
        self.QLineEdit_Programming_Error_Gain = QLineEdit()
        self.QLineEdit_Programming_Error_Offset = QLineEdit()
        self.QLineEdit_Readback_Error_Gain = QLineEdit()
        self.QLineEdit_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_DMM_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QComboBox_DUT.addItems(["Others", "Excavator", "EDU36311A", "Dolphin", "Hornbill", "SMU"])

        self.QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        self.QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        self.QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        self.QComboBox_set_Function.setEnabled(False)
        self.QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        self.QComboBox_set_PSU_Channel.setEnabled(True)
        self.QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        self.QComboBox_set_ELoad_Channel.setEnabled(True)
        self.QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])
        self.QComboBox_Voltage_Sense.setEnabled(True)

        #Power
        QLabel_Power = QLabel()
        QLabel_Power.setText("Rated Power (W):")
        self.QLineEdit_Power = QLineEdit()

        QLabel_rshunt = QLabel()
        QLabel_rshunt.setText("Shunt Resistance Value (ohm):")
        self.QLineEdit_rshunt = QLineEdit()

        # Current Sweep
        QLabel_minCurrent = QLabel()
        QLabel_maxCurrent = QLabel()
        QLabel_current_step_size = QLabel()
        QLabel_minCurrent.setText("Initial Current (A):")
        QLabel_maxCurrent.setText("Final Current (A):")
        QLabel_current_step_size.setText("Step Size:")
        self.QLineEdit_minCurrent = QLineEdit()
        self.QLineEdit_maxCurrent = QLineEdit()
        self.QLineEdit_current_stepsize = QLineEdit()

        # Voltage Sweep
        QLabel_minVoltage = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_step_size = QLabel()
        QLabel_minVoltage.setText("Initial Voltage (V):")
        QLabel_maxVoltage.setText("Final Voltage (V):")
        QLabel_voltage_step_size.setText("Step Size:")

        self.QLineEdit_minVoltage = QLineEdit()
        self.QLineEdit_maxVoltage = QLineEdit()
        self.QLineEdit_voltage_stepsize = QLineEdit()

        #Loop & Delay
        QLabel_noofloop = QLabel()
        QLabel_noofloop.setText("No. of Data Collection:")
        self.QComboBox_noofloop = QComboBox()
        self.QComboBox_noofloop.addItems(["1","2","3","4","5","6","7","8","9","10"])

        QLabel_updatedelay = QLabel()
        QLabel_updatedelay.setText("Delay Time (second) :(Default=50ms)")
        self.QComboBox_updatedelay = QComboBox()
        self.QComboBox_updatedelay.addItems(["0.0","0.8","1.0","2.0","3.0", "4.0"])
        
        #Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QVBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
        #save_path_layout.addWidget(QLineEdit_Save_Path)  # QLineEdit for the path
        save_path_layout.addWidget(self.QCheckBox_Report_Widget)  # Checkbox for "Generate Excel Report"
        save_path_layout.addWidget(self.QCheckBox_Image_Widget)  # Checkbox for "Show Graph"
        save_path_layout.addWidget(QCheckBox_Lock_Widget)  # Checkbox for "Show Graph"

        #Execute Layout + Outputbox in Right Container
        Right_container = QVBoxLayout()
        exec_layout_box = QHBoxLayout()
        exec_layout = QFormLayout()

        #exec_layout.addRow(Desp0)
        exec_layout.addWidget(self.OutputBox)
        exec_layout.addRow(QPushButton_Widget0)

        exec_layout.addRow(QPushButton_Widget3)
        exec_layout.addRow(QPushButton_Widget2)
        exec_layout.addRow(QPushButton_Widget1)   


        exec_layout_box.addLayout(exec_layout)
 
        Right_container.addLayout(save_path_layout)
        Right_container.addLayout(exec_layout_box)

        #Setting Form Layout with Left Container
        Left_container = QHBoxLayout()
        setting_layout = QFormLayout()

        setting_layout.addRow(Desp1)
        setting_layout.addRow(self.QCheckbox_Simulation_Mode)
        setting_layout.addRow(QLabel_DUT, self.QComboBox_DUT)
        setting_layout.addRow(QLabel_Voltage_Accuracy_Mode)
        setting_layout.addRow(self.QCheckBox_Current_Accuracy_FULL_Widget)
        setting_layout.addRow(self.QCheckBox_Current_Accuracy_200uA_Widget)
        setting_layout.addRow(self.QCheckBox_Current_Accuracy_2mA_Widget)
        setting_layout.addRow(QPushButton_Widget4)
        setting_layout.addRow(QLabel_DUT, self.QComboBox_DUT)
        setting_layout.addRow(QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        setting_layout.addRow(QLabel_DMM_VisaAddress, self.QLineEdit_DMM_VisaAddress)
        setting_layout.addRow(QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        setting_layout.addRow(QLabel_DMM_Instrument, self.QComboBox_DMM_Instrument)
        setting_layout.addRow(Desp2)
        setting_layout.addRow(QLabel_set_PSU_Channel, self.QComboBox_set_PSU_Channel)
        setting_layout.addRow(QLabel_set_ELoad_Channel, self.QComboBox_set_ELoad_Channel)
        setting_layout.addRow(QLabel_set_Function, self.QComboBox_set_Function)
        setting_layout.addRow(QLabel_Voltage_Sense, self.QComboBox_Voltage_Sense)
        setting_layout.addRow(QLabel_Programming_Error_Gain, self.QLineEdit_Programming_Error_Gain)
        setting_layout.addRow(QLabel_Programming_Error_Offset, self.QLineEdit_Programming_Error_Offset)
        setting_layout.addRow(QLabel_Readback_Error_Gain, self.QLineEdit_Readback_Error_Gain)
        setting_layout.addRow(QLabel_Readback_Error_Offset, self.QLineEdit_Readback_Error_Offset)
        setting_layout.addRow(Desp8)
        setting_layout.addRow(QLabel_External_Auxiliary_Equipment, self.QComboBox_External_Auxiliary_Equipment)
        setting_layout.addRow(Desp6)
        setting_layout.addRow(QLabel_Power, self.QLineEdit_Power)
        setting_layout.addRow(QLabel_rshunt, self.QLineEdit_rshunt)
        setting_layout.addRow(Desp3)
        setting_layout.addRow(QLabel_minVoltage, self.QLineEdit_minVoltage)
        setting_layout.addRow(QLabel_maxVoltage, self.QLineEdit_maxVoltage)
        setting_layout.addRow(QLabel_voltage_step_size, self.QLineEdit_voltage_stepsize)
        setting_layout.addRow(Desp4)
        setting_layout.addRow(QLabel_minCurrent, self.QLineEdit_minCurrent)
        setting_layout.addRow(QLabel_maxCurrent, self.QLineEdit_maxCurrent)
        setting_layout.addRow(QLabel_current_step_size, self.QLineEdit_current_stepsize)
        setting_layout.addRow(Desp5)
        setting_layout.addRow(QLabel_noofloop, self.QComboBox_noofloop)
        setting_layout.addRow(QLabel_updatedelay, self.QComboBox_updatedelay)

        Left_container.addLayout(setting_layout)

        #Main Layout
        Main_Layout = QHBoxLayout()
        Main_Layout.addLayout(Left_container,stretch= 2)
        Main_Layout.addLayout(Right_container,stretch = 1)
        self.setLayout(Main_Layout)
        scroll_area(self,Main_Layout)

        self.QCheckbox_Simulation_Mode.stateChanged.connect(self.simulation_mode_changed)   

        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddress.currentTextChanged.connect(self.DMM_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)

        self.QCheckBox_Current_Accuracy_FULL_Widget.stateChanged.connect(self.checkbox_state_Current_Accuracy_FULL)
        self.QCheckBox_Current_Accuracy_200uA_Widget.stateChanged.connect(self.checkbox_state_Current_Accuracy_200uA)
        self.QCheckBox_Current_Accuracy_2mA_Widget.stateChanged.connect(self.checkbox_state_Current_Accuracy_2mA)

        self.QComboBox_DUT.currentTextChanged.connect(self.DUT_changed)
        self.QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        self.QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        self.QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        self.QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)

        self.QComboBox_External_Auxiliary_Equipment.currentTextChanged.connect(self.External_Auxiliary_Equipment_changed)

        self.QLineEdit_Power.textEdited.connect(self.Power_changed)
        self.QLineEdit_rshunt.textEdited.connect(self.rshunt_changed)
        self.QComboBox_DUT.currentIndexChanged.connect(self.on_current_index_changed)

        self.QLineEdit_minVoltage.textEdited.connect(self.minVoltage_changed)
        self.QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)
        self.QLineEdit_minCurrent.textEdited.connect(self.minCurrent_changed)
        self.QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        self.QLineEdit_voltage_stepsize.textEdited.connect(self.voltage_step_size_changed)
        self.QLineEdit_current_stepsize.textEdited.connect(self.current_step_size_changed)
        self.QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        self.QComboBox_set_PSU_Channel.currentTextChanged.connect(self.PSU_Channel_changed)
        self.QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        self.QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        self.QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)

        self.QComboBox_noofloop.currentTextChanged.connect(self.noofloop_changed)
        self.QComboBox_updatedelay.currentTextChanged.connect(self.updatedelay_changed)

        self.QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)
        self.QCheckBox_Report_Widget.stateChanged.connect(self.checkbox_state_Report)
        self.QCheckBox_Image_Widget.stateChanged.connect(self.checkbox_state_Image)
        QCheckBox_Lock_Widget.stateChanged.connect(self.checkbox_state_lock)

        QPushButton_Widget0.clicked.connect(self.savepath)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget3.clicked.connect(self.estimateTime)
        QPushButton_Widget4.clicked.connect(self.doFind)

        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)
    
    def simulation_mode_changed(self, state):
        if state == Qt.Checked:
            self.simulation_mode = 2
        else:
            self.simulation_mode = 0

    def checkbox_state_Current_Accuracy_FULL(self, state):  #Shamman changes
        if state == Qt.Checked:
            self.checkbox_Current_Accuracy_FULL = 2
        else:
            self.checkbox_Current_Accuracy_FULL = 0
        
    def checkbox_state_Current_Accuracy_200uA(self, state):  #Shamman changes
        if state == Qt.Checked:
            self.checkbox_Current_Accuracy_200uA = 2
        else:
            self.checkbox_Current_Accuracy_200uA = 0

    def checkbox_state_Current_Accuracy_2mA(self, state):  #Shamman changes
        if state == Qt.Checked:
            self.checkbox_Current_Accuracy_2mA = 2
        else:
            self.checkbox_Current_Accuracy_2mA = 0

    def External_Auxiliary_Equipment_changed(self, s):
        if s == "None":
            self.relay_current = "None"
        else:
            self.relay_current = "RELAY"

    def on_current_index_changed(self):
        selected_text = self.QComboBox_DUT.currentText()
        self.update_selection(selected_text)
 
        self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
        self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
        self.QLineEdit_Programming_Error_Offset.setText(self.Programming_Error_Offset)
        self.QLineEdit_Readback_Error_Gain.setText(self.Readback_Error_Gain)
        self.QLineEdit_Readback_Error_Offset.setText(self.Readback_Error_Offset)
        self.QLineEdit_Power.setText(self.Power)
        self.QLineEdit_rshunt.setText(self.rshunt)
        self.QLineEdit_minVoltage.setText(self.minVoltage)
        self.QLineEdit_maxVoltage.setText(self.maxVoltage)
        self.QLineEdit_voltage_stepsize.setText(self.voltage_step_size)
        self.QLineEdit_minCurrent.setText(self.minCurrent)
        self.QLineEdit_maxCurrent.setText(self.maxCurrent)
        self.QLineEdit_current_stepsize.setText(self.current_step_size)

        self.QComboBox_set_PSU_Channel.setCurrentIndex(int(self.PSU_Channel))
        self.QComboBox_set_ELoad_Channel.setCurrentIndex(int(self.ELoad_Channel))
        self.QComboBox_Voltage_Sense.setCurrentText("4 Wire" if self.VoltageSense == "EXT" else "2 Wire")
        self.QComboBox_noofloop.setCurrentText(self.noofloop)
        self.QComboBox_updatedelay.setCurrentText(self.updatedelay)

    def update_selection(self, selected_text):
        """Update selected text and reload config file"""
        self.selected_text = selected_text
        self.load_data()

    def load_data(self):
        """Reads configuration file and returns a dictionary of key-value pairs."""
        config_data = {}
        if self.selected_text =="Excavator":
            self.config_file = os.path.join(config_folder,"config_Excavator.txt")
            
        elif self.selected_text =="Dolphin":
            self.config_file = os.path.join(config_folder,"config_Dolphin.txt")
            
        elif self.selected_text =="SMU":
            self.config_file = os.path.join(config_folder,"config_SMU.txt")

        else:
            self.config_file = os.path.join(config_folder,"config_Others.txt")

        try:
            with open(self.config_file, "r") as file: 
                for line in file:

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue 

                    if "=" in line:
                        key, value = line.strip().split("=")
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    # If savelocation has a valid value, do not overwrite it
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue 
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe("Config file not found. Using default values.")

        return config_data
    
    def rshunt_changed(self, value):
        self.rshunt = value

    def Power_changed(self, value):
        self.Power = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = GetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])
            
            # Add "None" option at the end
            self.QLineEdit_PSU_VisaAddress.addItem("None")
            self.QLineEdit_DMM_VisaAddress.addItem("None")
            self.QLineEdit_ELoad_VisaAddress.addItem("None")
                
        except:
            self.OutputBox.append("No Devices Found!!!")
        return   


    def updatedelay_changed(self, value):
        self.updatedelay = value
        #self.OutputBox.append(str(self.updatedelay))

    def noofloop_changed(self, value):
        self.noofloop = value
        #self.OutputBox.append(str(self.noofloop))

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s    

    def DMM_VisaAddress_changed(self, s):
        self.DMM2 = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s

    def DUT_changed(self, s):
        self.DUT = s

    def ELoad_Channel_changed(self, s):
       self.ELoad_Channel = s

    def PSU_Channel_changed(self, s):
       self.PSU_Channel = s

    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s

    def minVoltage_changed(self, s):
        self.minVoltage = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def minCurrent_changed(self, s):
        self.minCurrent = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def voltage_step_size_changed(self, s):
        self.voltage_step_size = s

    def current_step_size_changed(self, s):
        self.current_step_size = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.VoltageRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setInputZ(self, value):
        AdvancedSettingsList[3] = value

    def checkbox_state_Report(self, s):
        self.checkbox_data_Report = s

    def checkbox_state_Image(self, s):
        self.checkbox_data_Image = s

    def checkbox_state_lock(self, state):
        lockable_widgets = (QPushButton, QLineEdit, QTextEdit, QComboBox)

        for widget in self.findChildren(lockable_widgets):
            widget.setDisabled(state == 2)  # Disable if checkbox is checked

    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value

    def openDialog(self):
        dlg = AdvancedSetting_Current()
        dlg.exec()

    def estimateTime(self):
        
        #self.OutputBox.append(str(self.updatedelay))
        try:
            self.currloop = ((float(self.maxCurrent) - float(self.minCurrent))/ float(self.current_step_size)) + 1
            self.voltloop = ((float(self.maxVoltage) - float(self.minVoltage))/ float(self.voltage_step_size)) + 1

            if self.updatedelay == 0.0:
                constant = 0
                    
            else:
                constant = 1

            self.estimatetime = (self.currloop * self.voltloop *(0.2 + 0.8 + (float(self.updatedelay) * constant) )) * float(self.noofloop)
            self.OutputBox.append(f"{self.estimatetime} seconds")
            self.OutputBox.append("") 
        except Exception as e:
            QMessageBox.warning(self, "Warning", "Input cannot be empty!")
            return

    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

    def executeTest(self):
        global globalvv
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(
                self,
                "Test Already Running",
                "Wait for the active test to finish or stop it first.",
            )
            return
        try:
            for x in range (int(self.noofloop)):

                """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
                then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
                will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
                begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
                begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
                are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
                optionally export all the details into a CSV file or display a graph after the test is completed.

                """
            self.infoList = []
            self.dataList = []
            self.dataList2 = []
            self.setEnabled(False)

            dict = dictGenerator.input(
                V_Rating=self.Voltage_Rating,
 		        rshunt=self.rshunt,
                power=self.Power,
                estimatetime=self.estimatetime,
                updatedelay=self.updatedelay,
                readbackvoltage=self.readbackvoltage,
                readbackcurrent=self.readbackcurrent,
                noofloop=self.noofloop,
                Instrument=self.DMM_Instrument,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,
                unit=self.unit,
                minCurrent=self.minCurrent,
                maxCurrent=self.maxCurrent,
                current_step_size=self.current_step_size,
                minVoltage=self.minVoltage,
                maxVoltage=self.maxVoltage,
                voltage_step_size=self.voltage_step_size,
                selected_DUT=self.selected_text,
                PSU=self.PSU,
                DMM2=self.DMM2,
                OperationMode=self.SPOperationMode,
                ELoad=self.ELoad,
                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],
            )

            """ # Check for missing parameters
            missing = self.check_missing_params(dict)
            if missing:
                print_console_safe(f"The following parameters are missing or empty: {missing}")
                return"""

            """#Function: Visa Address Check & Run Test
            A = VisaResourceManager()
            flag, args = A.openRM(self.ELoad, self.PSU, self.DMM)
            if flag == 0:
                string = ""
                for item in args:
                    string = string + item

                QMessageBox.warning(self, "VISA IO ERROR", string)
                return"""
            if self.simulation_mode == 2:
                    QMessageBox.information(
                        self,
                        "Simulation Mode",
                        "Simulation Mode is enabled. No instruments are connected.",
                    )
                    
                    wb = load_workbook(self.output_file)
                    ws = wb["Data"]  # get the worksheet

                    graph_plot = Graph_Plotting(
                        wb,        # workbook
                        ws,        # sheet
                        self.output_file,
                        "Voltage",
                        1,
                        [8],       # graph_start_rows
                    )
                    graph_plot.execute_plot()

                    wb.save(self.output_file)  # save the charts back
                    sleep(2)

            else:
                QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
                                    )
        
                #Execute Voltage Measurement
                """relay_current = RelayController_Current()
                if self.relay_current == "RELAY":

                    relay_current.relay_on()
                else:
                    relay_current.relay_off()"""

                if self.DMM_Instrument == "Keysight":

                    if self.selected_text == "Dolphin":
                        if self.ELoad != "None" and self.DMM2 != "None":
                            print_console_safe("ELoad connected and DMM connected") #All connected
                            try:(
                                self.infoList,
                                self.dataList,
                                self.dataList2
                                ) = DolphinNewCurrentMeasurementwithELoad.executeCurrentMeasurementA(self, dict, self.PSU_Channel)
                        
                            except Exception as e:
                                QMessageBox.warning(self, "Error", str(e))
                                return

                        elif self.ELoad == "None" and self.DMM2 != "None":
                            print_console_safe("No ELoad connected and DMM connected") #No Eload connected but DMM connected
                            try:(
                                self.infoList,
                                self.dataList,
                                self.dataList2
                                ) = DolphinNewCurrentMeasurementNoELoadWithDMM.executeCurrentMeasurementA(self, dict, self.PSU_Channel)
                            
                            except Exception as e:
                                QMessageBox.warning(self, "Error", str(e))
                                return
                        
                        elif self.ELoad != "None" and self.DMM2 == "None":
                            print_console_safe("ELoad connected and No DMM connected") #Eload connected but no DMM connected
                            try:(
                                self.infoList,
                                self.dataList,
                                self.dataList2
                                ) = DolphinNewCurrentMeasurementwithELoadNoDMM.executeCurrentMeasurementA(self, dict, self.PSU_Channel)
                            except Exception as e:
                                QMessageBox.warning(self, "Error", str(e))
                                return
                        else :
                            print_console_safe("No ELoad connected and No DMM connected") #No Eload connected and
                            try:(
                                self.infoList,
                                self.dataList,
                                self.dataList2
                                ) = DolphinNewCurrentMeasurementNoELoadNoDMM.executeCurrentMeasurementA(self, dict, self.PSU_Channel)
                            except Exception as e:
                                QMessageBox.warning(self, "Error", str(e))
                                return
                            
                    elif self.selected_text == "Hornbill":
                        if self.ELoad != "None" and self.DMM != "None":
                            if self.checkbox_Current_Accuracy_FULL == 2:
                                print_console_safe("ELoad connected and DMM connected") #All connected
                                try:(
                                    self.infoList,   
                                    self.dataList,
                                    self.dataList2
                                ) = HornbillCurrentMeasurementwithELoad_IMON_FULL.Execute_Current_Accuracy_Current_Static(self, dict, self.PSU_Channel)
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return
                            if self.checkbox_Current_Accuracy_200uA == 2:
                                print_console_safe("ELoad connected and DMM connected") #All connected
                                try:(
                                    self.infoList,   
                                    self.dataList,
                                    self.dataList2
                                ) = HornbillCurrentMeasurementwithELoad_IMON_200uA.Execute_Current_Accuracy_Current_Static(self, dict, self.PSU_Channel)
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return   
                            if self.checkbox_Current_Accuracy_2mA == 2:
                                print_console_safe("ELoad connected and DMM connected") #All connected
                                try:(
                                    self.infoList,   
                                    self.dataList,
                                    self.dataList2
                                ) = HornbillCurrentMeasurementwithELoad_IMON_2mA.Execute_Current_Accuracy_Current_Static(self, dict, self.PSU_Channel)
                                except Exception as e:
                                    QMessageBox.warning(self, "Error", str(e))
                                    return 
                            
                        else:
                            print_console_safe("No ELoad connected and DMM connected") #No Eload connected but DMM connected
                            
                #relay_current.relay_off()
                            
                if x == (int(self.noofloop) - 1):   
                    self.OutputBox.append(my_result.getvalue())
                    self.OutputBox.append("Measurement is complete !")


                    if self.checkbox_data_Report == 2:
                        instrumentData(self.PSU, self.DMM2, self.ELoad)
                        datatoCSV_Accuracy(self.infoList, self.dataList, self.dataList2)
                        datatoGraph2(self.infoList, self.dataList, self.dataList2)
                        datatoGraph2.scatterCompareCurrent2(self, float(self.Programming_Error_Gain), float(self.Programming_Error_Offset), float(self.Readback_Error_Gain), float(self.Readback_Error_Offset), str(self.unit), float(self.Current_Rating))

                        A = xlreport(save_directory=self.savelocation, file_name=str(self.unit))
                        A.run()
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        df.to_csv(os.path.join(csv_folder,"config.csv"))

            
                    if self.checkbox_data_Image == 2:
                        self.image_dialog = image_Window()
                        self.image_dialog.setModal(True)
                        self.image_dialog.show()    



        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

            return
        finally:
            self.setEnabled(True)

class CV_LoadRegulationDialog(QDialog):
    """Class for configuring the Load Regulation under CV Mode DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.


    """

    def __init__(self):
        """ "Method declaring the Widgets, Signals & Slots for Load Regulation under CV Mode."""
        super().__init__()
         # Default Values
        self.Power_Rating = "2200"
        self.Current_Rating = "120"
        self.Voltage_Rating = "80"
        self.PSU = ""
        self.DMM = ""
        self.ELoad = ""
        self.ELoad_Channel = "1"
        self.PSU_Channel = "1"
        self.DMM_Instrument = "Keysight"
        self.Programming_Error_Offset = "0.0001"
        self.Programming_Error_Gain = "0.0001"
        self.Readback_Error_Offset = "0.0001"
        self.Readback_Error_Gain = "0.0001"
        
        self.SPOperationMode = "Independent"
        self.relay_voltage = "None"

        self.updatedelay = "5"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"

        self.setFunction = "Current"
        self.VoltageRes = "SLOW"
        self.VoltageSense = "EXT"
        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.Range = "Auto"
        self.Aperture = "10"
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.UpTime = "50"
        self.DownTime = "50"

        self.setWindowTitle("Load Regulation (CV)")
        self.image_window = None
        self.setWindowFlags(Qt.Window)

        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
         #Output Display
        self.OutputBox = QTextBrowser()
        self.OutputBox.append(f"{my_result.getvalue()}")
        self.OutputBox.append("")  # Empty line after each append

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp8 = QLabel()

        Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Specification:")
        Desp8.setText("External Auxiliary Equipment:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

       # Connections section
        QLabel_PSU_VisaAddress = QLabel()
        QLabel_DMM_VisaAddress = QLabel()
        QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DMM_Instrument = QLabel()
        QLabel_DUT = QLabel()
        QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        QLabel_DMM_VisaAddress.setText("Visa Address (DMM):")
        QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        QLabel_DUT.setText("DUT:")
        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        self.QComboBox_DMM_Instrument = QComboBox()
        self.QComboBox_DUT = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Power_Rating = QLabel()
        QLabel_Max_Voltage = QLabel()
        QLabel_Max_Current = QLabel()
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Power_Rating.setText("Power Rating (W):")
        QLabel_Max_Voltage.setText("Maximum Voltage (V):")
        QLabel_Max_Current.setText("Maximum Current (A):")
        QLabel_Programming_Error_Gain.setText("Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Readback Desired Specification (Offset):")

        # External Auxiliary Equipment section
        QLabel_External_Auxiliary_Equipment = QLabel()
        QLabel_External_Auxiliary_Equipment.setText("Relay")
        self.QComboBox_External_Auxiliary_Equipment = QComboBox()
        self.QComboBox_External_Auxiliary_Equipment.addItems(["None", "RELAY"])

        self.QComboBox_Voltage_Res = QComboBox()
        self.QComboBox_set_PSU_Channel = QComboBox()
        self.QComboBox_set_ELoad_Channel = QComboBox()
        self.QComboBox_Voltage_Sense = QComboBox()
        self.QComboBox_set_Function = QComboBox()
        self.QLineEdit_Power_Rating = QLineEdit()
        self.QLineEdit_Max_Voltage = QLineEdit()
        self.QLineEdit_Max_Current = QLineEdit()
        self.QLineEdit_Programming_Error_Gain = QLineEdit()
        self.QLineEdit_Programming_Error_Offset = QLineEdit()
        self.QLineEdit_Readback_Error_Gain = QLineEdit()
        self.QLineEdit_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_DMM_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QComboBox_DUT.addItems(["Others", "Excavator", "EDU36311A", "Dolphin", "Hornbill", "SMU"])

        self.QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        self.QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        self.QComboBox_set_Function.addItems(
            [
                "Current Priority✅",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )

        self.QComboBox_set_Function.setEnabled(False)
        self.QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        self.QComboBox_set_PSU_Channel.setEnabled(True)
        self.QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        self.QComboBox_set_ELoad_Channel.setEnabled(True)
        self.QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])
        self.QComboBox_Voltage_Sense.setEnabled(True)

        #Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QVBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
    
        #Execute Layout + Outputbox in Right Container
        Right_container = QVBoxLayout()
        exec_layout_box = QHBoxLayout()
        exec_layout = QFormLayout()

        #exec_layout.addRow(Desp0)
        exec_layout.addWidget(self.OutputBox)
        exec_layout.addRow(QPushButton_Widget0)

        exec_layout.addRow(QPushButton_Widget2)
        exec_layout.addRow(QPushButton_Widget1)   

        exec_layout_box.addLayout(exec_layout)
 
        Right_container.addLayout(save_path_layout)
        Right_container.addLayout(exec_layout_box)

        #Setting Form Layout with Left Container
        Left_container = QHBoxLayout()
        setting_layout = QFormLayout()

        setting_layout .addRow(Desp0)
        setting_layout.addRow(QPushButton_Widget0)

        setting_layout.addRow(Desp1)
        setting_layout.addRow(QPushButton_Widget4)
        setting_layout.addRow(QLabel_DUT, self.QComboBox_DUT)
        setting_layout.addRow(QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        setting_layout.addRow(QLabel_DMM_VisaAddress, self.QLineEdit_DMM_VisaAddress)
        setting_layout.addRow(QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        setting_layout.addRow(QLabel_DMM_Instrument, self.QComboBox_DMM_Instrument)
        setting_layout.addRow(Desp2)
        setting_layout.addRow(QLabel_set_PSU_Channel, self.QComboBox_set_PSU_Channel)
        setting_layout.addRow(QLabel_set_ELoad_Channel, self.QComboBox_set_ELoad_Channel)
        setting_layout.addRow(QLabel_set_Function, self.QComboBox_set_Function)
        setting_layout.addRow(QLabel_Voltage_Sense, self.QComboBox_Voltage_Sense)
        setting_layout.addRow(Desp8)
        setting_layout.addRow(QLabel_External_Auxiliary_Equipment, self.QComboBox_External_Auxiliary_Equipment)
        setting_layout.addRow(Desp3)
        setting_layout.addRow(QLabel_Power_Rating, self.QLineEdit_Power_Rating)
        setting_layout.addRow(QLabel_Max_Voltage, self.QLineEdit_Max_Voltage)
        setting_layout.addRow(QLabel_Max_Current, self.QLineEdit_Max_Current)
        setting_layout.addRow(QLabel_Programming_Error_Gain, self.QLineEdit_Programming_Error_Gain)
        setting_layout.addRow(QLabel_Programming_Error_Offset, self.QLineEdit_Programming_Error_Offset)
        setting_layout.addRow(QLabel_Readback_Error_Gain, self.QLineEdit_Readback_Error_Gain)
        setting_layout.addRow(QLabel_Readback_Error_Offset, self.QLineEdit_Readback_Error_Offset)
        setting_layout.addRow(QPushButton_Widget2)
        setting_layout.addRow(QPushButton_Widget1)
        setting_layout.addRow(self.OutputBox)
        Left_container.addLayout(setting_layout)
        #Main Layout
        Main_Layout = QHBoxLayout()
        Main_Layout.addLayout(Left_container,stretch= 2)
        Main_Layout.addLayout(Right_container,stretch = 1)
        self.setLayout(Main_Layout)
        scroll_area(self,Main_Layout)

       
        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)
        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddress.currentTextChanged.connect(self.DMM_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)
        self.QComboBox_DUT.currentTextChanged.connect(self.DUT_changed)
        self.QComboBox_set_PSU_Channel.currentTextChanged.connect(self.PSU_Channel_changed)
        self.QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        self.QComboBox_External_Auxiliary_Equipment.currentTextChanged.connect(self.External_Auxiliary_Equipment_changed)
        self.QComboBox_DUT.currentIndexChanged.connect(self.on_current_index_changed)
        self.QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        self.QLineEdit_Power_Rating.textEdited.connect(self.Power_Rating_changed)
        self.QLineEdit_Max_Current.textEdited.connect(self.Max_Current_changed)
        self.QLineEdit_Max_Voltage.textEdited.connect(self.Max_Voltage_changed)
        self.QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        self.QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        self.QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        self.QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)
        self.QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        self.QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)
        self.QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)
        QPushButton_Widget4.clicked.connect(self.doFind)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget0.clicked.connect(self.savepath)
    
    def External_Auxiliary_Equipment_changed(self, s):
        if s == "None":
            self.relay_voltage = "None"
        else:
            self.relay_voltage = "RELAY"

    def on_current_index_changed(self):
        selected_text = self.QComboBox_DUT.currentText()
        self.update_selection(selected_text)
        self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
        self.QLineEdit_Programming_Error_Offset.setText(self.Programming_Error_Offset)
        self.QLineEdit_Readback_Error_Gain.setText(self.Readback_Error_Gain)
        self.QLineEdit_Readback_Error_Offset.setText(self.Readback_Error_Offset)
        self.QLineEdit_Power_Rating.setText(self.Power_Rating)
        self.QLineEdit_Max_Voltage.setText(self.Voltage_Rating)
        self.QLineEdit_Max_Current.setText(self.Current_Rating)
        self.QComboBox_set_PSU_Channel.setCurrentIndex(int(self.PSU_Channel))
        self.QComboBox_set_ELoad_Channel.setCurrentIndex(int(self.ELoad_Channel))
        self.QComboBox_Voltage_Sense.setCurrentText("4 Wire" if self.VoltageSense == "EXT" else "2 Wire")


    
    def update_selection(self, selected_text):
        """Update selected text and reload config file"""
        self.selected_text = selected_text
        self.load_data()
    
    
    def load_data(self):
        """Reads configuration file and returns a dictionary of key-value pairs."""
        config_data = {}
        if self.selected_text =="Excavator":
            self.config_file = os.path.join(config_folder,"config_Excavator.txt")
            
        elif self.selected_text =="Dolphin":
            self.config_file = os.path.join(config_folder,"config_Dolphin.txt")
            
        elif self.selected_text =="SMU":
            self.config_file = os.path.join(config_folder,"config_SMU.txt")

        else:
            self.config_file = os.path.join(config_folder,"config_Others.txt")

        try:
            with open(self.config_file, "r") as file: 
                for line in file:

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue 

                    if "=" in line:
                        key, value = line.strip().split("=")
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    # If savelocation has a valid value, do not overwrite it
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue 
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe("Config file not found. Using default values.")

        return config_data


    def DUT_changed(self, s):
        self.DUT = s
    
    def PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = GetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])
            
            # Add "None" option at the end
            self.QLineEdit_PSU_VisaAddress.addItem("None")
            self.QLineEdit_DMM_VisaAddress.addItem("None")
            self.QLineEdit_ELoad_VisaAddress.addItem("None")
                
        except:
            self.OutputBox.append("No Devices Found!!!")
        return 
    
    def updatedelay_changed(self, value):
        self.updatedelay = value
        self.OutputBox.append(str(self.updatedelay))

    def RangeChanged(self, s):
        AdvancedSettingsList[0] = s

    def ApertureChanged(self, s):
        AdvancedSettingsList[1] = s

    def AutoZeroChanged(self, s):
        AdvancedSettingsList[2] = s

    def InputZChanged(self, s):
        AdvancedSettingsList[3] = s

    def UpTimeChanged(self, s):
        AdvancedSettingsList[4] = s

    def DownTimeChanged(self, s):
        AdvancedSettingsList[5] = s

    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s

    def Power_Rating_changed(self, s):
        self.Power_Rating = s

    def Max_Current_changed(self, s):
        self.Current_Rating = s

    def Max_Voltage_changed(self, s):
        self.Voltage_Rating = s

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s

    def DMM_VisaAddress_changed(self, s):
        self.DMM = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s

    #def ELoad_Channel_changed(self, s):
       # self.ELoad_Channel = s

    #def PSU_Channel_changed(self, s):
        #self.PSU_Channel = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.VoltageRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setInputZ(self, value):
        AdvancedSettingsList[3] = value
    
    def checkbox_state_Report(self, s):
        self.checkbox_data_Report = s

    def checkbox_state_Image(self, s):
        self.checkbox_data_Image = s


    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value
    
    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

    def executeTest(self):
        try:
            """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
            then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
            will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
            begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
            begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
            are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
            optionally export all the details into a CSV file or display a graph after the test is completed.

                """
            dict = dictGenerator.input(
                
                savedir=self.savelocation,
                Instrument=self.DMM_Instrument,
                #Error_Gain=self.Error_Gain,
                #Error_Offset=self.Error_Offset,
                updatedelay=self.updatedelay,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,
                V_Rating=self.Voltage_Rating,
                I_Rating=self.Current_Rating,
                P_Rating=self.Power_Rating,
                power = self.Power_Rating,
                maxVoltage=self.Voltage_Rating,
                maxCurrent=self.Current_Rating,
                Load_Programming_Error_Gain=self.Programming_Error_Gain,
                Load_Programming_Error_Offset=self.Programming_Error_Offset,
                Load_Readback_Error_Gain=self.Readback_Error_Gain,  
                PSU=self.PSU,
                DMM=self.DMM,
                ELoad=self.ELoad,
                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                OperationMode=self.SPOperationMode,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],
            )
            QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
            )
            #Execute Voltage Measurement
            relay_voltage = RelayController_Voltage()
            if self.relay_voltage == "RELAY":

                relay_voltage.relay_on()
            else:
                relay_voltage.relay_off()

            if self.DMM_Instrument == "Keysight":
                if self.selected_text == "Dolphin":
                    if self.ELoad != "None" and self.DMM != "None":
                        try:
                            DolphinLoadRegulationwithELoad.executeCV_LoadRegulationA(self, dict)
                        except Exception as e:
                            QMessageBox.warning(self, "Error", str(e))
                            exit()
                    elif self.ELoad == "None" and self.DMM != "None":
                        self.OutputBox.append("No ELoad connected and DMM connected") #No Eload connected but DMM connected)
                        return
                    elif self.ELoad != "None" and self.DMM == "None":
                        self.OutputBox.append("ELoad connected and No DMM connected") #Eload connected
                        return
                    else :
                        self.OutputBox.append("No ELoad connected and No DMM connected") #No Eload connected and No DMM connected
                        return

            relay_voltage.relay_off()
            self.OutputBox.append(my_result.getvalue())
            self.OutputBox.append("Measurement is complete !")
        
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
    

    def openDialog(self):
        dlg = AdvancedSetting_Voltage()
        dlg.exec()

class CC_LoadRegulationDialog(QDialog):
    """Class for configuring the  Load Regulation under CC Mode DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.

    """

    def __init__(self):
        """ "Method declaring the Widgets, Signals & Slots for Load Regulation under CV Mode."""
        super().__init__()

        # Default Values
        self.rshunt = "0.01"
        self.Power_Rating = "2200"
        self.Power = self.Power_Rating
        self.Current_Rating = "100"
        self.Voltage_Rating = "80"
        self.maxCurrent = self.Current_Rating
        self.maxVoltage = self.Voltage_Rating
        self.PSU = ""
        self.DMM2 = ""
        self.ELoad = ""
        self.ELoad_Channel = ""
        self.PSU_Channel = ""
        self.DMM_Instrument = "Keysight"
        self.DMM_Instrument = "Keysight"
        self.Programming_Error_Offset = "0.0005"
        self.Programming_Error_Gain = "0.0005"
        self.Readback_Error_Offset = "0.0005"
        self.Readback_Error_Gain = "0.0005"

        self.SPOperationMode = "Independent"
        self.relay_current = "None"

        self.updatedelay = "4"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"

        self.setFunction = "Voltage"
        self.VoltageRes = "SLOW"
        self.VoltageSense = "INT"
        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.Range = "Auto"
        self.Aperture = "10"
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.Terminal = "3A"
        self.UpTime = "50"
        self.DownTime = "50"

        self.setWindowTitle("Load Regulation (CC)")
        self.image_window = None
        self.setWindowFlags(Qt.Window)

        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
        
         #Output Display
        self.OutputBox = QTextBrowser()
        self.OutputBox.append(f"{my_result.getvalue()}")
        self.OutputBox.append("")  # Empty line after each append

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp8 = QLabel()
        
        Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Specification:")
        Desp8.setText("External Auxiliary Equipment:")

         #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location:")

       # Connections section
        QLabel_PSU_VisaAddress = QLabel()
        QLabel_DMM_VisaAddress = QLabel()
        QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DMM_Instrument = QLabel()
        QLabel_DUT = QLabel()
        QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        QLabel_DMM_VisaAddress.setText("Visa Address (DMM):")
        QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        QLabel_DUT.setText("DUT:")
        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        self.QComboBox_DMM_Instrument = QComboBox()
        self.QComboBox_DUT = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Power_Rating = QLabel()
        QLabel_Max_Voltage = QLabel()
        QLabel_Max_Current = QLabel()
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Power_Rating.setText("Power Rating (W):")
        QLabel_Max_Voltage.setText("Maximum Voltage (V):")
        QLabel_Max_Current.setText("Maximum Current (A):")
        QLabel_Programming_Error_Gain.setText("Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Readback Desired Specification (Offset):")

        #Shunt
        QLabel_rshunt = QLabel()
        QLabel_rshunt.setText("Shunt Resistance Value (ohm):")
        self.QLineEdit_rshunt = QLineEdit()

        # External Auxiliary Equipment section
        QLabel_External_Auxiliary_Equipment = QLabel()
        QLabel_External_Auxiliary_Equipment.setText("Relay")
        self.QComboBox_External_Auxiliary_Equipment = QComboBox()
        self.QComboBox_External_Auxiliary_Equipment.addItems(["None", "RELAY"])

        self.QComboBox_Voltage_Res = QComboBox()
        self.QComboBox_set_PSU_Channel = QComboBox()
        self.QComboBox_set_ELoad_Channel = QComboBox()
        self.QComboBox_Voltage_Sense = QComboBox()
        self.QComboBox_set_Function = QComboBox()
        self.QLineEdit_Power_Rating = QLineEdit()
        self.QLineEdit_Max_Voltage = QLineEdit()
        self.QLineEdit_Max_Current = QLineEdit()
        self.QLineEdit_Programming_Error_Gain = QLineEdit()
        self.QLineEdit_Programming_Error_Offset = QLineEdit()
        self.QLineEdit_Readback_Error_Gain = QLineEdit()
        self.QLineEdit_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_DMM_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QComboBox_DUT.addItems(["Others", "Excavator", "EDU36311A", "Dolphin", "Hornbill", "SMU"])

        self.QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        self.QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        self.QComboBox_set_Function.addItems(
            [
                "Current Priority✅",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )

        self.QComboBox_set_Function.setEnabled(False)
        self.QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        self.QComboBox_set_PSU_Channel.setEnabled(True)
        self.QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        self.QComboBox_set_ELoad_Channel.setEnabled(True)
        self.QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])
        self.QComboBox_Voltage_Sense.setEnabled(True)
        
        #Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QVBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
    
        #Execute Layout + Outputbox in Right Container
        Right_container = QVBoxLayout()
        exec_layout_box = QHBoxLayout()
        exec_layout = QFormLayout()

        #exec_layout.addRow(Desp0)
        exec_layout.addWidget(self.OutputBox)
        exec_layout.addRow(QPushButton_Widget0)

        exec_layout.addRow(QPushButton_Widget2)
        exec_layout.addRow(QPushButton_Widget1)   

        exec_layout_box.addLayout(exec_layout)
 
        Right_container.addLayout(save_path_layout)
        Right_container.addLayout(exec_layout_box)

        #Setting Form Layout with Left Container
        Left_container = QHBoxLayout()
        setting_layout = QFormLayout()

        setting_layout .addRow(Desp0)
        setting_layout.addRow(QPushButton_Widget0)

        setting_layout.addRow(Desp1)
        setting_layout.addRow(QPushButton_Widget4)
        setting_layout.addRow(QLabel_DUT, self.QComboBox_DUT)
        setting_layout.addRow(QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        setting_layout.addRow(QLabel_DMM_VisaAddress, self.QLineEdit_DMM_VisaAddress)
        setting_layout.addRow(QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        setting_layout.addRow(QLabel_DMM_Instrument, self.QComboBox_DMM_Instrument)
        setting_layout.addRow(Desp2)
        setting_layout.addRow(QLabel_set_PSU_Channel, self.QComboBox_set_PSU_Channel)
        setting_layout.addRow(QLabel_set_ELoad_Channel, self.QComboBox_set_ELoad_Channel)
        setting_layout.addRow(QLabel_set_Function, self.QComboBox_set_Function)
        setting_layout.addRow(QLabel_Voltage_Sense, self.QComboBox_Voltage_Sense)
        setting_layout.addRow(QLabel_rshunt, self.QLineEdit_rshunt)
        setting_layout.addRow(Desp8)
        setting_layout.addRow(QLabel_External_Auxiliary_Equipment, self.QComboBox_External_Auxiliary_Equipment)
        setting_layout.addRow(Desp3)
        setting_layout.addRow(QLabel_Power_Rating, self.QLineEdit_Power_Rating)
        setting_layout.addRow(QLabel_Max_Voltage, self.QLineEdit_Max_Voltage)
        setting_layout.addRow(QLabel_Max_Current, self.QLineEdit_Max_Current)
        setting_layout.addRow(QLabel_Programming_Error_Gain, self.QLineEdit_Programming_Error_Gain)
        setting_layout.addRow(QLabel_Programming_Error_Offset, self.QLineEdit_Programming_Error_Offset)
        setting_layout.addRow(QLabel_Readback_Error_Gain, self.QLineEdit_Readback_Error_Gain)
        setting_layout.addRow(QLabel_Readback_Error_Offset, self.QLineEdit_Readback_Error_Offset)
        setting_layout.addRow(QPushButton_Widget2)
        setting_layout.addRow(QPushButton_Widget1)
        setting_layout.addRow(self.OutputBox)
        Left_container.addLayout(setting_layout)
        #Main Layout
        Main_Layout = QHBoxLayout()
        Main_Layout.addLayout(Left_container,stretch= 2)
        Main_Layout.addLayout(Right_container,stretch = 1)
        self.setLayout(Main_Layout)
        scroll_area(self,Main_Layout)

        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)
        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddress.currentTextChanged.connect(self.DMM_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)
        self.QComboBox_DUT.currentTextChanged.connect(self.DUT_changed)
        self.QComboBox_set_PSU_Channel.currentTextChanged.connect(self.PSU_Channel_changed)
        self.QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        self.QLineEdit_rshunt.textEdited.connect(self.rshunt_changed)
        self.QComboBox_External_Auxiliary_Equipment.currentTextChanged.connect(self.External_Auxiliary_Equipment_changed)
        self.QComboBox_DUT.currentIndexChanged.connect(self.on_current_index_changed)
        self.QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        self.QLineEdit_Power_Rating.textEdited.connect(self.Power_Rating_changed)
        self.QLineEdit_Max_Current.textEdited.connect(self.Max_Current_changed)
        self.QLineEdit_Max_Voltage.textEdited.connect(self.Max_Voltage_changed)
        self.QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        self.QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        self.QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        self.QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)
        self.QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        self.QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)
        self.QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)
        QPushButton_Widget4.clicked.connect(self.doFind)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget0.clicked.connect(self.savepath)
    
    def External_Auxiliary_Equipment_changed(self, s):
        if s == "None":
            self.relay_current = "None"
        else:
            self.relay_current = "RELAY"

    def on_current_index_changed(self):
        selected_text = self.QComboBox_DUT.currentText()
        self.update_selection(selected_text)
        self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
        self.QLineEdit_Programming_Error_Offset.setText(self.Programming_Error_Offset)
        self.QLineEdit_Readback_Error_Gain.setText(self.Readback_Error_Gain)
        self.QLineEdit_Readback_Error_Offset.setText(self.Readback_Error_Offset)
        self.QLineEdit_Power_Rating.setText(self.Power_Rating)
        self.QLineEdit_Max_Voltage.setText(self.Voltage_Rating)
        self.QLineEdit_Max_Current.setText(self.Current_Rating)
        self.QComboBox_set_PSU_Channel.setCurrentIndex(int(self.PSU_Channel))
        self.QComboBox_set_ELoad_Channel.setCurrentIndex(int(self.ELoad_Channel))
        self.QComboBox_Voltage_Sense.setCurrentText("4 Wire" if self.VoltageSense == "EXT" else "2 Wire")
        self.QLineEdit_rshunt.setText(self.rshunt)

    def rshunt_changed(self, value):
        self.rshunt = value

    def update_selection(self, selected_text):
        """Update selected text and reload config file"""
        self.selected_text = selected_text
        self.load_data()
    
    
    def load_data(self):
        """Reads configuration file and returns a dictionary of key-value pairs."""
        config_data = {}
        if self.selected_text =="Excavator":
            self.config_file = os.path.join(config_folder,"config_Excavator.txt")
            
        elif self.selected_text =="Dolphin":
            self.config_file = os.path.join(config_folder,"config_Dolphin.txt")
            
        elif self.selected_text =="SMU":
            self.config_file = os.path.join(config_folder,"config_SMU.txt")

        else:
            self.config_file = os.path.join(config_folder,"config_Others.txt")

        try:
            with open(self.config_file, "r") as file: 
                for line in file:

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue 

                    if "=" in line:
                        key, value = line.strip().split("=")
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    # If savelocation has a valid value, do not overwrite it
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue 
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe("Config file not found. Using default values.")

        return config_data


    def DUT_changed(self, s):
        self.DUT = s
    
    def PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s
    
    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = GetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])
            
            # Add "None" option at the end
            self.QLineEdit_PSU_VisaAddress.addItem("None")
            self.QLineEdit_DMM_VisaAddress.addItem("None")
            self.QLineEdit_ELoad_VisaAddress.addItem("None")
                
        except:
            self.OutputBox.append("No Devices Found!!!")
        return 
    
    def updatedelay_changed(self, value):
        self.updatedelay = value
        self.OutputBox.append(str(self.updatedelay))
    
    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s


    def Power_Rating_changed(self, s):
        self.Power_Rating = s

    def Max_Current_changed(self, s):
        self.Current_Rating = s

    def Max_Voltage_changed(self, s):
        self.Voltage_Rating = s

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s

    def DMM_VisaAddress_changed(self, s):
        self.DMM2 = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.CurrentRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.CurrentSense = "INT"
        elif s == "4 Wire":
            self.CurrentSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setTerminal(self, value):
        AdvancedSettingsList[3] = value

    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value
    
    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

    def executeTest(self):
        try:
            """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
            then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
            will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
            begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
            begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
            are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
            optionally export all the details into a CSV file or display a graph after the test is completed.

            """
            dict = dictGenerator.input(
                rshunt=self.rshunt,
                savedir=self.savelocation,
                Instrument=self.DMM_Instrument,
                #Error_Gain=self.Error_Gain,
                #Error_Offset=self.Error_Offset,
                updatedelay=self.updatedelay,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,
                V_Rating=self.Voltage_Rating,
                I_Rating=self.Current_Rating,
                P_Rating=self.Power_Rating,
                power=self.Power,
                maxVoltage=self.Voltage_Rating,
                maxCurrent=self.Current_Rating,
                PSU=self.PSU,
                DMM2=self.DMM2,
                ELoad=self.ELoad,
                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                OperationMode=self.SPOperationMode,
                #CurrentSense=self.CurrentSense,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                #Terminal=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],
            )
            QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
            )
            #Execute Voltage Measurement
            relay_current = RelayController_Current()
            if self.relay_current == "RELAY":

                relay_current.relay_on()
            else:
                relay_current.relay_off()

            if self.DMM_Instrument == "Keysight":
                if self.selected_text == "Dolphin":
                    if self.ELoad != "None" and self.DMM2 != "None":
                        try:
                            DolphinLoadRegulationwithELoad.executeCC_LoadRegulationA(self, dict)

                        except Exception as e:
                            QMessageBox.warning(self, "Error", str(e))
                            return

            relay_current.relay_off()
            self.OutputBox.append(my_result.getvalue())
            self.OutputBox.append("Measurement is complete !")

        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
            return

    def openDialog(self):
        dlg = AdvancedSetting_Current()
        dlg.exec()

class TransientRecoveryTime(QDialog):

    """Class for configuring the Transient Recovery Time DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.

    """

    def __init__(self):
        """ "Method declaring the Widgets, Signals & Slots for Transient Recovery Time."""
        super().__init__()

        # Default Values
        self.selected_text = "Dolphin"
        self.PSU = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        self.OSC = "USB0::0x0957::0x17B0::MY52060151::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        self.ELoad_Channel = "1"
        self.PSU_Channel = "1"
        self.OSC_Channel = "1"
        self.setFunction = "Current"
        self.relay_voltage = "None"

        self.VoltageSense = "EXT"
        self.Power_Rating = "160"
        self.Current_Rating = "120"
        self.Voltage_Rating = "80"
        self.maxCurrent = "10"
        self.maxVoltage = "80"

        self.Channel_CouplingMode = "AC"
        self.Trigger_CouplingMode = "DC"
        self.Trigger_Mode = "EDGE"
        self.Trigger_SweepMode = "NORMAL"
        self.Trigger_SlopeMode = "EITHer"
        self.TimeScale = "0.01"
        self.VerticalScale = "0.00001"
        self.I_Step = ""
        self.V_Settling_Band = "0.8"
        self.T_Settling_Band = "0.001"
        self.Probe_Setting = "X10"
        self.Acq_Type = "AVERage"

        self.checkbox_SpecialCase = 2
        self.checkbox_NormalCase = 2

        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"


        self.setWindowTitle("Transient Recovery Time")

        QPushButton_Widget00 = QPushButton()
        QPushButton_Widget00.setText("Save Path")
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Find Instruments")
        QPushButton_Widget = QPushButton()
        QPushButton_Widget.setText("Execute Test")
        QCheckBox_SpecialCase_Widget = QCheckBox()
        QCheckBox_SpecialCase_Widget.setText("Special Case (0% <-> 100%)")
        QCheckBox_SpecialCase_Widget.setCheckState(Qt.Checked)
        QCheckBox_NormalCase_Widget = QCheckBox()
        QCheckBox_NormalCase_Widget.setText("Normal Case (50% <-> 100%)")
        QCheckBox_NormalCase_Widget.setCheckState(Qt.Checked)
        
        layout1 = QFormLayout()
        self.OutputBox = QTextBrowser()

        self.OutputBox.append(my_result.getvalue())

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()

        Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Specification:")
        Desp4.setText("Oscilloscope Settings:")
        Desp5.setText("Perform Test:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        # Connections
        self.QLabel_PSU_VisaAddress = QLabel()
        self.QLabel_OSC_VisaAddress = QLabel()
        self.QLabel_ELoad_VisaAddress = QLabel()
 
        self.QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        self.QLabel_OSC_VisaAddress.setText("Visa Address (OSC):")
        self.QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_OSC_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()

        # General Settings
        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_OSC_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])

        QLabel_OSC_Display_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Power_Rating = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_rated = QLabel()
        QLabel_current_rated = QLabel()
        QLabel_maxCurrent = QLabel()

        QLabel_V_Settling_Band = QLabel()
        QLabel_T_Settling_Band = QLabel()
        QLabel_Probe_Setting = QLabel()
        QLabel_Acq_Type = QLabel()


        """QLabel_ELoad_Display_Channel.setText("Display Channel (Eload):")
        QLabel_PSU_Display_Channel.setText("Display Channel (PSU):")"""
        QLabel_OSC_Display_Channel.setText("Display Channel (OSC)")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Power_Rating.setText("Power Rating (W):")
        QLabel_maxVoltage.setText("Testing Voltage (V):")
        QLabel_voltage_rated.setText("DUT Rated Voltage:")
        QLabel_current_rated.setText("DUT Rated Current:")
        QLabel_maxCurrent.setText("Testing Current (A):")
        
        QLabel_V_Settling_Band.setText("Settling Band Voltage (V) / Error Band:")
        QLabel_T_Settling_Band.setText("Settling Band Time (s):")
        QLabel_Probe_Setting.setText("Probe Setting Ratio:")
        QLabel_Acq_Type.setText("Acquire Mode:")


        """QLineEdit_ELoad_Display_Channel = QLineEdit()
        QLineEdit_PSU_Display_Channel = QLineEdit()"""
        QLineEdit_OSC_Display_Channel = QLineEdit()
        QComboBox_Voltage_Sense = QComboBox()
        QComboBox_set_Function = QComboBox()
        QLineEdit_Power_Rating = QLineEdit()
        QLineEdit_maxVoltage = QLineEdit()
        QLineEdit_voltage_rated = QLineEdit()
        QLineEdit_current_rated = QLineEdit()
        QLineEdit_maxCurrent = QLineEdit()
        
        QLineEdit_V_Settling_Band = QLineEdit()
        QLineEdit_T_Settling_Band = QLineEdit()
        QComboBox_Probe_Setting = QComboBox()
        QComboBox_Acq_Type = QComboBox()

        QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        QComboBox_set_Function.setEnabled(False)
        QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])

        QComboBox_Probe_Setting.addItems(["X1", "X10", "X20", "X100"])
        QComboBox_Acq_Type.addItems(["NORMal", "PEAK", "AVERage", "HRESolution"])

        # Oscilloscope Settings
        QLabel_Channel_CouplingMode = QLabel()
        QLabel_Trigger_Mode = QLabel()
        QLabel_Trigger_CouplingMode = QLabel()
        QLabel_Trigger_SweepMode = QLabel()
        QLabel_Trigger_SlopeMode = QLabel()
        QLabel_TimeScale = QLabel()
        QLabel_VerticalScale = QLabel()

        QLabel_Channel_CouplingMode.setText("Coupling Mode (Channel)")
        QLabel_Trigger_Mode.setText("Trigger Mode:")
        QLabel_Trigger_CouplingMode.setText("Coupling Mode (Trigger):")
        QLabel_Trigger_SweepMode.setText("Trigger Sweep Mode:")
        QLabel_Trigger_SlopeMode.setText("Trigger Slope Mode:")
        QLabel_TimeScale.setText("Time Scale:")
        QLabel_VerticalScale.setText("Vertical Scale:")

        QComboBox_Channel_CouplingMode = QComboBox()
        QComboBox_Trigger_Mode = QComboBox()
        QComboBox_Trigger_CouplingMode = QComboBox()
        QComboBox_Trigger_SweepMode = QComboBox()
        QComboBox_Trigger_SlopeMode = QComboBox()
        QLineEdit_TimeScale = QLineEdit()
        QLineEdit_VerticalScale = QLineEdit()

        QComboBox_Channel_CouplingMode.addItems(["AC", "DC"])
        QComboBox_Trigger_Mode.addItems(["EDGE", "IIC", "EBUR"])
        QComboBox_Trigger_CouplingMode.addItems(["AC", "DC"])
        QComboBox_Trigger_SweepMode.addItems(["NORMAL", "AUTO"])
        QComboBox_Trigger_SlopeMode.addItems(["ALT", "POS", "NEG", "EITH"])

        QComboBox_Channel_CouplingMode.setEnabled(True)
        QComboBox_Trigger_Mode.setEnabled(True)
        QComboBox_Trigger_CouplingMode.setEnabled(True)
        QComboBox_Trigger_SweepMode.setEnabled(True)
        QComboBox_Trigger_SlopeMode.setEnabled(True)
        QComboBox_Probe_Setting.setEnabled(True)
        QComboBox_Acq_Type.setEnabled(True)

        # Create a horizontal layout for the "Save Path" and checkboxes
        performtest_layout = QHBoxLayout()
        performtest_layout.addWidget(Desp5)  # QLabel for "Save Path"
        performtest_layout.addWidget(QCheckBox_SpecialCase_Widget)  # Checkbox for "Generate Excel Report"
        performtest_layout.addWidget(QCheckBox_NormalCase_Widget)  # Checkbox for "Show Graph"

        layout1.addRow(Desp0)
        layout1.addRow(QPushButton_Widget00)
        layout1.addRow(self.OutputBox)
        layout1.addRow(QPushButton_Widget0)
        layout1.addRow(Desp1)
        layout1.addRow(self.QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        layout1.addRow(self.QLabel_OSC_VisaAddress, self.QLineEdit_OSC_VisaAddress)
        layout1.addRow(self.QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        layout1.addRow(Desp2)
        """layout1.addRow(QLabel_ELoad_Display_Channel, QLineEdit_ELoad_Display_Channel)
        layout1.addRow(QLabel_PSU_Display_Channel, QLineEdit_PSU_Display_Channel)"""
        layout1.addRow(QLabel_OSC_Display_Channel, QLineEdit_OSC_Display_Channel)
        layout1.addRow(QLabel_set_Function, QComboBox_set_Function)
        layout1.addRow(QLabel_Voltage_Sense, QComboBox_Voltage_Sense)
        layout1.addRow(Desp3)
        layout1.addRow(QLabel_Power_Rating, QLineEdit_Power_Rating)
        layout1.addRow(QLabel_voltage_rated, QLineEdit_voltage_rated)
        layout1.addRow(QLabel_current_rated, QLineEdit_current_rated)
        layout1.addRow(QLabel_maxVoltage, QLineEdit_maxVoltage)
        layout1.addRow(QLabel_maxCurrent, QLineEdit_maxCurrent)
        
        layout1.addRow(QLabel_V_Settling_Band, QLineEdit_V_Settling_Band)
        layout1.addRow(QLabel_T_Settling_Band, QLineEdit_T_Settling_Band)
        layout1.addRow(Desp4)
        layout1.addRow(QLabel_Probe_Setting, QComboBox_Probe_Setting)
        layout1.addRow(QLabel_Acq_Type, QComboBox_Acq_Type)
        layout1.addRow(QLabel_Channel_CouplingMode, QComboBox_Channel_CouplingMode)
        layout1.addRow(QLabel_Trigger_CouplingMode, QComboBox_Trigger_CouplingMode)
        layout1.addRow(QLabel_Trigger_Mode, QComboBox_Trigger_Mode)
        layout1.addRow(QLabel_Trigger_SweepMode, QComboBox_Trigger_SweepMode)
        layout1.addRow(QLabel_Trigger_SlopeMode, QComboBox_Trigger_SlopeMode)
        layout1.addRow(QLabel_TimeScale, QLineEdit_TimeScale)
        layout1.addRow(QLabel_VerticalScale, QLineEdit_VerticalScale)
        layout1.addRow(performtest_layout)
        layout1.addRow(QPushButton_Widget)
        self.setLayout(layout1)

        
        QPushButton_Widget0.clicked.connect(self.doFind)
        QPushButton_Widget.clicked.connect(self.executeTest)
        

        QLineEdit_V_Settling_Band.textEdited.connect(self.V_Settling_Band_changed)
        QLineEdit_T_Settling_Band.textEdited.connect(self.T_Settling_Band_changed)

        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_OSC_VisaAddress.currentTextChanged.connect(self.OSC_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)
        """QLineEdit_ELoad_Display_Channel.textEdited.connect(self.ELoad_Channel_changed)
        QLineEdit_PSU_Display_Channel.textEdited.connect(self.PSU_Channel_changed)"""
        QLineEdit_OSC_Display_Channel.textEdited.connect(self.OSC_Channel_changed)

        QLineEdit_Power_Rating.textEdited.connect(self.Power_Rating_changed)
        QLineEdit_voltage_rated.textEdited.connect(self.Voltage_Rating_changed)
        QLineEdit_current_rated.textEdited.connect(self.Current_Rating_changed)

        QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)
        QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        QComboBox_Voltage_Sense.currentTextChanged.connect(
            self.set_VoltageSense_changed
        )
        QComboBox_Channel_CouplingMode.currentTextChanged.connect(
            self.Channel_CouplingMode_changed
        )
        QComboBox_Trigger_CouplingMode.currentTextChanged.connect(
            self.Trigger_CouplingMode_changed
        )
        QComboBox_Trigger_Mode.currentTextChanged.connect(self.Trigger_Mode_changed)
        QComboBox_Trigger_SweepMode.currentTextChanged.connect(
            self.Trigger_SweepMode_changed
        )
        QComboBox_Trigger_SlopeMode.currentTextChanged.connect(
            self.Trigger_SlopeMode_changed
        )
        QComboBox_Probe_Setting.currentTextChanged.connect(
            self.Probe_Setting_changed
        )
        QComboBox_Acq_Type.currentTextChanged.connect(
            self.Acq_Type_changed
        )
        QLineEdit_TimeScale.textEdited.connect(self.TimeScale_changed)
        QLineEdit_VerticalScale.textEdited.connect(self.VerticalScale_changed)

        QCheckBox_SpecialCase_Widget.stateChanged.connect(self.checkbox_state_SpecialCase)
        QCheckBox_NormalCase_Widget.stateChanged.connect(self.checkbox_state_NormalCase)

        QPushButton_Widget00.clicked.connect(self.savepath)

    def Voltage_Rating_changed(self, value):
        self.Voltage_Rating = value
    
    def Current_Rating_changed(self, value):
        self.Current_Rating = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_OSC_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = GetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_OSC_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])
            
            # Add "None" option at the end
            self.QLineEdit_PSU_VisaAddress.addItem("None")
            self.QLineEdit_OSC_VisaAddress.addItem("None")
            self.QLineEdit_ELoad_VisaAddress.addItem("None")

        except:
            self.OutputBox.append("No Devices Found!!!")
        return 
    
    def checkbox_state_SpecialCase(self, s):
        self.checkbox_SpecialCase = s

    def checkbox_state_NormalCase(self, s):
        self.checkbox_NormalCase = s
    
    def PSU_VisaAddress_changed(self, s):
        self.PSU = s

    def OSC_VisaAddress_changed(self, s):
        self.OSC = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s
    
    def ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s

    def PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def OSC_Channel_changed(self, s):
        self.OSC_Channel = s

    def Power_Rating_changed(self, s):
        self.Power_Rating = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def I_Step_changed(self, s):
        self.I_Step = s

    def T_Settling_Band_changed(self, s):
        self.T_Settling_Band = s

    def V_Settling_Band_changed(self, s):
        self.V_Settling_Band = s

    def Channel_CouplingMode_changed(self, s):
        self.Channel_CouplingMode = s

    def Trigger_CouplingMode_changed(self, s):
        self.Trigger_CouplingMode = s

    def Trigger_Mode_changed(self, s):
        self.Trigger_Mode = s

    def Trigger_SweepMode_changed(self, s):
        self.Trigger_SweepMode = s

    def Trigger_SlopeMode_changed(self, s):
        self.Trigger_SlopeMode = s
    
    def Probe_Setting_changed(self, s):
        self.Probe_Setting = s
    
    def Acq_Type_changed(self, s):
        self.Acq_Type = s

    def TimeScale_changed(self, s):
        self.TimeScale = s

    def VerticalScale_changed(self, s):
        self.VerticalScale = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))


    def executeTest(self):
        """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
        then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
        will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
        begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
        begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
        are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
        optionally export all the details into a CSV file or display a graph after the test is completed.

        """
        dict = dictGenerator.input(
            savedir=self.savelocation,
            Instrument="Keysight",
            PSU=self.PSU,
            OSC=self.OSC,
            ELoad=self.ELoad,
            V_Rating=self.Voltage_Rating,
            I_Rating=self.Current_Rating,
            P_Rating=self.Power_Rating,
            power = self.Power_Rating,
            maxCurrent=self.maxCurrent,
            maxVoltage=self.maxVoltage,
            ELoad_Channel=self.ELoad_Channel,
            PSU_Channel=self.PSU_Channel,
            OSC_Channel=self.OSC_Channel,
            VoltageSense=self.VoltageSense,
            setFunction=self.setFunction,
            Channel_CouplingMode=self.Channel_CouplingMode,
            Trigger_Mode=self.Trigger_Mode,
            Trigger_CouplingMode=self.Trigger_CouplingMode,
            Trigger_SweepMode=self.Trigger_SweepMode,
            Trigger_SlopeMode=self.Trigger_SlopeMode,
            Probe_Setting=self.Probe_Setting,
            Acq_Type=self.Acq_Type,
            TimeScale=self.TimeScale,
            VerticalScale=self.VerticalScale,
            
            V_Settling_Band=self.V_Settling_Band,
            T_Settling_Band=self.T_Settling_Band,
        )
        QMessageBox.warning(
            self,
            "In Progress",
            "Measurement will start now , please do not close the main window until test is completed",
        )

        #Execute Voltage Measurement
        relay_voltage = RelayController_Voltage()
        if self.relay_voltage == "RELAY":
            relay_voltage.relay_on()
        else:
            relay_voltage.relay_off()

        
        if self.selected_text == "Dolphin":
            if self.ELoad != "None" and self.OSC != "None":
                print_console_safe("ELoad connected and DMM connected") #All connected
                try:
                    if self.checkbox_SpecialCase == 2:
                        DolphinRiseFallTimewithELoad.executeA(self, dict)
                    
                    if self.checkbox_NormalCase == 2:
                        DolphinRiseFallTimewithELoad.executeB(self, dict)

                
                except Exception as e:
                    print_console_safe(e)
                    QMessageBox.warning(self, "Error", str(e))
                    exit()

        self.OutputBox.append(my_result.getvalue())
        self.OutputBox.append("Measurement is complete !")

class TransientRecoveryTimeWithCurrentSensor(QDialog):

    """Class for configuring the Transient Recovery Time DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.

    """

    def __init__(self):
        """ "Method declaring the Widgets, Signals & Slots for Transient Recovery Time."""
        super().__init__()

        # Default Values
        self.PSU = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        self.OSC = "USB0::0x0957::0x17B0::MY52060151::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        self.ELoad_Channel = "1"
        self.PSU_Channel = "2"
        self.DUT_OSC_Channel = "CHANNEL1"
        self.CurrentTrigger_OSC_Channel = "CHANNEL2"
        self.TriggerSource = "CHANNEL2"
        self.setFunction = "Current"

        self.VoltageSense = "EXT"
        self.Power_Rating = "160"
        self.Current_Rating = "120"
        self.Voltage_Rating = "80"
        self.maxCurrent = "10"
        self.maxVoltage = "80"

        self.relay_voltage = "RELAY"
        
        #From OSC view
        self.DUT_Channel_Unit = "VOLT"
        self.DUT_Channel_Offset = "0"
    
        self.DUT_Channel_CouplingMode = "AC"
        self.DUT_Trigger_CouplingMode = "DC"
        self.DUT_Trigger_Mode = "EDGE"
        self.DUT_Trigger_SweepMode = "NORMAL"
        self.DUT_Trigger_SlopeMode = "EITHer"
        self.DUT_TimeScale = "0.01"
        self.DUT_VerticalScale = "0.00001"
        self.DUT_V_Settling_Band = "0.8"
        self.DUT_T_Settling_Band = "0.001"
        self.DUT_Probe_Setting = "X10"
        self.DUT_Acq_Type = "AVERage"
        self.DUT_Unit = "VOLT"


        self.CurrentTrigger_Channel_Unit = "AMPere"
        self.CurrentTrigger_Channel_Offset = "20"
        self.CurrentTrigger_Channel_CouplingMode = "DC"
        self.CurrentTrigger_Trigger_CouplingMode = "DC"
        self.CurrentTrigger_Trigger_Mode = "EDGE"
        self.CurrentTrigger_Trigger_SweepMode = "NORMAL"
        self.CurrentTrigger_Trigger_SlopeMode = "EITHer"
        self.CurrentTrigger_TimeScale = "0.01"
        self.CurrentTrigger_VerticalScale = "20"
        self.CurrentTrigger_V_Settling_Band = "8"
        self.CurrentTrigger_T_Settling_Band = "0.001"
        self.CurrentTrigger_Probe_Setting = "X100"
        self.CurrentTrigger_Acq_Type = "AVERage"
        self.CurrentTrigger_Unit = "AMPere"


        self.I_Step = ""
        

        self.checkbox_SpecialCase = 0
        self.checkbox_NormalCase = 0
        self.checkbox_CurrentCase = 2

        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"

        self.setWindowTitle("Transient Recovery Time with Current Probe")

        QPushButton_Widget00 = QPushButton()
        QPushButton_Widget00.setText("Save Path")
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Find Instruments")
        QPushButton_Widget = QPushButton()
        QPushButton_Widget.setText("Execute Test")
        QCheckBox_SpecialCase_Widget = QCheckBox()
        QCheckBox_SpecialCase_Widget.setText("Special Case (0% <-> 100%)")
        QCheckBox_SpecialCase_Widget.setCheckState(Qt.Unchecked)
        QCheckBox_NormalCase_Widget = QCheckBox()
        QCheckBox_NormalCase_Widget.setText("Normal Case without Current Probe (50% <-> 100%)")
        QCheckBox_NormalCase_Widget.setCheckState(Qt.Unchecked)
        QCheckBox_CurrentCase_Widget = QCheckBox()
        QCheckBox_CurrentCase_Widget.setText("Normal Case with Current Probe (50% <-> 100%)")
        QCheckBox_CurrentCase_Widget.setCheckState(Qt.Checked)

        
        layout1 = QFormLayout()
        self.OutputBox = QTextBrowser()

        self.OutputBox.append(my_result.getvalue())

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()
        Desp6 = QLabel()
        Desp8 = QLabel()

        Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)
        Desp6.setFont(desp_font)
        Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Specification:")
        Desp4.setText("DUT Oscilloscope Settings:")
        Desp5.setText("Perform Test:")
        Desp6.setText("Current Trigger Oscilloscope Settings:")
        Desp8.setText("External Auxiliary Equipment:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        # Connections
        QLabel_PSU_VisaAddress = QLabel()
        QLabel_OSC_VisaAddress = QLabel()
        QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DUT = QLabel()

        QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        QLabel_OSC_VisaAddress.setText("Visa Address (OSC):")
        QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DUT.setText("DUT:")

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_OSC_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        self.QComboBox_DUT = QComboBox()

        # General Settings
        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_OSC_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QComboBox_DUT.addItems(["Others", "Excavator", "EDU36311A", "Dolphin", "Hornbill", "SMU"])

        QLabel_OSC_DUT_Display_Channel = QLabel() #DUT Source 
        QLabel_OSC_CurrentTrigger_Display_Channel = QLabel() #Current Trigger
        QLabel_Trigger_Source_Display_Channel = QLabel()

        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Power_Rating = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_rated = QLabel()
        QLabel_current_rated = QLabel()
        QLabel_maxCurrent = QLabel()

        QLabel_DUT_V_Settling_Band = QLabel()
        QLabel_DUT_T_Settling_Band = QLabel()
        QLabel_DUT_Probe_Setting = QLabel()
        QLabel_DUT_Acq_Type = QLabel()

        QLabel_CurrentTrigger_V_Settling_Band = QLabel()
        QLabel_CurrentTrigger_T_Settling_Band = QLabel()
        QLabel_CurrentTrigger_Probe_Setting = QLabel()
        QLabel_CurrentTrigger_Acq_Type = QLabel()

        QLabel_OSC_DUT_Display_Channel.setText("DUT Display Channel (OSC)")
        QLabel_OSC_CurrentTrigger_Display_Channel.setText("Current Trigger Display Channel (OSC)")
        QLabel_Trigger_Source_Display_Channel.setText("Trigger Source Channel (OSC)")
        
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Power_Rating.setText("Power Rating (W):")
        QLabel_maxVoltage.setText("Testing Voltage (V):")
        QLabel_voltage_rated.setText("DUT Rated Voltage:")
        QLabel_current_rated.setText("DUT Rated Current:")
        QLabel_maxCurrent.setText("Testing Current (A):")
        
        QLabel_DUT_V_Settling_Band.setText("DUT Settling Band Voltage (V) / Error Band:")
        QLabel_DUT_T_Settling_Band.setText("DUT Settling Band Time (s):")
        QLabel_DUT_Probe_Setting.setText("DUT Probe Setting Ratio:")
        QLabel_DUT_Acq_Type.setText("DUT Acquire Mode:")

        QLabel_CurrentTrigger_V_Settling_Band.setText("CurrentTrigger Settling Band Voltage (V) / Error Band:")
        QLabel_CurrentTrigger_T_Settling_Band.setText("CurrentTrigger Settling Band Time (s):")
        QLabel_CurrentTrigger_Probe_Setting.setText("CurrentTrigger Probe Setting Ratio:")
        QLabel_CurrentTrigger_Acq_Type.setText("CurrentTrigger Acquire Mode:")

        self.QLineEdit_OSC_DUT_Display_Channel = QComboBox()
        self.QLineEdit_OSC_CurrentTrigger_Display_Channel = QComboBox()
        self.QComboBox_Trigger_Source_Display_Channel = QComboBox()

        self.QLineEdit_OSC_DUT_Display_Channel.addItems(["CHANNEL1", "CHANNEL2", "CHANNEL3", "CHANNEL4"])
        self.QLineEdit_OSC_CurrentTrigger_Display_Channel.addItems(["CHANNEL1", "CHANNEL2", "CHANNEL3", "CHANNEL4"])
        self.QComboBox_Trigger_Source_Display_Channel.addItems(["CHANNEL1", "CHANNEL2", "CHANNEL3", "CHANNEL4"])

        # External Auxiliary Equipment section
        QLabel_External_Auxiliary_Equipment = QLabel()
        QLabel_External_Auxiliary_Equipment.setText("Relay")
        self.QComboBox_External_Auxiliary_Equipment = QComboBox()
        self.QComboBox_External_Auxiliary_Equipment.addItems(["None", "RELAY"])

        self.QComboBox_set_PSU_Channel = QComboBox()
        self.QComboBox_set_ELoad_Channel = QComboBox()
        self.QComboBox_Voltage_Sense = QComboBox()
        self.QComboBox_set_Function = QComboBox()
        self.QLineEdit_Power_Rating = QLineEdit()
        self.QLineEdit_maxVoltage = QLineEdit()
        self.QLineEdit_voltage_rated = QLineEdit()
        self.QLineEdit_current_rated = QLineEdit()
        self.QLineEdit_maxCurrent = QLineEdit()

        self.QLineEdit_DUT_V_Settling_Band = QLineEdit()
        self.QLineEdit_DUT_T_Settling_Band = QLineEdit()
        self.QComboBox_DUT_Probe_Setting = QComboBox()
        self.QComboBox_DUT_Acq_Type = QComboBox()

        self.QLineEdit_CurrentTrigger_V_Settling_Band = QLineEdit()
        self.QLineEdit_CurrentTrigger_T_Settling_Band = QLineEdit()
        self.QComboBox_CurrentTrigger_Probe_Setting = QComboBox()
        self.QComboBox_CurrentTrigger_Acq_Type = QComboBox()

        self.QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        self.QComboBox_set_Function.setEnabled(False)
        self.QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])

        self.QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        self.QComboBox_set_PSU_Channel.setEnabled(True)
        self.QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        self.QComboBox_set_ELoad_Channel.setEnabled(True)

        self.QComboBox_DUT_Probe_Setting.addItems(["X1", "X10", "X20", "X100"])
        self.QComboBox_DUT_Acq_Type.addItems(["NORMal", "PEAK", "AVERage", "HRESolution"])

        self.QComboBox_CurrentTrigger_Probe_Setting.addItems(["X1", "X10", "X20", "X100"])
        self.QComboBox_CurrentTrigger_Acq_Type.addItems(["NORMal", "PEAK", "AVERage", "HRESolution"])

        # DUT Oscilloscope Settings
        QLabel_DUT_Channel_Unit = QLabel()
        QLabel_DUT_Channel_Offset = QLabel()
        QLabel_DUT_Channel_CouplingMode = QLabel()
        QLabel_DUT_Trigger_Mode = QLabel()
        QLabel_DUT_Trigger_CouplingMode = QLabel()
        QLabel_DUT_Trigger_SweepMode = QLabel()
        QLabel_DUT_Trigger_SlopeMode = QLabel()
        QLabel_DUT_TimeScale = QLabel()
        QLabel_DUT_VerticalScale = QLabel()

        QLabel_DUT_Channel_Unit.setText("Unit (DUT Channel)")
        QLabel_DUT_Channel_Offset.setText("Offset (DUT Channel)")
        QLabel_DUT_Channel_CouplingMode.setText("Coupling Mode (Channel)")
        QLabel_DUT_Trigger_Mode.setText("Trigger Mode:")
        QLabel_DUT_Trigger_CouplingMode.setText("Coupling Mode (Trigger):")
        QLabel_DUT_Trigger_SweepMode.setText("Trigger Sweep Mode:")
        QLabel_DUT_Trigger_SlopeMode.setText("Trigger Slope Mode:")
        QLabel_DUT_TimeScale.setText("Time Scale:")
        QLabel_DUT_VerticalScale.setText("Vertical Scale:(DUT(V))")
        
        self.QComboBox_DUT_Channel_Unit = QComboBox()
        self.QLineEdit_DUT_Channel_Offset = QLineEdit()
        self.QComboBox_DUT_Channel_CouplingMode = QComboBox()
        self.QComboBox_DUT_Trigger_Mode = QComboBox()
        self.QComboBox_DUT_Trigger_CouplingMode = QComboBox()
        self.QComboBox_DUT_Trigger_SweepMode = QComboBox()
        self.QComboBox_DUT_Trigger_SlopeMode = QComboBox()
        self.QLineEdit_DUT_TimeScale = QLineEdit()
        self.QLineEdit_DUT_VerticalScale = QLineEdit()

        self.QComboBox_DUT_Channel_Unit.addItems(["VOLT", "AMPere"])
        self.QComboBox_DUT_Channel_CouplingMode.addItems(["AC", "DC"])
        self.QComboBox_DUT_Trigger_Mode.addItems(["EDGE", "IIC", "EBUR"])
        self.QComboBox_DUT_Trigger_CouplingMode.addItems(["AC", "DC"])
        self.QComboBox_DUT_Trigger_SweepMode.addItems(["NORMAL", "AUTO"])
        self.QComboBox_DUT_Trigger_SlopeMode.addItems(["ALT", "POS", "NEG", "EITH"])

        self.QComboBox_DUT_Channel_Unit.setEnabled(True)
        self.QComboBox_DUT_Channel_CouplingMode.setEnabled(True)
        self.QComboBox_DUT_Trigger_Mode.setEnabled(False)
        self.QComboBox_DUT_Trigger_CouplingMode.setEnabled(False)
        self.QComboBox_DUT_Trigger_SweepMode.setEnabled(False)
        self.QComboBox_DUT_Trigger_SlopeMode.setEnabled(False)
        self.QComboBox_DUT_Probe_Setting.setEnabled(True)
        self.QComboBox_DUT_Acq_Type.setEnabled(False)

         # Current Trigger Oscilloscope Settings
        QLabel_CurrentTrigger_Channel_Unit = QLabel()
        QLabel_CurrentTrigger_Channel_Offset = QLabel()
        QLabel_CurrentTrigger_Channel_CouplingMode = QLabel()
        QLabel_CurrentTrigger_Trigger_Mode = QLabel()
        QLabel_CurrentTrigger_Trigger_CouplingMode = QLabel()
        QLabel_CurrentTrigger_Trigger_SweepMode = QLabel()
        QLabel_CurrentTrigger_Trigger_SlopeMode = QLabel()
        QLabel_CurrentTrigger_TimeScale = QLabel()
        QLabel_CurrentTrigger_VerticalScale = QLabel()

        QLabel_CurrentTrigger_Channel_Unit.setText("Unit (Current Trigger Channel)")
        QLabel_CurrentTrigger_Channel_Offset.setText("Offset (Current Trigger Channel)")
        QLabel_CurrentTrigger_Channel_CouplingMode.setText("Coupling Mode (Channel)")
        QLabel_CurrentTrigger_Trigger_Mode.setText("Trigger Mode:")
        QLabel_CurrentTrigger_Trigger_CouplingMode.setText("Coupling Mode (Trigger):")
        QLabel_CurrentTrigger_Trigger_SweepMode.setText("Trigger Sweep Mode:")
        QLabel_CurrentTrigger_Trigger_SlopeMode.setText("Trigger Slope Mode:")
        QLabel_CurrentTrigger_TimeScale.setText("Time Scale:")
        QLabel_CurrentTrigger_VerticalScale.setText("Vertical Scale:(Current(A))")

        self.QComboBox_CurrentTrigger_Channel_Unit = QComboBox()
        self.QLineEdit_CurrentTrigger_Channel_Offset = QLineEdit()
        self.QComboBox_CurrentTrigger_Channel_CouplingMode = QComboBox()
        self.QComboBox_CurrentTrigger_Trigger_Mode = QComboBox()
        self.QComboBox_CurrentTrigger_Trigger_CouplingMode = QComboBox()
        self.QComboBox_CurrentTrigger_Trigger_SweepMode = QComboBox()
        self.QComboBox_CurrentTrigger_Trigger_SlopeMode = QComboBox()
        self.QLineEdit_CurrentTrigger_TimeScale = QLineEdit()
        self.QLineEdit_CurrentTrigger_VerticalScale = QLineEdit()

        self.QComboBox_CurrentTrigger_Channel_Unit.addItems([ "AMPere","VOLT"])
        self.QComboBox_CurrentTrigger_Channel_CouplingMode.addItems(["DC","AC"])
        self.QComboBox_CurrentTrigger_Trigger_Mode.addItems(["EDGE", "IIC", "EBUR"])
        self.QComboBox_CurrentTrigger_Trigger_CouplingMode.addItems(["AC", "DC"])
        self.QComboBox_CurrentTrigger_Trigger_SweepMode.addItems(["NORMAL", "AUTO"])
        self.QComboBox_CurrentTrigger_Trigger_SlopeMode.addItems(["ALT", "POS", "NEG", "EITH"])

        self.QComboBox_CurrentTrigger_Channel_CouplingMode.setEnabled(False)
        self.QComboBox_CurrentTrigger_Trigger_Mode.setEnabled(True)
        self.QComboBox_CurrentTrigger_Trigger_CouplingMode.setEnabled(True)
        self.QComboBox_CurrentTrigger_Trigger_SweepMode.setEnabled(True)
        self.QComboBox_CurrentTrigger_Trigger_SlopeMode.setEnabled(True)
        self.QComboBox_CurrentTrigger_Probe_Setting.setEnabled(True)
        self.QComboBox_CurrentTrigger_Acq_Type.setEnabled(True)


        # Create a horizontal layout for the "Save Path" and checkboxes
        Rated_Power_Voltage_Current_Layout = QHBoxLayout()
        Rated_Power_Voltage_Current_Layout.addWidget(QLabel_Power_Rating)  # QLabel for "Save Path"
        Rated_Power_Voltage_Current_Layout.addWidget(self.QLineEdit_Power_Rating)  # Checkbox for "Generate Excel Report"
        Rated_Power_Voltage_Current_Layout.addWidget(QLabel_voltage_rated)  # Checkbox for "Show Graph"
        Rated_Power_Voltage_Current_Layout.addWidget(self.QLineEdit_voltage_rated)  # Checkbox for "Show Graph"
        Rated_Power_Voltage_Current_Layout.addWidget(QLabel_current_rated)  # QLabel for "Save Path"
        Rated_Power_Voltage_Current_Layout.addWidget(self.QLineEdit_current_rated)  # Checkbox for "Generate Excel Report"

        Test_Voltage_Current_Layout = QHBoxLayout()
        Test_Voltage_Current_Layout.addWidget(QLabel_maxVoltage)  # QLabel for "Save Path"
        Test_Voltage_Current_Layout.addWidget(self.QLineEdit_maxVoltage)  # Checkbox for "Generate Excel Report"
        Test_Voltage_Current_Layout.addWidget(QLabel_maxCurrent)  # Checkbox for "Show Graph"
        Test_Voltage_Current_Layout.addWidget(self.QLineEdit_maxCurrent)  # Checkbox for "Show Graph"

        DUT_Error_Band_layout = QHBoxLayout()
        DUT_Error_Band_layout.addWidget(QLabel_DUT_V_Settling_Band)  # QLabel for "Save Path"
        DUT_Error_Band_layout.addWidget(self.QLineEdit_DUT_V_Settling_Band)  # Checkbox for "Generate Excel Report"
        DUT_Error_Band_layout.addWidget(QLabel_DUT_T_Settling_Band)  # Checkbox for "Show Graph"
        DUT_Error_Band_layout.addWidget(self.QLineEdit_DUT_T_Settling_Band)  # Checkbox for "Show Graph"


        performtest_layout = QHBoxLayout()
        performtest_layout.addWidget(Desp5)  # QLabel for "Save Path"
        performtest_layout.addWidget(QCheckBox_SpecialCase_Widget)  # Checkbox for "Generate Excel Report"
        performtest_layout.addWidget(QCheckBox_NormalCase_Widget)  # Checkbox for "Show Graph"
        performtest_layout.addWidget(QCheckBox_CurrentCase_Widget)  # Checkbox for "Show Graph"

        Channel_CouplingMode_layout = QHBoxLayout()
        Channel_CouplingMode_layout.addWidget(QLabel_DUT_Channel_CouplingMode)  # QLabel for "Save Path"
        Channel_CouplingMode_layout.addWidget(self.QComboBox_DUT_Channel_CouplingMode)  # Checkbox for "Generate Excel Report"
        Channel_CouplingMode_layout.addWidget(QLabel_CurrentTrigger_Channel_CouplingMode)  # Checkbox for "Show Graph"
        Channel_CouplingMode_layout.addWidget(self.QComboBox_CurrentTrigger_Channel_CouplingMode)  # Checkbox for "Show Graph"

        Trigger_Mode_layout = QHBoxLayout()
        Trigger_Mode_layout.addWidget(QLabel_DUT_Trigger_Mode)  # Checkbox for "Show Graph"
        Trigger_Mode_layout.addWidget(self.QComboBox_DUT_Trigger_Mode)  # Checkbox for "Show Graph"
        Trigger_Mode_layout.addWidget(QLabel_CurrentTrigger_Trigger_Mode)  # QLabel for "Save Path"
        Trigger_Mode_layout.addWidget(self.QComboBox_CurrentTrigger_Trigger_Mode)  # Checkbox for "Generate Excel Report"

        Unit_Channel_layout = QHBoxLayout()
        Unit_Channel_layout.addWidget(QLabel_DUT_Channel_Unit)  # QLabel for "Save Path"
        Unit_Channel_layout.addWidget(self.QComboBox_DUT_Channel_Unit)  # Checkbox for "Generate Excel Report"
        Unit_Channel_layout.addWidget(QLabel_CurrentTrigger_Channel_Unit)  # Checkbox for "Show Graph"
        Unit_Channel_layout.addWidget(self.QComboBox_CurrentTrigger_Channel_Unit)  # Checkbox for "Show Graph"

        Offset_Channel_layout = QHBoxLayout()
        Offset_Channel_layout.addWidget(QLabel_DUT_Channel_Offset)  # Checkbox for "Show Graph"
        Offset_Channel_layout.addWidget(self.QLineEdit_DUT_Channel_Offset)  # Checkbox for "Show Graph"
        Offset_Channel_layout.addWidget(QLabel_CurrentTrigger_Channel_Offset)  # QLabel for "Save Path"
        Offset_Channel_layout.addWidget(self.QLineEdit_CurrentTrigger_Channel_Offset)  # Checkbox for "Generate Excel Report"

        Trigger_CouplingMode_layout = QHBoxLayout()
        Trigger_CouplingMode_layout.addWidget(QLabel_DUT_Trigger_CouplingMode)  # QLabel for "Save Path"
        Trigger_CouplingMode_layout.addWidget(self.QComboBox_DUT_Trigger_CouplingMode)  # Checkbox for "Generate Excel Report"
        Trigger_CouplingMode_layout.addWidget(QLabel_CurrentTrigger_Trigger_CouplingMode)  # Checkbox for "Show Graph"
        Trigger_CouplingMode_layout.addWidget(self.QComboBox_CurrentTrigger_Trigger_CouplingMode)  # Checkbox for "Show Graph"

        Trigger_SweepMode_layout = QHBoxLayout()
        Trigger_SweepMode_layout.addWidget(QLabel_DUT_Trigger_SweepMode)  # Checkbox for "Show Graph"  
        Trigger_SweepMode_layout.addWidget(self.QComboBox_DUT_Trigger_SweepMode)  # Checkbox for "Show Graph"
        Trigger_SweepMode_layout.addWidget(QLabel_CurrentTrigger_Trigger_SweepMode)  # QLabel for "Save Path"
        Trigger_SweepMode_layout.addWidget(self.QComboBox_CurrentTrigger_Trigger_SweepMode)  # Checkbox for "Generate Excel Report"

        Trigger_SlopeMode_layout = QHBoxLayout()
        Trigger_SlopeMode_layout.addWidget(QLabel_DUT_Trigger_SlopeMode)  # QLabel for "Save Path"
        Trigger_SlopeMode_layout.addWidget(self.QComboBox_DUT_Trigger_SlopeMode)  # Checkbox for "Generate Excel Report"
        Trigger_SlopeMode_layout.addWidget(QLabel_CurrentTrigger_Trigger_SlopeMode)  # Checkbox for "Show Graph"
        Trigger_SlopeMode_layout.addWidget(self.QComboBox_CurrentTrigger_Trigger_SlopeMode)  # Checkbox for "Show Graph"

        Probe_Setting_layout = QHBoxLayout()
        Probe_Setting_layout.addWidget(QLabel_DUT_Probe_Setting)  # QLabel for "Save Path"
        Probe_Setting_layout.addWidget(self.QComboBox_DUT_Probe_Setting)  # Checkbox for "Generate Excel Report"
        Probe_Setting_layout.addWidget(QLabel_CurrentTrigger_Probe_Setting)  # Checkbox for "Show Graph"
        Probe_Setting_layout.addWidget(self.QComboBox_CurrentTrigger_Probe_Setting)  # Checkbox for "Show Graph"

        TimeScale_layout = QHBoxLayout()
        TimeScale_layout.addWidget(QLabel_DUT_TimeScale)  # Checkbox for "Show Graph"
        TimeScale_layout.addWidget(self.QLineEdit_DUT_TimeScale)  # Checkbox for "Show Graph"
        #TimeScale_layout.addWidget(QLabel_CurrentTrigger_TimeScale)  # QLabel for "Save Path"
        #TimeScale_layout.addWidget(QLineEdit_CurrentTrigger_TimeScale)  # Checkbox for "Generate Excel Report"

        VerticalScale_layout = QHBoxLayout()
        VerticalScale_layout.addWidget(QLabel_DUT_VerticalScale)  # QLabel for "Save Path"
        VerticalScale_layout.addWidget(self.QLineEdit_DUT_VerticalScale)  # Checkbox for "Generate Excel Report"
        #VerticalScale_layout.addWidget(QLabel_CurrentTrigger_VerticalScale)  # Checkbox for "Show Graph"
        #VerticalScale_layout.addWidget(QLineEdit_CurrentTrigger_VerticalScale)  # Checkbox for "Show Graph"


        layout1.addRow(Desp0)
        layout1.addRow(QPushButton_Widget00)
        layout1.addRow(self.OutputBox)
        layout1.addRow(QPushButton_Widget0)
        layout1.addRow(Desp1)
        layout1.addRow(QLabel_DUT, self.QComboBox_DUT)
        layout1.addRow(QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        layout1.addRow(QLabel_OSC_VisaAddress, self.QLineEdit_OSC_VisaAddress)
        layout1.addRow(QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_set_Function, self.QComboBox_set_Function)
        layout1.addRow(QLabel_Voltage_Sense, self.QComboBox_Voltage_Sense)
        layout1.addRow(QLabel_set_PSU_Channel, self.QComboBox_set_PSU_Channel)
        layout1.addRow(QLabel_set_ELoad_Channel, self.QComboBox_set_ELoad_Channel)
        layout1.addRow(QLabel_OSC_DUT_Display_Channel, self.QLineEdit_OSC_DUT_Display_Channel)
        layout1.addRow(QLabel_OSC_CurrentTrigger_Display_Channel, self.QLineEdit_OSC_CurrentTrigger_Display_Channel)
        layout1.addRow(QLabel_Trigger_Source_Display_Channel, self.QComboBox_Trigger_Source_Display_Channel)
        layout1.addRow(Desp8)
        layout1.addRow(QLabel_External_Auxiliary_Equipment, self.QComboBox_External_Auxiliary_Equipment)
        layout1.addRow(Desp3)
        layout1.addRow(Rated_Power_Voltage_Current_Layout)
        layout1.addRow(Test_Voltage_Current_Layout)
        layout1.addRow(DUT_Error_Band_layout)
        layout1.addRow(QLabel_CurrentTrigger_V_Settling_Band, self.QLineEdit_CurrentTrigger_V_Settling_Band)

  
        layout1.addRow(Desp4)#DUT Oscilloscope Settings

        layout1.addRow(Unit_Channel_layout)
        layout1.addRow(Offset_Channel_layout)
        layout1.addRow(Channel_CouplingMode_layout)
        layout1.addRow(Trigger_Mode_layout)
        layout1.addRow(Trigger_CouplingMode_layout)
        layout1.addRow(Trigger_SweepMode_layout)
        layout1.addRow(Trigger_SlopeMode_layout)
        layout1.addRow(Probe_Setting_layout)
        layout1.addRow(TimeScale_layout)
        layout1.addRow(VerticalScale_layout)
        layout1.addRow(QLabel_CurrentTrigger_VerticalScale, self.QLineEdit_CurrentTrigger_VerticalScale)
        
        

        layout1.addRow(performtest_layout)
        layout1.addRow(QPushButton_Widget)
        self.setLayout(layout1)

        
        QPushButton_Widget0.clicked.connect(self.doFind)
        QPushButton_Widget.clicked.connect(self.executeTest)
    
        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_OSC_VisaAddress.currentTextChanged.connect(self.OSC_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)
        self.QComboBox_DUT.currentTextChanged.connect(self.on_current_index_changed)
        self.QComboBox_set_PSU_Channel.currentTextChanged.connect(self.PSU_Channel_changed)
        self.QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        self.QComboBox_External_Auxiliary_Equipment.currentTextChanged.connect(self.External_Auxiliary_Equipment_changed)
        """QLineEdit_ELoad_Display_Channel.textEdited.connect(self.ELoad_Channel_changed)
        QLineEdit_PSU_Display_Channel.textEdited.connect(self.PSU_Channel_changed)"""

        self.QLineEdit_OSC_DUT_Display_Channel.currentTextChanged.connect(self.DUT_OSC_Channel_changed)
        self.QLineEdit_OSC_CurrentTrigger_Display_Channel.currentTextChanged.connect(self.CurrentTrigger_OSC_Channel_changed)
        self.QComboBox_Trigger_Source_Display_Channel.currentTextChanged.connect(self.TriggerSource_changed)
        
        self.QLineEdit_Power_Rating.textEdited.connect(self.Power_Rating_changed)
        self.QLineEdit_voltage_rated.textEdited.connect(self.Voltage_Rating_changed)
        self.QLineEdit_current_rated.textEdited.connect(self.Current_Rating_changed)

        self.QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        self.QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)

        self.QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)

        self.QComboBox_Voltage_Sense.currentTextChanged.connect(
            self.set_VoltageSense_changed
        )
        #DUT Oscilloscope Settings
        self.QLineEdit_DUT_V_Settling_Band.textEdited.connect(self.DUT_V_Settling_Band_changed)
        self.QLineEdit_DUT_T_Settling_Band.textEdited.connect(self.DUT_T_Settling_Band_changed)
        self.QLineEdit_CurrentTrigger_V_Settling_Band.textEdited.connect(self.CurrentTrigger_V_Settling_Band_changed)
        #self.QLineEdit_CurrentTrigger_T_Settling_Band.textEdited.connect(self.T_Settling_Band_changed)

        self.QComboBox_DUT_Channel_Unit.currentTextChanged.connect(
            self.DUT_Channel_Unit_changed
        )

        self.QLineEdit_DUT_Channel_Offset.textEdited.connect(self.DUT_Channel_Offset_changed)

        self.QComboBox_DUT_Channel_CouplingMode.currentTextChanged.connect(
            self.DUT_Channel_CouplingMode_changed
        )
        self.QComboBox_DUT_Trigger_CouplingMode.currentTextChanged.connect(
            self.DUT_Trigger_CouplingMode_changed
        )
        self.QComboBox_DUT_Trigger_Mode.currentTextChanged.connect(self.DUT_Trigger_Mode_changed)
        self.QComboBox_DUT_Trigger_SweepMode.currentTextChanged.connect(
            self.DUT_Trigger_SweepMode_changed
        )
        self.QComboBox_DUT_Trigger_SlopeMode.currentTextChanged.connect(
            self.DUT_Trigger_SlopeMode_changed
        )
        self.QComboBox_DUT_Probe_Setting.currentTextChanged.connect(
            self.DUT_Probe_Setting_changed
        )
        self.QComboBox_DUT_Acq_Type.currentTextChanged.connect(
            self.DUT_Acq_Type_changed
        )
        self.QLineEdit_DUT_TimeScale.textEdited.connect(self.DUT_TimeScale_changed)

        self.QLineEdit_DUT_VerticalScale.textEdited.connect(self.DUT_VerticalScale_changed)

        ###Current Trigger
        self.QComboBox_CurrentTrigger_Channel_Unit.currentTextChanged.connect(
            self.CurrentTrigger_Channel_Unit_changed
        )

        self.QLineEdit_CurrentTrigger_Channel_Offset.textEdited.connect(self.CurrentTrigger_Channel_Offset_changed)

        self.QComboBox_CurrentTrigger_Channel_CouplingMode.currentTextChanged.connect(
            self.CurrentTrigger_Channel_CouplingMode_changed
        )
        self.QComboBox_CurrentTrigger_Trigger_CouplingMode.currentTextChanged.connect(
            self.CurrentTrigger_Trigger_CouplingMode_changed
        )
        self.QComboBox_CurrentTrigger_Trigger_Mode.currentTextChanged.connect(self.CurrentTrigger_Trigger_Mode_changed)
        self.QComboBox_CurrentTrigger_Trigger_SweepMode.currentTextChanged.connect(
            self.CurrentTrigger_Trigger_SweepMode_changed
        )
        self.QComboBox_CurrentTrigger_Trigger_SlopeMode.currentTextChanged.connect(
            self.CurrentTrigger_Trigger_SlopeMode_changed
        )   
        self.QComboBox_CurrentTrigger_Probe_Setting.currentTextChanged.connect(
            self.CurrentTrigger_Probe_Setting_changed
        )   
        self.QComboBox_CurrentTrigger_Acq_Type.currentTextChanged.connect(
            self.CurrentTrigger_Acq_Type_changed
        ) 
        

        self.QLineEdit_CurrentTrigger_VerticalScale.textEdited.connect(self.CurrentTrigger_VerticalScale_changed)

        QCheckBox_SpecialCase_Widget.stateChanged.connect(self.checkbox_state_SpecialCase)
        QCheckBox_NormalCase_Widget.stateChanged.connect(self.checkbox_state_NormalCase)
        QCheckBox_CurrentCase_Widget.stateChanged.connect(self.checkbox_state_CurrentCase)

        QPushButton_Widget00.clicked.connect(self.savepath)

    def External_Auxiliary_Equipment_changed(self, s):
        if s == "None":
            self.relay_voltage = "None"
        else:
            self.relay_voltage = "RELAY"

    def on_current_index_changed(self):
        selected_text = self.QComboBox_DUT.currentText()
        self.update_selection(selected_text)
        self.QLineEdit_PSU_VisaAddress.setCurrentText(self.PSU)
        self.QLineEdit_OSC_VisaAddress.setCurrentText(self.OSC)
        self.QLineEdit_ELoad_VisaAddress.setCurrentText(self.ELoad)
        self.QComboBox_Voltage_Sense.setCurrentText("4 Wire" if self.VoltageSense == "EXT" else "2 Wire")
        self.QComboBox_set_PSU_Channel.setCurrentIndex(int(self.PSU_Channel))
        self.QComboBox_set_ELoad_Channel.setCurrentIndex(int(self.ELoad_Channel))
        self.QLineEdit_OSC_DUT_Display_Channel.setCurrentText(self.DUT_OSC_Channel)
        self.QLineEdit_OSC_CurrentTrigger_Display_Channel.setCurrentText(self.CurrentTrigger_OSC_Channel)
        self.QComboBox_Trigger_Source_Display_Channel.setCurrentText(self.TriggerSource)
        self.QComboBox_External_Auxiliary_Equipment.setCurrentText(self.relay_voltage)
        self.QLineEdit_Power_Rating.setText(self.Power_Rating)
        self.QLineEdit_voltage_rated.setText(self.Voltage_Rating)
        self.QLineEdit_current_rated.setText(self.Current_Rating)
        self.QLineEdit_maxVoltage.setText(self.maxVoltage)
        self.QLineEdit_maxCurrent.setText(self.maxCurrent)
        self.QLineEdit_DUT_V_Settling_Band.setText(self.DUT_V_Settling_Band)
        self.QLineEdit_DUT_T_Settling_Band.setText(self.DUT_T_Settling_Band)
        self.QLineEdit_CurrentTrigger_V_Settling_Band.setText(self.CurrentTrigger_V_Settling_Band)
        self.QComboBox_DUT_Channel_Unit.setCurrentText(self.DUT_Channel_Unit)
        self.QLineEdit_DUT_Channel_Offset.setText(self.DUT_Channel_Offset)
        self.QComboBox_DUT_Channel_CouplingMode.setCurrentText(self.DUT_Channel_CouplingMode)
        self.QComboBox_DUT_Trigger_Mode.setCurrentText(self.DUT_Trigger_Mode)
        self.QComboBox_DUT_Trigger_CouplingMode.setCurrentText(self.DUT_Trigger_CouplingMode)
        self.QComboBox_DUT_Trigger_SweepMode.setCurrentText(self.DUT_Trigger_SweepMode)
        self.QComboBox_DUT_Trigger_SlopeMode.setCurrentText(self.DUT_Trigger_SlopeMode)
        self.QComboBox_DUT_Probe_Setting.setCurrentText(self.DUT_Probe_Setting)
        self.QComboBox_DUT_Acq_Type.setCurrentText(self.DUT_Acq_Type)
        self.QLineEdit_DUT_TimeScale.setText(self.DUT_TimeScale)
        self.QLineEdit_DUT_VerticalScale.setText(self.DUT_VerticalScale)
        self.QComboBox_CurrentTrigger_Channel_Unit.setCurrentText(self.CurrentTrigger_Channel_Unit)
        self.QLineEdit_CurrentTrigger_Channel_Offset.setText(self.CurrentTrigger_Channel_Offset)
        self.QComboBox_CurrentTrigger_Channel_CouplingMode.setCurrentText(self.CurrentTrigger_Channel_CouplingMode)
        self.QComboBox_CurrentTrigger_Trigger_Mode.setCurrentText(self.CurrentTrigger_Trigger_Mode)
        self.QComboBox_CurrentTrigger_Trigger_CouplingMode.setCurrentText(self.CurrentTrigger_Trigger_CouplingMode)
        self.QComboBox_CurrentTrigger_Trigger_SweepMode.setCurrentText(self.CurrentTrigger_Trigger_SweepMode)
        self.QComboBox_CurrentTrigger_Trigger_SlopeMode.setCurrentText(self.CurrentTrigger_Trigger_SlopeMode)
        self.QComboBox_CurrentTrigger_Probe_Setting.setCurrentText(self.CurrentTrigger_Probe_Setting)
        self.QComboBox_CurrentTrigger_Acq_Type.setCurrentText(self.CurrentTrigger_Acq_Type)
        self.QLineEdit_CurrentTrigger_VerticalScale.setText(self.CurrentTrigger_VerticalScale)


    def update_selection(self, selected_text):
        """Update selected text and reload config file"""
        self.selected_text = selected_text
        self.load_data()
    
    
    def load_data(self):
        """Reads configuration file and returns a dictionary of key-value pairs."""
        config_data = {}
        if self.selected_text =="Excavator":
            self.config_file = os.path.join(config_folder,"config_Excavator.txt")
            
        elif self.selected_text =="Dolphin":
            self.config_file = os.path.join(config_folder,"config_Dolphin.txt")
            
        elif self.selected_text =="SMU":
            self.config_file = os.path.join(config_folder,"config_SMU.txt")

        else:
            self.config_file = os.path.join(config_folder,"config_Others.txt")

        try:
            with open(self.config_file, "r") as file: 
                for line in file:

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue 

                    if "=" in line:
                        key, value = line.strip().split("=")
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    # If savelocation has a valid value, do not overwrite it
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue 
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe("Config file not found. Using default values.")

        return config_data


    def DUT_changed(self, s):
        self.DUT = s
    
    def CurrentTrigger_VerticalScale_changed(self, s):
        self.CurrentTrigger_VerticalScale = s

    def checkbox_state_CurrentCase(self, s):
        self.checkbox_CurrentCase = s
    
    def DUT_Channel_Offset_changed(self, s):
        self.DUT_Channel_Offset = s
    
    def CurrentTrigger_Channel_Offset_changed(self, s):
        self.CurrentTrigger_Channel_Offset = s

    def DUT_Channel_Unit_changed(self, s):
        self.DUT_Channel_Unit = s
    
    def CurrentTrigger_Channel_Unit_changed(self, s):
        self.CurrentTrigger_Channel_Unit = s

    def TriggerSource_changed(self, s):
        self.TriggerSource = s

    def Voltage_Rating_changed(self, value):
        self.Voltage_Rating = value
    
    def Current_Rating_changed(self, value):
        self.Current_Rating = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_OSC_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()

            discovery = NewGetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            instrument_roles = discovery.roles

            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_OSC_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])

            if 'PSU' in instrument_roles:
                psu_index = self.visaIdList.index(instrument_roles['PSU'])
                self.QLineEdit_PSU_VisaAddress.setCurrentIndex(psu_index)

            if 'ELOAD' in instrument_roles:
                eload_index = self.visaIdList.index(instrument_roles['ELOAD'])
                self.QLineEdit_ELoad_VisaAddress.setCurrentIndex(eload_index)

            if 'SCOPE' in instrument_roles:
                osc_index = self.visaIdList.index(instrument_roles['SCOPE'])
                self.QLineEdit_OSC_VisaAddress.setCurrentIndex(osc_index)
            
           
        except:
            self.OutputBox.append("No Devices Found!!!")
        return   
    
    def checkbox_state_SpecialCase(self, s):
        self.checkbox_SpecialCase = s

    def checkbox_state_NormalCase(self, s):
        self.checkbox_NormalCase = s
    
    def PSU_VisaAddress_changed(self, s):
        self.PSU = s

    def OSC_VisaAddress_changed(self, s):
        self.OSC = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s
    
    def ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s

    def PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def DUT_OSC_Channel_changed(self, s):
        self.DUT_OSC_Channel = s
    
    def CurrentTrigger_OSC_Channel_changed(self, s):
        self.CurrentTrigger_OSC_Channel = s

    def Power_Rating_changed(self, s):
        self.Power_Rating = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def I_Step_changed(self, s):
        self.I_Step = s

    def CurrentTrigger_T_Settling_Band_changed(self, s):
        self.CurrentTrigger_T_Settling_Band = s
    
    def CurrentTrigger_V_Settling_Band_changed(self, s):
        self.CurrentTrigger_V_Settling_Band = s
    
    def CurrentTrigger_Channel_CouplingMode_changed(self, s):
        self.CurrentTrigger_Channel_CouplingMode = s
    
    def CurrentTrigger_Trigger_CouplingMode_changed(self, s):
        self.CurrentTrigger_Trigger_CouplingMode = s
    
    def CurrentTrigger_Trigger_Mode_changed(self, s):
        self.CurrentTrigger_Trigger_Mode = s

    def CurrentTrigger_Trigger_SweepMode_changed(self, s):
        self.CurrentTrigger_Trigger_SweepMode = s
    
    def CurrentTrigger_Trigger_SlopeMode_changed(self, s):
        self.CurrentTrigger_Trigger_SlopeMode = s
    
    def CurrentTrigger_Probe_Setting_changed(self, s):
        self.CurrentTrigger_Probe_Setting = s
    
    def CurrentTrigger_Acq_Type_changed(self, s):
        self.CurrentTrigger_Acq_Type = s

#########################################################

    def DUT_T_Settling_Band_changed(self, s):
        self.DUT_T_Settling_Band = s

    def DUT_V_Settling_Band_changed(self, s):
        self.DUT_V_Settling_Band = s

    def DUT_Channel_CouplingMode_changed(self, s):
        self.DUT_Channel_CouplingMode = s

    def DUT_Trigger_CouplingMode_changed(self, s):
        self.DUT_Trigger_CouplingMode = s

    def DUT_Trigger_Mode_changed(self, s):
        self.DUT_Trigger_Mode = s

    def DUT_Trigger_SweepMode_changed(self, s):
        self.DUT_Trigger_SweepMode = s

    def DUT_Trigger_SlopeMode_changed(self, s):
        self.DUT_Trigger_SlopeMode = s
    
    def DUT_Probe_Setting_changed(self, s):
        self.DUT_Probe_Setting = s
    
    def DUT_Acq_Type_changed(self, s):
        self.DUT_Acq_Type = s

    def DUT_TimeScale_changed(self, s):
        self.DUT_TimeScale = s

    def DUT_VerticalScale_changed(self, s):
        self.DUT_VerticalScale = s



    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))


    def executeTest(self):
        """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
        then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
        will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
        begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
        begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
        are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
        optionally export all the details into a CSV file or display a graph after the test is completed.

        """
        dict = dictGenerator.input(
            savedir=self.savelocation,
            Instrument="Keysight",
            PSU=self.PSU,
            OSC=self.OSC,
            ELoad=self.ELoad,
            V_Rating=self.Voltage_Rating,
            I_Rating=self.Current_Rating,
            P_Rating=self.Power_Rating,
            power=self.Power_Rating,
            maxCurrent=self.maxCurrent,
            maxVoltage=self.maxVoltage,
            ELoad_Channel=self.ELoad_Channel,
            PSU_Channel=self.PSU_Channel,

            DUT_OSC_Channel=self.DUT_OSC_Channel,
            CurrentTrigger_OSC_Channel=self.CurrentTrigger_OSC_Channel,

            TriggerSource=self.TriggerSource,

            VoltageSense=self.VoltageSense,
            setFunction=self.setFunction,
            
            CurrentTrigger_Channel_Unit=self.CurrentTrigger_Channel_Unit,
            CurrentTrigger_Channel_Offset=self.CurrentTrigger_Channel_Offset,
            CurrentTrigger_Channel_CouplingMode=self.CurrentTrigger_Channel_CouplingMode,
            CurrentTrigger_Trigger_Mode=self.CurrentTrigger_Trigger_Mode,
            CurrentTrigger_Trigger_CouplingMode=self.CurrentTrigger_Trigger_CouplingMode,
            CurrentTrigger_Trigger_SweepMode=self.CurrentTrigger_Trigger_SweepMode,
            CurrentTrigger_Trigger_SlopeMode=self.CurrentTrigger_Trigger_SlopeMode,
            CurrentTrigger_Probe_Setting=self.CurrentTrigger_Probe_Setting,
            CurrentTrigger_Acq_Type=self.CurrentTrigger_Acq_Type,
            CurrentTrigger_VerticalScale=self.CurrentTrigger_VerticalScale,
           
            DUT_Channel_Unit=self.DUT_Channel_Unit,
            DUT_Channel_Offset=self.DUT_Channel_Offset,
            DUT_Channel_CouplingMode=self.DUT_Channel_CouplingMode,
            DUT_Trigger_Mode=self.DUT_Trigger_Mode,
            DUT_Trigger_CouplingMode=self.DUT_Trigger_CouplingMode,
            DUT_Trigger_SweepMode=self.DUT_Trigger_SweepMode,
            DUT_Trigger_SlopeMode=self.DUT_Trigger_SlopeMode,
            DUT_Probe_Setting=self.DUT_Probe_Setting,
            DUT_Acq_Type=self.DUT_Acq_Type,
            DUT_TimeScale=self.DUT_TimeScale,
            DUT_VerticalScale=self.DUT_VerticalScale,
            
            DUT_V_Settling_Band=self.DUT_V_Settling_Band,
            DUT_T_Settling_Band=self.DUT_T_Settling_Band,
            CurrentTrigger_V_Settling_Band=self.CurrentTrigger_V_Settling_Band,
        )
        QMessageBox.warning(
            self,
            "In Progress",
            "Measurement will start now , please do not close the main window until test is completed",
        )

        #Execute Voltage Measurement
        relay_voltage = RelayController_Voltage()
        if self.relay_voltage == "RELAY":
            relay_voltage.relay_on()
        else:
            relay_voltage.relay_off()
        try:
            """if self.checkbox_SpecialCase == 2:
                RiseFallTime.executeA(self, dict)
            
            if self.checkbox_NormalCase == 2:
                RiseFallTime.executeB(self, dict)"""

            if self.checkbox_CurrentCase == 2:
                DolphinRiseFallTimewithELoad.executeC(self, dict)

        except Exception as e:
            print_console_safe(e)
            QMessageBox.warning(self, "Error", str(e))
            exit()

        self.OutputBox.append(my_result.getvalue())
        self.OutputBox.append("Measurement is complete !")

class ProgrammingSpeed(QDialog):
    """Class for configuring the Programming Speed DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.

    """

    def __init__(self):
        """ "Method declaring the Widgets, Signals & Slots for Programming Speed."""
        super().__init__()

         # Default Values
        self.PSU = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        self.OSC = "USB0::0x0957::0x17B0::MY52060151::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        self.ELoad_Channel = "1"
        self.PSU_Channel = "2"
        self.DUT_OSC_Channel = "CHANNEL1"
        self.TriggerSource = "CHANNEL1"
        self.setFunction = "Current"

        self.VoltageSense = "EXT"
        self.Power_Rating = "160"
        self.Current_Rating = "120"
        self.Voltage_Rating = "80"
        self.maxCurrent = "10"
        self.maxVoltage = "80"

        self.relay_voltage = "RELAY"
        
        #From OSC view
        self.DUT_Channel_Unit = "VOLT"
        self.DUT_Channel_Offset = "0"
    
        self.DUT_Channel_CouplingMode = "DC"
        self.DUT_Trigger_CouplingMode = "DC"
        self.DUT_Trigger_Mode = "EDGE"
        self.DUT_Trigger_SweepMode = "NORMAL"
        self.DUT_Trigger_SlopeMode = "EITHer"
        self.DUT_TimeScale = "0.01"
        self.DUT_VerticalScale = "0.00001"
        self.DUT_V_Settling_Band = "0.8"
        self.DUT_T_Settling_Band = "0.001"
        self.DUT_Probe_Setting = "X10"
        self.DUT_Acq_Type = "AVERage"
        self.DUT_Unit = "VOLT"

        #CV Programming Speed Response Limits
        self.Programming_Response_Up_NoLoad = "0.008"
        self.Programming_Response_Up_FullLoad = "0.008"
        self.Programming_Response_Down_NoLoad = "0.5"
        self.Programming_Response_Down_FullLoad = "0.06"

        #Delay Time
        self.updatedelay = "3"

        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"


        self.setWindowTitle("Programming Speed")

        QPushButton_Widget00 = QPushButton()
        QPushButton_Widget00.setText("Save Path")
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Find Instruments")
        QPushButton_Widget = QPushButton()
        QPushButton_Widget.setText("Execute Test")

        
    
        layout1 = QFormLayout()
        self.OutputBox = QTextBrowser()
        self.OutputBox.append(my_result.getvalue())

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp8 = QLabel()

        Desp0.setFont(desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Specification/Test Settings:")
        Desp4.setText("DUT Oscilloscope Settings:")
        Desp8.setText("External Auxiliary Equipment Settings:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        # Connections
        QLabel_PSU_VisaAddress = QLabel()
        QLabel_OSC_VisaAddress = QLabel()
        QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DUT = QLabel()

        QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        QLabel_OSC_VisaAddress.setText("Visa Address (OSC):")
        QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DUT.setText("DUT:")

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_OSC_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        self.QComboBox_DUT = QComboBox()

        # General Settings
        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_OSC_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QComboBox_DUT.addItems(["Others", "Excavator", "EDU36311A", "Dolphin", "Hornbill", "SMU"])

        QLabel_OSC_DUT_Display_Channel = QLabel() #DUT Source 
        QLabel_Trigger_Source_Display_Channel = QLabel()

        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Power_Rating = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_rated = QLabel()
        QLabel_current_rated = QLabel()
        QLabel_maxCurrent = QLabel()

        QLabel_DUT_V_Settling_Band = QLabel()
        QLabel_DUT_T_Settling_Band = QLabel()
        QLabel_DUT_Probe_Setting = QLabel()
        QLabel_DUT_Acq_Type = QLabel()

        QLabel_OSC_DUT_Display_Channel.setText("DUT Display Channel (OSC)")
        QLabel_Trigger_Source_Display_Channel.setText("Trigger Source Channel (OSC)")
        
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Power_Rating.setText("Power Rating (W):")
        QLabel_maxVoltage.setText("Testing Voltage (V):")
        QLabel_voltage_rated.setText("DUT Rated Voltage:")
        QLabel_current_rated.setText("DUT Rated Current:")
        QLabel_maxCurrent.setText("Testing Current (A):")
        
        """QLabel_DUT_V_Settling_Band.setText("DUT Settling Band Voltage (V) / Error Band:")
        QLabel_DUT_T_Settling_Band.setText("DUT Settling Band Time (s):")"""
        QLabel_DUT_Probe_Setting.setText("DUT Probe Setting Ratio:")
        QLabel_DUT_Acq_Type.setText("DUT Acquire Mode:")

        self.QLineEdit_OSC_DUT_Display_Channel = QComboBox()
        self.QComboBox_Trigger_Source_Display_Channel = QComboBox()

        self.QLineEdit_OSC_DUT_Display_Channel.addItems(["CHANNEL1", "CHANNEL2", "CHANNEL3", "CHANNEL4"])
        self.QComboBox_Trigger_Source_Display_Channel.addItems(["CHANNEL1", "CHANNEL2", "CHANNEL3", "CHANNEL4"])

        # External Auxiliary Equipment section
        QLabel_External_Auxiliary_Equipment = QLabel()
        QLabel_External_Auxiliary_Equipment.setText("Relay")
        self.QComboBox_External_Auxiliary_Equipment = QComboBox()
        self.QComboBox_External_Auxiliary_Equipment.addItems(["None", "RELAY"])

        self.QComboBox_set_PSU_Channel = QComboBox()
        self.QComboBox_set_ELoad_Channel = QComboBox()
        self.QComboBox_Voltage_Sense = QComboBox()
        self.QComboBox_set_Function = QComboBox()
        self.QLineEdit_Power_Rating = QLineEdit()
        self.QLineEdit_maxVoltage = QLineEdit()
        self.QLineEdit_voltage_rated = QLineEdit()
        self.QLineEdit_current_rated = QLineEdit()
        self.QLineEdit_maxCurrent = QLineEdit()

        self.QLineEdit_DUT_V_Settling_Band = QLineEdit()
        self.QLineEdit_DUT_T_Settling_Band = QLineEdit()
        self.QComboBox_DUT_Probe_Setting = QComboBox()
        self.QComboBox_DUT_Acq_Type = QComboBox()

        self.QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        self.QComboBox_set_Function.setEnabled(False)
        self.QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])

        self.QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        self.QComboBox_set_PSU_Channel.setEnabled(True)
        self.QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        self.QComboBox_set_ELoad_Channel.setEnabled(True)

        self.QComboBox_DUT_Probe_Setting.addItems(["X1", "X10", "X20", "X100"])
        self.QComboBox_DUT_Acq_Type.addItems(["NORMal", "PEAK", "AVERage", "HRESolution"])

         # DUT Oscilloscope Settings
        QLabel_DUT_Channel_Unit = QLabel()
        QLabel_DUT_Channel_Offset = QLabel()
        QLabel_DUT_Channel_CouplingMode = QLabel()
        QLabel_DUT_Trigger_Mode = QLabel()
        QLabel_DUT_Trigger_CouplingMode = QLabel()
        QLabel_DUT_Trigger_SweepMode = QLabel()
        QLabel_DUT_Trigger_SlopeMode = QLabel()
        QLabel_DUT_TimeScale = QLabel()
        QLabel_DUT_VerticalScale = QLabel()

        QLabel_DUT_Channel_Unit.setText("Unit (DUT Channel)")
        QLabel_DUT_Channel_Offset.setText("Offset (DUT Channel)")
        QLabel_DUT_Channel_CouplingMode.setText("Coupling Mode (Channel)")
        QLabel_DUT_Trigger_Mode.setText("Trigger Mode:")
        QLabel_DUT_Trigger_CouplingMode.setText("Coupling Mode (Trigger):")
        QLabel_DUT_Trigger_SweepMode.setText("Trigger Sweep Mode:")
        QLabel_DUT_Trigger_SlopeMode.setText("Trigger Slope Mode:")
        QLabel_DUT_TimeScale.setText("Time Scale:")
        QLabel_DUT_VerticalScale.setText("Vertical Scale:(DUT(V))")
        
        self.QComboBox_DUT_Channel_Unit = QComboBox()
        self.QLineEdit_DUT_Channel_Offset = QLineEdit()
        self.QComboBox_DUT_Channel_CouplingMode = QComboBox()
        self.QComboBox_DUT_Trigger_Mode = QComboBox()
        self.QComboBox_DUT_Trigger_CouplingMode = QComboBox()
        self.QComboBox_DUT_Trigger_SweepMode = QComboBox()
        self.QComboBox_DUT_Trigger_SlopeMode = QComboBox()
        self.QLineEdit_DUT_TimeScale = QLineEdit()
        self.QLineEdit_DUT_VerticalScale = QLineEdit()

        self.QComboBox_DUT_Channel_Unit.addItems(["VOLT", "AMPere"])
        self.QComboBox_DUT_Channel_CouplingMode.addItems(["DC", "AC"])
        self.QComboBox_DUT_Trigger_Mode.addItems(["EDGE", "IIC", "EBUR"])
        self.QComboBox_DUT_Trigger_CouplingMode.addItems(["DC", "AC"])
        self.QComboBox_DUT_Trigger_SweepMode.addItems(["NORMAL", "AUTO"])
        self.QComboBox_DUT_Trigger_SlopeMode.addItems(["ALT", "POS", "NEG", "EITH"])

        self.QComboBox_DUT_Channel_Unit.setEnabled(True)
        self.QComboBox_DUT_Channel_CouplingMode.setEnabled(True)
        self.QComboBox_DUT_Trigger_Mode.setEnabled(False)
        self.QComboBox_DUT_Trigger_CouplingMode.setEnabled(False)
        self.QComboBox_DUT_Trigger_SweepMode.setEnabled(False)
        self.QComboBox_DUT_Trigger_SlopeMode.setEnabled(False)
        self.QComboBox_DUT_Probe_Setting.setEnabled(True)
        self.QComboBox_DUT_Acq_Type.setEnabled(False)

        QLabel_Programming_Response_Up_NoLoad= QLabel()
        QLabel_Programming_Response_Up_FullLoad= QLabel()
        QLabel_Programming_Response_Down_NoLoad= QLabel()
        QLabel_Programming_Response_Down_FullLoad= QLabel()

        QLabel_Programming_Response_Up_NoLoad.setText("Programming Response Limit (Up-NoLoad)")
        QLabel_Programming_Response_Up_FullLoad.setText("Programming Response Limit (Up-FullLoad)")
        QLabel_Programming_Response_Down_NoLoad.setText("Programming Response Limit (Down-NoLoad)")
        QLabel_Programming_Response_Down_FullLoad.setText("Programming Response Limit (Down-FullLoad)")

        self.QLineEdit_Programming_Response_Up_NoLoad  = QLineEdit()
        self.QLineEdit_Programming_Response_Up_FullLoad  = QLineEdit()
        self.QLineEdit_Programming_Response_Down_NoLoad = QLineEdit()
        self.QLineEdit_Programming_Response_Down_FullLoad = QLineEdit()
      



        # Create a horizontal layout for the "Save Path" and checkboxes
        Rated_Power_Voltage_Current_Layout = QHBoxLayout()
        Rated_Power_Voltage_Current_Layout.addWidget(QLabel_Power_Rating)  # QLabel for "Save Path"
        Rated_Power_Voltage_Current_Layout.addWidget(self.QLineEdit_Power_Rating)  # Checkbox for "Generate Excel Report"
        Rated_Power_Voltage_Current_Layout.addWidget(QLabel_voltage_rated)  # Checkbox for "Show Graph"
        Rated_Power_Voltage_Current_Layout.addWidget(self.QLineEdit_voltage_rated)  # Checkbox for "Show Graph"
        Rated_Power_Voltage_Current_Layout.addWidget(QLabel_current_rated)  # QLabel for "Save Path"
        Rated_Power_Voltage_Current_Layout.addWidget(self.QLineEdit_current_rated)  # Checkbox for "Generate Excel Report"

        Test_Voltage_Current_Layout = QHBoxLayout()
        Test_Voltage_Current_Layout.addWidget(QLabel_maxVoltage)  # QLabel for "Save Path"
        Test_Voltage_Current_Layout.addWidget(self.QLineEdit_maxVoltage)  # Checkbox for "Generate Excel Report"
        Test_Voltage_Current_Layout.addWidget(QLabel_maxCurrent)  # Checkbox for "Show Graph"
        Test_Voltage_Current_Layout.addWidget(self.QLineEdit_maxCurrent)  # Checkbox for "Show Graph"


        Channel_CouplingMode_layout = QHBoxLayout()
        Channel_CouplingMode_layout.addWidget(QLabel_DUT_Channel_CouplingMode)  # QLabel for "Save Path"
        Channel_CouplingMode_layout.addWidget(self.QComboBox_DUT_Channel_CouplingMode)  # Checkbox for "Generate Excel Report"

        Trigger_Mode_layout = QHBoxLayout()
        Trigger_Mode_layout.addWidget(QLabel_DUT_Trigger_Mode)  # Checkbox for "Show Graph"
        Trigger_Mode_layout.addWidget(self.QComboBox_DUT_Trigger_Mode)  # Checkbox for "Show Graph"

        Unit_Channel_layout = QHBoxLayout()
        Unit_Channel_layout.addWidget(QLabel_DUT_Channel_Unit)  # QLabel for "Save Path"
        Unit_Channel_layout.addWidget(self.QComboBox_DUT_Channel_Unit)  # Checkbox for "Generate Excel Report"

        Offset_Channel_layout = QHBoxLayout()
        Offset_Channel_layout.addWidget(QLabel_DUT_Channel_Offset)  # Checkbox for "Show Graph"
        Offset_Channel_layout.addWidget(self.QLineEdit_DUT_Channel_Offset)  # Checkbox for "Show Graph"

        Trigger_CouplingMode_layout = QHBoxLayout()
        Trigger_CouplingMode_layout.addWidget(QLabel_DUT_Trigger_CouplingMode)  # QLabel for "Save Path"
        Trigger_CouplingMode_layout.addWidget(self.QComboBox_DUT_Trigger_CouplingMode)  # Checkbox for "Generate Excel Report"

        Trigger_SweepMode_layout = QHBoxLayout()
        Trigger_SweepMode_layout.addWidget(QLabel_DUT_Trigger_SweepMode)  # Checkbox for "Show Graph"  
        Trigger_SweepMode_layout.addWidget(self.QComboBox_DUT_Trigger_SweepMode)  # Checkbox for "Show Graph"

        Trigger_SlopeMode_layout = QHBoxLayout()
        Trigger_SlopeMode_layout.addWidget(QLabel_DUT_Trigger_SlopeMode)  # QLabel for "Save Path"
        Trigger_SlopeMode_layout.addWidget(self.QComboBox_DUT_Trigger_SlopeMode)  # Checkbox for "Generate Excel Report"

        Probe_Setting_layout = QHBoxLayout()
        Probe_Setting_layout.addWidget(QLabel_DUT_Probe_Setting)  # QLabel for "Save Path"
        Probe_Setting_layout.addWidget(self.QComboBox_DUT_Probe_Setting)  # Checkbox for "Generate Excel Report"

        TimeScale_layout = QHBoxLayout()
        TimeScale_layout.addWidget(QLabel_DUT_TimeScale)  # Checkbox for "Show Graph"
        TimeScale_layout.addWidget(self.QLineEdit_DUT_TimeScale)  # Checkbox for "Show Graph"

        VerticalScale_layout = QHBoxLayout()
        VerticalScale_layout.addWidget(QLabel_DUT_VerticalScale)  # QLabel for "Save Path"
        VerticalScale_layout.addWidget(self.QLineEdit_DUT_VerticalScale)  # Checkbox for "Generate Excel Report"

        Programming_Response_layout = QHBoxLayout()
        Programming_Response_layout.addWidget(QLabel_Programming_Response_Up_NoLoad)
        Programming_Response_layout.addWidget(self.QLineEdit_Programming_Response_Up_NoLoad)
        Programming_Response_layout.addWidget(QLabel_Programming_Response_Up_FullLoad)
        Programming_Response_layout.addWidget(self.QLineEdit_Programming_Response_Up_FullLoad)
        Programming_Response_layout.addWidget(QLabel_Programming_Response_Down_NoLoad)
        Programming_Response_layout.addWidget(self.QLineEdit_Programming_Response_Down_NoLoad)
        Programming_Response_layout.addWidget(QLabel_Programming_Response_Down_FullLoad)
        Programming_Response_layout.addWidget(self.QLineEdit_Programming_Response_Down_FullLoad)


        layout1.addRow(Desp0)
        layout1.addRow(QPushButton_Widget00)
        layout1.addRow(self.OutputBox)
        layout1.addRow(QPushButton_Widget0)
        layout1.addRow(Desp1)
        layout1.addRow(QLabel_DUT, self.QComboBox_DUT)
        layout1.addRow(QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        layout1.addRow(QLabel_OSC_VisaAddress, self.QLineEdit_OSC_VisaAddress)
        layout1.addRow(QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_set_Function, self.QComboBox_set_Function)
        layout1.addRow(QLabel_Voltage_Sense, self.QComboBox_Voltage_Sense)
        layout1.addRow(QLabel_set_PSU_Channel, self.QComboBox_set_PSU_Channel)
        layout1.addRow(QLabel_set_ELoad_Channel, self.QComboBox_set_ELoad_Channel)
        layout1.addRow(QLabel_OSC_DUT_Display_Channel, self.QLineEdit_OSC_DUT_Display_Channel)
        layout1.addRow(QLabel_Trigger_Source_Display_Channel, self.QComboBox_Trigger_Source_Display_Channel)
        layout1.addRow(Desp8)
        layout1.addRow(QLabel_External_Auxiliary_Equipment, self.QComboBox_External_Auxiliary_Equipment)
        layout1.addRow(Desp3)
        layout1.addRow(Rated_Power_Voltage_Current_Layout)
        layout1.addRow(Test_Voltage_Current_Layout)
        layout1.addRow(Programming_Response_layout)
        
        layout1.addRow(Desp4)#DUT Oscilloscope Settings

        layout1.addRow(Unit_Channel_layout)
        layout1.addRow(Offset_Channel_layout)
        layout1.addRow(Channel_CouplingMode_layout)
        layout1.addRow(Trigger_Mode_layout)
        layout1.addRow(Trigger_CouplingMode_layout)
        layout1.addRow(Trigger_SweepMode_layout)
        layout1.addRow(Trigger_SlopeMode_layout)
        layout1.addRow(Probe_Setting_layout)
        layout1.addRow(TimeScale_layout)
        layout1.addRow(VerticalScale_layout)

        layout1.addRow(QPushButton_Widget)
        self.setLayout(layout1)

        QPushButton_Widget0.clicked.connect(self.doFind)
        QPushButton_Widget.clicked.connect(self.executeTest)
    
        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_OSC_VisaAddress.currentTextChanged.connect(self.OSC_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)
        self.QComboBox_DUT.currentTextChanged.connect(self.on_current_index_changed)
        self.QComboBox_set_PSU_Channel.currentTextChanged.connect(self.PSU_Channel_changed)
        self.QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        self.QComboBox_External_Auxiliary_Equipment.currentTextChanged.connect(self.External_Auxiliary_Equipment_changed)

        self.QLineEdit_OSC_DUT_Display_Channel.currentTextChanged.connect(self.DUT_OSC_Channel_changed)
        self.QComboBox_Trigger_Source_Display_Channel.currentTextChanged.connect(self.TriggerSource_changed)
        
        self.QLineEdit_Power_Rating.textEdited.connect(self.Power_Rating_changed)
        self.QLineEdit_voltage_rated.textEdited.connect(self.Voltage_Rating_changed)
        self.QLineEdit_current_rated.textEdited.connect(self.Current_Rating_changed)

        self.QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        self.QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)

        self.QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)

        self.QComboBox_Voltage_Sense.currentTextChanged.connect(
            self.set_VoltageSense_changed
        )
        # Programming Response Settings
        self.QLineEdit_Programming_Response_Up_NoLoad.textEdited.connect(self.Programming_Response_Up_NoLoad_changed)
        self.QLineEdit_Programming_Response_Up_FullLoad.textEdited.connect(self.Programming_Response_Up_FullLoad_changed)
        self.QLineEdit_Programming_Response_Down_NoLoad.textEdited.connect(self.Programming_Response_Down_NoLoad_changed)
        self.QLineEdit_Programming_Response_Down_FullLoad.textEdited.connect(self.Programming_Response_Down_FullLoad_changed)

        #DUT Oscilloscope Settings

        self.QComboBox_DUT_Channel_Unit.currentTextChanged.connect(
            self.DUT_Channel_Unit_changed
        )

        self.QLineEdit_DUT_Channel_Offset.textEdited.connect(self.DUT_Channel_Offset_changed)

        self.QComboBox_DUT_Channel_CouplingMode.currentTextChanged.connect(
            self.DUT_Channel_CouplingMode_changed
        )
        self.QComboBox_DUT_Trigger_CouplingMode.currentTextChanged.connect(
            self.DUT_Trigger_CouplingMode_changed
        )
        self.QComboBox_DUT_Trigger_Mode.currentTextChanged.connect(self.DUT_Trigger_Mode_changed)
        self.QComboBox_DUT_Trigger_SweepMode.currentTextChanged.connect(
            self.DUT_Trigger_SweepMode_changed
        )
        self.QComboBox_DUT_Trigger_SlopeMode.currentTextChanged.connect(
            self.DUT_Trigger_SlopeMode_changed
        )
        self.QComboBox_DUT_Probe_Setting.currentTextChanged.connect(
            self.DUT_Probe_Setting_changed
        )
        self.QComboBox_DUT_Acq_Type.currentTextChanged.connect(
            self.DUT_Acq_Type_changed
        )
        self.QLineEdit_DUT_TimeScale.textEdited.connect(self.DUT_TimeScale_changed)

        self.QLineEdit_DUT_VerticalScale.textEdited.connect(self.DUT_VerticalScale_changed)

        QPushButton_Widget00.clicked.connect(self.savepath)
    
    def Programming_Response_Up_NoLoad_changed(self, s):
        self.Programming_Response_Up_NoLoad = s

    def Programming_Response_Up_FullLoad_changed(self, s):
        self.Programming_Response_Up_FullLoad = s

    def Programming_Response_Down_NoLoad_changed(self, s):
        self.Programming_Response_Down_NoLoad = s

    def Programming_Response_Down_FullLoad_changed(self, s):
        self.Programming_Response_Down_FullLoad = s
    
    def External_Auxiliary_Equipment_changed(self, s):
        if s == "None":
            self.relay_voltage = "None"
        else:
            self.relay_voltage = "RELAY"

    def on_current_index_changed(self):
        selected_text = self.QComboBox_DUT.currentText()
        self.update_selection(selected_text)
        self.QLineEdit_PSU_VisaAddress.setCurrentText(self.PSU)
        self.QLineEdit_OSC_VisaAddress.setCurrentText(self.OSC)
        self.QLineEdit_ELoad_VisaAddress.setCurrentText(self.ELoad)
        self.QComboBox_Voltage_Sense.setCurrentText("4 Wire" if self.VoltageSense == "EXT" else "2 Wire")
        self.QComboBox_set_PSU_Channel.setCurrentIndex(int(self.PSU_Channel))
        self.QComboBox_set_ELoad_Channel.setCurrentIndex(int(self.ELoad_Channel))
        self.QLineEdit_OSC_DUT_Display_Channel.setCurrentText(self.DUT_OSC_Channel)
        self.QComboBox_Trigger_Source_Display_Channel.setCurrentText(self.TriggerSource)
        self.QComboBox_External_Auxiliary_Equipment.setCurrentText(self.relay_voltage)
        self.QLineEdit_Power_Rating.setText(self.Power_Rating)
        self.QLineEdit_voltage_rated.setText(self.Voltage_Rating)
        self.QLineEdit_current_rated.setText(self.Current_Rating)
        self.QLineEdit_maxVoltage.setText(self.maxVoltage)
        self.QLineEdit_maxCurrent.setText(self.maxCurrent)
        
        self.QLineEdit_Programming_Response_Up_NoLoad.setText(self.Programming_Response_Up_NoLoad)
        self.QLineEdit_Programming_Response_Up_FullLoad.setText(self.Programming_Response_Up_FullLoad)
        self.QLineEdit_Programming_Response_Down_NoLoad.setText(self.Programming_Response_Down_NoLoad)
        self.QLineEdit_Programming_Response_Down_FullLoad.setText(self.Programming_Response_Down_FullLoad)

        self.QComboBox_DUT_Channel_Unit.setCurrentText(self.DUT_Channel_Unit)
        self.QLineEdit_DUT_Channel_Offset.setText(self.DUT_Channel_Offset)
        self.QComboBox_DUT_Channel_CouplingMode.setCurrentText(self.DUT_Channel_CouplingMode)
        self.QComboBox_DUT_Trigger_Mode.setCurrentText(self.DUT_Trigger_Mode)
        self.QComboBox_DUT_Trigger_CouplingMode.setCurrentText(self.DUT_Trigger_CouplingMode)
        self.QComboBox_DUT_Trigger_SweepMode.setCurrentText(self.DUT_Trigger_SweepMode)
        self.QComboBox_DUT_Trigger_SlopeMode.setCurrentText(self.DUT_Trigger_SlopeMode)
        self.QComboBox_DUT_Probe_Setting.setCurrentText(self.DUT_Probe_Setting)
        self.QComboBox_DUT_Acq_Type.setCurrentText(self.DUT_Acq_Type)
        self.QLineEdit_DUT_TimeScale.setText(self.DUT_TimeScale)
        self.QLineEdit_DUT_VerticalScale.setText(self.DUT_VerticalScale)


    def update_selection(self, selected_text):
        """Update selected text and reload config file"""
        self.selected_text = selected_text
        self.load_data()
    
    
    def load_data(self):
        """Reads configuration file and returns a dictionary of key-value pairs."""
        config_data = {}
        if self.selected_text =="Excavator":
            self.config_file = os.path.join(config_folder,"config_Excavator.txt")
            
        elif self.selected_text =="Dolphin":
            self.config_file = os.path.join(config_folder,"config_Dolphin.txt")
            
        elif self.selected_text =="SMU":
            self.config_file = os.path.join(config_folder,"config_SMU.txt")

        else:
            self.config_file = os.path.join(config_folder,"config_Others.txt")

        try:
            with open(self.config_file, "r") as file: 
                for line in file:

                    if not line or line.startswith("#") or line.startswith("//"):
                        continue 

                    if "=" in line:
                        key, value = line.strip().split("=")
                        config_data[key.strip()] = value.strip()

            for key, value in config_data.items():
                if key == "savelocation":
                    # If savelocation has a valid value, do not overwrite it
                    if self.savelocation and self.savelocation != "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing":
                        continue 
                    else:
                        setattr(self, key, value)
                else:
                    setattr(self, key, value)

        except FileNotFoundError:
            print_console_safe("Config file not found. Using default values.")

        return config_data


    def DUT_changed(self, s):
        self.DUT = s
    
    def DUT_Channel_Offset_changed(self, s):
        self.DUT_Channel_Offset = s
    
    def CurrentTrigger_Channel_Offset_changed(self, s):
        self.CurrentTrigger_Channel_Offset = s

    def DUT_Channel_Unit_changed(self, s):
        self.DUT_Channel_Unit = s
    
    def CurrentTrigger_Channel_Unit_changed(self, s):
        self.CurrentTrigger_Channel_Unit = s

    def TriggerSource_changed(self, s):
        self.TriggerSource = s

    def Voltage_Rating_changed(self, value):
        self.Voltage_Rating = value
    
    def Current_Rating_changed(self, value):
        self.Current_Rating = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_OSC_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()

            discovery = NewGetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            instrument_roles = discovery.roles

            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_OSC_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])

            if 'PSU' in instrument_roles:
                psu_index = self.visaIdList.index(instrument_roles['PSU'])
                self.QLineEdit_PSU_VisaAddress.setCurrentIndex(psu_index)

            if 'ELOAD' in instrument_roles:
                eload_index = self.visaIdList.index(instrument_roles['ELOAD'])
                self.QLineEdit_ELoad_VisaAddress.setCurrentIndex(eload_index)

            if 'SCOPE' in instrument_roles:
                osc_index = self.visaIdList.index(instrument_roles['SCOPE'])
                self.QLineEdit_OSC_VisaAddress.setCurrentIndex(osc_index)
            
           
        except:
            self.OutputBox.append("No Devices Found!!!")
        return   
    
    def PSU_VisaAddress_changed(self, s):
        self.PSU = s

    def OSC_VisaAddress_changed(self, s):
        self.OSC = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s
    
    def ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s

    def PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def DUT_OSC_Channel_changed(self, s):
        self.DUT_OSC_Channel = s

    def Power_Rating_changed(self, s):
        self.Power_Rating = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def I_Step_changed(self, s):
        self.I_Step = s

#########################################################

    def DUT_T_Settling_Band_changed(self, s):
        self.DUT_T_Settling_Band = s

    def DUT_V_Settling_Band_changed(self, s):
        self.DUT_V_Settling_Band = s

    def DUT_Channel_CouplingMode_changed(self, s):
        self.DUT_Channel_CouplingMode = s

    def DUT_Trigger_CouplingMode_changed(self, s):
        self.DUT_Trigger_CouplingMode = s

    def DUT_Trigger_Mode_changed(self, s):
        self.DUT_Trigger_Mode = s

    def DUT_Trigger_SweepMode_changed(self, s):
        self.DUT_Trigger_SweepMode = s

    def DUT_Trigger_SlopeMode_changed(self, s):
        self.DUT_Trigger_SlopeMode = s
    
    def DUT_Probe_Setting_changed(self, s):
        self.DUT_Probe_Setting = s
    
    def DUT_Acq_Type_changed(self, s):
        self.DUT_Acq_Type = s

    def DUT_TimeScale_changed(self, s):
        self.DUT_TimeScale = s

    def DUT_VerticalScale_changed(self, s):
        self.DUT_VerticalScale = s



    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))


    def executeTest(self):
        """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
        then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
        will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
        begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
        begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
        are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
        optionally export all the details into a CSV file or display a graph after the test is completed.

        """
        dict = dictGenerator.input(
            savedir=self.savelocation,
            Instrument="Keysight",
            PSU=self.PSU,
            OSC=self.OSC,
            ELoad=self.ELoad,
            V_Rating=self.Voltage_Rating,
            I_Rating=self.Current_Rating,
            P_Rating=self.Power_Rating,
            power=self.Power_Rating,
            maxCurrent=self.maxCurrent,
            maxVoltage=self.maxVoltage,
            ELoad_Channel=self.ELoad_Channel,
            PSU_Channel=self.PSU_Channel,
            DUT_OSC_Channel=self.DUT_OSC_Channel,
            TriggerSource=self.TriggerSource,
            VoltageSense=self.VoltageSense,
            setFunction=self.setFunction,
            DUT_Channel_Unit=self.DUT_Channel_Unit,
            DUT_Channel_Offset=self.DUT_Channel_Offset,
            DUT_Channel_CouplingMode=self.DUT_Channel_CouplingMode,
            DUT_Trigger_Mode=self.DUT_Trigger_Mode,
            DUT_Trigger_CouplingMode=self.DUT_Trigger_CouplingMode,
            DUT_Trigger_SweepMode=self.DUT_Trigger_SweepMode,
            DUT_Trigger_SlopeMode=self.DUT_Trigger_SlopeMode,
            DUT_Probe_Setting=self.DUT_Probe_Setting,
            DUT_Acq_Type=self.DUT_Acq_Type,
            DUT_TimeScale=self.DUT_TimeScale,
            DUT_VerticalScale=self.DUT_VerticalScale,
            Response_Up_NoLoad=self.Programming_Response_Up_NoLoad,
            Response_Up_FullLoad=self.Programming_Response_Up_FullLoad,
            Response_Down_NoLoad=self.Programming_Response_Down_NoLoad,
            Response_Down_FullLoad=self.Programming_Response_Down_FullLoad,
            updatedelay = self.updatedelay,
        )
        QMessageBox.warning(
            self,
            "In Progress",
            "Measurement will start now , please do not close the main window until test is completed",
        )


        try:
            DolphinProgrammingResponse.execute(self, dict)
        
        except Exception as e:
            print_console_safe(e)
            QMessageBox.warning(self, "Error", str(e))
            exit()

        self.OutputBox.append(my_result.getvalue())
        self.OutputBox.append("Measurement is complete !")

class PowerMeasurementDialog(QDialog):
    """Class for configuring the voltage measurement DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.


    """

    def __init__(self):
        """Method where Widgets, Signals and Slots are defined in the GUI for Voltage Measurement"""
        super().__init__()
        # Initialize default CSV path
        self.DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerdata.csv"  # or use a dynamic path if needed
        self.IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/powerChart.png"
        self.ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powererror.csv"
        self.image_dialog = None

        
        # Default Values
        #self.PSU_VisaAddress = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        #self.DMM_VisaAddress = "USB0::0x2A8D::0x1601::MY60094787::0::INSTR"
        #self.ELoad_VisaAddress = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        self.Power = "2200"
        self.Power_Rating = "6000"
        self.powerini = "0"
        self.powerfin = "60"
        self.power_step_size = "10"

        self.rshunt = "0.01"
        self.powerloop = "1"
        self.currloop = "1"
        self.voltloop = "1"
        self.estimatetime = "0"
        self.readbackvoltage = "0"
        self.readbackcurrent = "0"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"
        self.noofloop = "1"
        self.updatedelay = "1"
        self.unit = "VOLTAGE"
        self.Programming_Error_Offset = "0.005"
        self.Programming_Error_Gain = "0.005"
        self.Readback_Error_Offset = "0.005"
        self.Readback_Error_Gain = "0.005"

        self.minCurrent = "1"
        self.maxCurrent = "10"
        self.current_step_size = "1"
        self.minVoltage = "1"
        self.maxVoltage = "80"
        self.voltage_step_size = "1"

        self.PSU = "USB0::0x2A8D::0xDA04::CN24350083::0::INSTR"
        self.DMM = "USB0::0x2A8D::0x0201::MY57702128::0::INSTR"
        self.DMM2 = "USB0::0x2A8D::0x0201::MY54701197::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"

        self.ELoad_Channel = "0"
        self.PSU_Channel = "1"
        self.DMM_Instrument = "Keysight"

        self.setFunction = "Current" #ELoad In CC Mode
        self.VoltageRes = "SLOW"
        self.VoltageSense = "INT"
        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.Range = "Auto"
        self.Aperture = "10"
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.UpTime = "50"
        self.DownTime = "50"

        # Ensure DATA_CSV_PATH is defined before use
        if not hasattr(self, 'DATA_CSV_PATH') or not self.DATA_CSV_PATH:
            QMessageBox.warning(self, "Error", "CSV path is not set.")

        self.setWindowTitle("Power Measurement")
        self.image_window = None
        self.setWindowFlags(Qt.Window)

        # create find button 
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget3 = QPushButton()
        QPushButton_Widget3.setText("Estimate Data Collection Time")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
        QCheckBox_Report_Widget = QCheckBox()
        QCheckBox_Report_Widget.setText("Generate Excel Report")
        QCheckBox_Report_Widget.setCheckState(Qt.Checked)
        QCheckBox_Image_Widget = QCheckBox()
        QCheckBox_Image_Widget.setText("Show Graph")
        QCheckBox_Image_Widget.setCheckState(Qt.Checked)

        layout1 = QFormLayout()
        self.OutputBox = QTextBrowser()
        self.OutputBox.setFixedSize(800, 20)  # Set width to 200 and height to 100
        self.OutputBox.append(my_result.getvalue())

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()
        Desp6 = QLabel()
        Desp7 = QLabel()

        Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)
        Desp6.setFont(desp_font)
        Desp7.setFont(desp_font)


        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Source Voltage Limit")
        Desp4.setText("Source Current Limit")
        Desp5.setText("No. of Collection:")
        Desp6.setText("Source Power Limit [W]")
        Desp7.setText("Resistive Shunt Value Setting")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        # Connections
        self.QLabel_PSU_VisaAddress = QLabel()
        self.QLabel_DMM_VisaAddressforVoltage = QLabel()
        self.QLabel_DMM_VisaAddressforCurrent = QLabel()
        self.QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DMM_Instrument = QLabel()
        self.QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        self.QLabel_DMM_VisaAddressforVoltage.setText("Visa Address (DMM for Measure Voltage):")
        self.QLabel_DMM_VisaAddressforCurrent.setText("Visa Address (DMM for Measure Current Shunt):")
        self.QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        #QLineEdit_PSU_VisaAddress = QLineEdit()
        #QLineEdit_DMM_VisaAddress = QLineEdit()
        #QLineEdit_ELoad_VisaAddress = QLineEdit()

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddressforVoltage = QComboBox()
        self.QLineEdit_DMM_VisaAddressforCurrent = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        QComboBox_DMM_Instrument = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        QLabel_set_PSU_Channel.setText("Mode(PSU):")
        QLabel_set_ELoad_Channel.setText("Mode(Eload):")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Programming_Error_Gain.setText("Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Readback Desired Specification (Offset):")

        QComboBox_Voltage_Res = QComboBox()
        #QLineEdit_ELoad_Display_Channel = QLineEdit()
        #QLineEdit_PSU_Display_Channel = QLineEdit()
        QComboBox_set_PSU_Channel = QComboBox()
        QComboBox_set_ELoad_Channel = QComboBox()
        QComboBox_set_Function = QComboBox()
        QComboBox_Voltage_Sense = QComboBox()
        QLineEdit_Programming_Error_Gain = QLineEdit()
        QLineEdit_Programming_Error_Offset = QLineEdit()
        QLineEdit_Readback_Error_Gain = QLineEdit()
        QLineEdit_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_DMM_VisaAddressforVoltage.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_DMM_VisaAddressforCurrent.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])


        QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
                "Power Priority,"
            ]
        )
        QComboBox_set_Function.setEnabled(True)

        QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        QComboBox_set_PSU_Channel.setEnabled(True)

        QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        QComboBox_set_ELoad_Channel.setEnabled(True)
        
        QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])
        QComboBox_Voltage_Sense.setEnabled(True)

        #Power
        QLabel_Power = QLabel()
        QLabel_Power.setText("Rated Power (W):")
        QLineEdit_Power = QLineEdit()
        
        QLabel_PowerINI = QLabel()
        QLabel_PowerINI.setText("Initial Power (W):")
        QLineEdit_PowerINI = QLineEdit()

        QLabel_PowerFIN = QLabel()
        QLabel_PowerFIN.setText("Final Power (W):")
        QLineEdit_PowerFIN = QLineEdit()


        QLabel_power_step_size = QLabel()
        QLabel_power_step_size.setText("Step Size:")
        QLineEdit_power_step_size = QLineEdit()
        
        QLabel_rshunt = QLabel()
        QLabel_rshunt.setText("Shunt Resistance Value (ohm):")
        QLineEdit_rshunt = QLineEdit()



        # Current Sweep
        QLabel_minCurrent = QLabel()
        QLabel_maxCurrent = QLabel()
        QLabel_current_step_size = QLabel()
        QLabel_minCurrent.setText("Initial Current (A):")
        QLabel_maxCurrent.setText("Final Current (A):")
        QLabel_current_step_size.setText("Step Size:")

        QLineEdit_minCurrent = QLineEdit()
        QLineEdit_maxCurrent = QLineEdit()
        QLineEdit_current_stepsize = QLineEdit()

        # Voltage Sweep
        QLabel_minVoltage = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_step_size = QLabel()
        #QLabel_minVoltage.setText("Initial Voltage (V):")
        QLabel_maxVoltage.setText("Final Voltage (V): By default VMAX")
        #QLabel_voltage_step_size.setText("Step Size:")

        QLineEdit_minVoltage = QLineEdit()
        QLineEdit_maxVoltage = QLineEdit()
        QLineEdit_voltage_stepsize = QLineEdit()

        #Loop
        QLabel_noofloop = QLabel()
        QLabel_noofloop.setText("No. of Data Collection:")
        QComboBox_noofloop = QComboBox()
        QComboBox_noofloop.addItems(["1","2","3","4","5","6","7","8","9","10"])

        QLabel_updatedelay = QLabel()
        QLabel_updatedelay.setText("Delay Time (second) :(Default=50ms)")
        QComboBox_updatedelay = QComboBox()
        QComboBox_updatedelay.addItems(["0.0","0.8","1.0","2.0","3.0", "4.0"])
        
        # Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QHBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
        #save_path_layout.addWidget(QLineEdit_Save_Path)  # QLineEdit for the path
        save_path_layout.addWidget(QCheckBox_Report_Widget)  # Checkbox for "Generate Excel Report"
        save_path_layout.addWidget(QCheckBox_Image_Widget)  # Checkbox for "Show Graph"

        # Add the combined layout to the main layout
        layout1.addRow(save_path_layout)

        # Rest of your layout remains unchanged
        layout1.addRow(self.OutputBox)
        #layout1.addRow(Desp0)
        layout1.addRow(QPushButton_Widget0)

        layout1.addRow(Desp1)
        layout1.addRow(QPushButton_Widget4)
        layout1.addRow(self.QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        layout1.addRow(self.QLabel_DMM_VisaAddressforVoltage, self.QLineEdit_DMM_VisaAddressforVoltage)
        layout1.addRow(self.QLabel_DMM_VisaAddressforCurrent, self.QLineEdit_DMM_VisaAddressforCurrent)
        layout1.addRow(self.QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        layout1.addRow(QLabel_DMM_Instrument, QComboBox_DMM_Instrument)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_set_PSU_Channel, QComboBox_set_PSU_Channel)
        layout1.addRow(QLabel_set_ELoad_Channel, QComboBox_set_ELoad_Channel)
        layout1.addRow(QLabel_set_Function, QComboBox_set_Function)
        layout1.addRow(QLabel_Voltage_Sense, QComboBox_Voltage_Sense)
        layout1.addRow(QLabel_Programming_Error_Gain, QLineEdit_Programming_Error_Gain)
        layout1.addRow(QLabel_Programming_Error_Offset, QLineEdit_Programming_Error_Offset)
        layout1.addRow(QLabel_Readback_Error_Gain, QLineEdit_Readback_Error_Gain)
        layout1.addRow(QLabel_Readback_Error_Offset, QLineEdit_Readback_Error_Offset)

        layout1.addRow(Desp6)
        layout1.addRow(QLabel_Power, QLineEdit_Power)
        layout1.addRow(QLabel_PowerINI, QLineEdit_PowerINI)
        layout1.addRow(QLabel_PowerFIN, QLineEdit_PowerFIN)
        layout1.addRow(QLabel_power_step_size, QLineEdit_power_step_size)

        layout1.addRow(Desp7)
        layout1.addRow(QLabel_rshunt, QLineEdit_rshunt)

        layout1.addRow(Desp3)
        #layout1.addRow(QLabel_minVoltage, QLineEdit_minVoltage)
        layout1.addRow(QLabel_maxVoltage, QLineEdit_maxVoltage)
        #layout1.addRow(QLabel_voltage_step_size, QLineEdit_voltage_stepsize)
        layout1.addRow(Desp4)
        #layout1.addRow(QLabel_minCurrent, QLineEdit_minCurrent)
        layout1.addRow(QLabel_maxCurrent, QLineEdit_maxCurrent)
        #layout1.addRow(QLabel_current_step_size, QLineEdit_current_stepsize)
        layout1.addRow(Desp5)
        layout1.addRow(QLabel_noofloop, QComboBox_noofloop)
        layout1.addRow(QLabel_updatedelay, QComboBox_updatedelay)

        layout1.addRow(QPushButton_Widget3)
        layout1.addRow(QPushButton_Widget2)
        layout1.addRow(QPushButton_Widget1)        
        
        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)

        #Set window and scroll area
        self.setLayout(layout1)
        scroll_area(self,layout1)
        
        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddressforVoltage.currentTextChanged.connect(self.DMM_VisaAddressforVoltage_changed)
        self.QLineEdit_DMM_VisaAddressforCurrent.currentTextChanged.connect(self.DMM_VisaAddressforCurrent_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)

        #QLineEdit_ELoad_Display_Channel.textEdited.connect(self.ELoad_Channel_changed)
        #QLineEdit_PSU_Display_Channel.textEdited.connect(self.PSU_Channel_changed)
        QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)

        QLineEdit_Power.textEdited.connect(self.Power_changed)
        QLineEdit_PowerINI.textEdited.connect(self.PowerINI_changed)
        QLineEdit_PowerFIN.textEdited.connect(self.PowerFIN_changed)
        QLineEdit_power_step_size.textEdited.connect(self.power_step_size_changed)

        QLineEdit_rshunt.textEdited.connect(self.rshunt_changed)

        QLineEdit_minVoltage.textEdited.connect(self.minVoltage_changed)
        QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)
        QLineEdit_minCurrent.textEdited.connect(self.minCurrent_changed)
        QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        QLineEdit_voltage_stepsize.textEdited.connect(self.voltage_step_size_changed)
        QLineEdit_current_stepsize.textEdited.connect(self.current_step_size_changed)
        
        QComboBox_set_PSU_Channel.currentTextChanged.connect(self.set_PSU_Channel_changed)
        QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)

        QComboBox_noofloop.currentTextChanged.connect(self.noofloop_changed)
        QComboBox_updatedelay.currentTextChanged.connect(self.updatedelay_changed)

        QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)
        QCheckBox_Report_Widget.stateChanged.connect(self.checkbox_state_Report)
        QCheckBox_Image_Widget.stateChanged.connect(self.checkbox_state_Image)

        QPushButton_Widget0.clicked.connect(self.savepath)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget3.clicked.connect(self.estimateTime)
        QPushButton_Widget4.clicked.connect(self.doFind)
     

    def set_PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def ELoad_Channel_changed(self, s):
       self.ELoad_Channel = s 

    def Power_changed(self, value):
        self.Power = value
    
    def PowerINI_changed(self, value):
        self.powerini= value
    
    def PowerFIN_changed(self, value):
        self.powerfin = value

    def power_step_size_changed(self, value):
        self.power_step_size = value

    def rshunt_changed(self, value):
        self.rshunt = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddressforVoltage.clear()
            self.QLineEdit_DMM_VisaAddressforCurrent.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = GetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddressforVoltage.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddressforCurrent.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])
                
        except:
            self.OutputBox.append("No Devices Found!!!")
        return   



    def updatedelay_changed(self, value):
        self.updatedelay = value
        self.OutputBox.append(str(self.updatedelay))

    def noofloop_changed(self, value):
        self.noofloop = value
        self.OutputBox.append(str(self.noofloop))

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s    

    def DMM_VisaAddressforVoltage_changed(self, s):
        self.DMM = s
    
    def DMM_VisaAddressforCurrent_changed(self, s):
        self.DMM2 = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s

    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s

    def minVoltage_changed(self, s):
        self.minVoltage = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def minCurrent_changed(self, s):
        self.minCurrent = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def voltage_step_size_changed(self, s):
        self.voltage_step_size = s

    def current_step_size_changed(self, s):
        self.current_step_size = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"
        
        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.VoltageRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setInputZ(self, value):
        AdvancedSettingsList[3] = value

    def checkbox_state_Report(self, s):
        self.checkbox_data_Report = s

    def checkbox_state_Image(self, s):
        self.checkbox_data_Image = s

    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value

    def openDialog(self):
        dlg = AdvancedSetting_Voltage()
        dlg.exec()

    def estimateTime(self):
        
        self.OutputBox.append(str(self.updatedelay))
           
        self.powerloop = ((float(self.powerfin) - float(self.powerini))/ float(self.power_step_size)) + 1

        if self.updatedelay == 0.0:
            constant = 0
                
        else:
            constant = 1

        self.estimatetime = (self.powerloop *(0.2 + 0.8 + (float(self.updatedelay) * constant) )) * float(self.noofloop)
        self.OutputBox.append(str(self.estimatetime) + "seconds")
    

    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

         

    def executeTest(self):
        global globalvv

        for x in range (int(self.noofloop)):

            """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
            then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
            will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
            begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
            begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
            are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
            optionally export all the details into a CSV file or display a graph after the test is completed.

            """
            self.infoList = []
            self.dataList = []
            self.dataList2 = []

            dict = dictGenerator.input(
                P_Rating=self.Power_Rating,
                power=self.Power,
                rshunt = self.rshunt,
                powerini=self.powerini,
                powerfin=self.powerfin,
                power_step_size = self.power_step_size,

                estimatetime=self.estimatetime,
                updatedelay=self.updatedelay,
                readbackvoltage=self.readbackvoltage,
                readbackcurrent=self.readbackcurrent,
                noofloop=self.noofloop,
                Instrument=self.DMM_Instrument,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,
                unit=self.unit,
                #minCurrent=self.minCurrent,
                maxCurrent=self.maxCurrent,
                #current_step_size=self.current_step_size,
                #minVoltage=self.minVoltage,
                maxVoltage=self.maxVoltage,
                #voltage_step_size=self.voltage_step_size,

                PSU=self.PSU,
                DMM=self.DMM,
                DMM2=self.DMM2,
                ELoad=self.ELoad,

                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],
            )

            if x == 0:
                QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
                                    )
                """globalvv = dict["estimatetime"]
                loading_thread = threading.Thread(target=self.tkinter_loading_screen, args=(globalvv,))
                loading_thread.start()"""
                
                
                                      
            for i in [dict]:
                if i == "":
                    QMessageBox.warning(
                        self, "Error", "One of the parameters are not filled in"
                    )
                    break
            else:
                A = VisaResourceManager()
                flag, args = A.openRM(self.ELoad, self.PSU, self.DMM, self.DMM2)

                if flag == 0:
                    string = ""
                    for item in args:
                        string = string + item

                    QMessageBox.warning(self, "VISA IO ERROR", string)
                    exit()

                if self.setFunction == "Current":
                    try:(
                        infoList,
                        dataList,
                        dataList2
                        ) = PowerMeasurement.executePowerMeasurementA(self, dict)
                    
                    
                    except Exception as e:
                        QMessageBox.warning(self, "Error", str(e))
                        exit()

                if self.setFunction == "Voltage":
                    try:(
                        infoList,
                        dataList,
                        dataList2
                        ) = PowerMeasurement.executePowerMeasurementB(self, dict)
                    
                    
                    except Exception as e:
                        QMessageBox.warning(self, "Error", str(e))
                        exit()

                if x == (int(self.noofloop) - 1):   
                    self.OutputBox.append(my_result.getvalue())
                    self.OutputBox.append("Measurement is complete !")


                if self.checkbox_data_Report == 2:
                    powerinstrumentData(self.PSU, self.DMM, self.DMM2, self.ELoad)
                    datatoCSV_PowerAccuracy(infoList, dataList, dataList2)
                    datatoGraph3(infoList, dataList,dataList2)
                    datatoGraph3.scatterComparePower(self, float(self.Programming_Error_Gain), float(self.Programming_Error_Offset), float(self.Readback_Error_Gain), float(self.Readback_Error_Offset), str(self.setFunction), float(self.Power_Rating) )

                    if self.setFunction == "Current":
                        A = xlreportpower(save_directory=self.savelocation, file_name="Power_CC")
                        A.run()
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        df.to_csv("C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerconfig.csv")

                    elif self.setFunction == "Voltage":
                        A = xlreportpower(save_directory=self.savelocation, file_name="Power_CV")
                        A.run()
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        df.to_csv("C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerconfig.csv")


                if x == (int(self.noofloop) - 1):  
                    if self.checkbox_data_Image == 2:
                        self.image_dialog = image_Window2()
                        self.image_dialog.setModal(True)
                        self.image_dialog.show()
    
    def tkinter_loading_screen(self, x):
            def stop_loading():
                progress_bar.stop()
                loading_label.config(text="Loading Complete!")

            def update_countdown(remaining_time):
                if remaining_time > 0:
                    loading_label.config(text=f"Loading... {remaining_time} seconds remaining")
                    root.after(1000, update_countdown, remaining_time - 1)
                else:
                    stop_loading()

            def start_loading():
                progress_bar.start()
                update_countdown(float(x))  # Start the countdown

            sleep(2)
            root = tk.Tk()
            root.title("Loading Screen")

            loading_label = tk.Label(root, text="Loading...", font=("Helvetica", 16))
            loading_label.pack(pady=20)

            progress_bar = ttk.Progressbar(root, orient="horizontal", mode="indeterminate", length=280)
            progress_bar.pack(pady=20)
            start_loading()

            root.mainloop()

class BundleMeasurementVoltageDialog(QDialog):
    """Class for configuring the voltage measurement DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.


    """

    def __init__(self):
        """Method where Widgets, Signals and Slots are defined in the GUI for Voltage Measurement"""
        super().__init__()
        # Initialize default CSV path
        self.DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/data.csv"  # or use a dynamic path if needed
        self.IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/Chart.png"
        self.ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/error.csv"
        self.image_dialog = None
        
        # Default Values
        #self.PSU_VisaAddress = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        #self.DMM_VisaAddress = "USB0::0x2A8D::0x1601::MY60094787::0::INSTR"
        #self.ELoad_VisaAddress = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        
        self.Power_Rating = "6000"
        self.Current_Rating = "24"
        self.Voltage_Rating = "800"
        self.Power = "2000"
        self.currloop = "1"
        self.voltloop = "1"
        self.estimatetime = "0"
        self.readbackvoltage = "0"
        self.readbackcurrent = "0"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"
        self.noofloop = "1"
        self.updatedelay = "3.0"
        self.unit = "VOLTAGE"
        self.Programming_Error_Offset = "0.0003"
        self.Programming_Error_Gain = "0.0003"
        self.Readback_Error_Offset = "0.0003"
        self.Readback_Error_Gain = "0.0003"
        self.minCurrent = "0"
        self.maxCurrent = "24"
        self.current_step_size = "1"
        self.minVoltage = "0"
        self.maxVoltage = "800"
        self.voltage_step_size = "80"
        self.PSU = "USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"
        self.OSC = "USB0::0x0957::0x17B0::MY52060151::0::INSTR"
        self.DMM = "USB0::0x2A8D::0x1601::MY60094787::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"
        self.OSC_Channel = "1"
        self.DMM_Instrument = "Keysight"
        self.DUT = "Others"

        self.PSU_Channel = "1"
        self.ELoad_Channel = "1"
        self.setFunction = "Current" #Set Eload in CC Mode
        self.VoltageRes = "SLOW"
        self.VoltageSense = "INT"

        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.checkbox_test_VoltageAccuracy = 2
        self.checkbox_test_VoltageLoadRegulation = 2
        self.checkbox_test_TransientRecovery = 2

        self.Range = "Auto"
        self.Aperture = "0.2"
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.UpTime = "50"
        self.DownTime = "50"

        self.Channel_CouplingMode = "AC"
        self.Trigger_CouplingMode = "DC"
        self.Trigger_Mode = "EDGE"
        self.Trigger_SweepMode = "NORMAL"
        self.Trigger_SlopeMode = "EITHer"
        self.TimeScale = "0.01"
        self.VerticalScale = "0.00001"
        self.I_Step = ""
        self.V_Settling_Band = "0.8"
        self.T_Settling_Band = "0.001"
        self.Probe_Setting = "X10"
        self.Acq_Type = "AVERage"

        self.checkbox_SpecialCase = 2
        self.checkbox_NormalCase = 2
        
        # Ensure DATA_CSV_PATH is defined before use
        if not hasattr(self, 'DATA_CSV_PATH') or not self.DATA_CSV_PATH:
            QMessageBox.warning(self, "Error", "CSV path is not set.")

        self.setWindowTitle("Bundle Measurement")
        # Enable maximize button
        self.setWindowFlags(Qt.Window)

        # Allow resizing
        # self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        # self.setMinimumSize(800, 600)
        self.image_window = None
        self.setStyleSheet("font-size: 15px;")

        # create find button 
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget3 = QPushButton()
        QPushButton_Widget3.setText("Estimate Data Collection Time")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
        QCheckBox_Report_Widget = QCheckBox()
        QCheckBox_Report_Widget.setText("Generate Excel Report")
        QCheckBox_Report_Widget.setCheckState(Qt.Checked)
        QCheckBox_Image_Widget = QCheckBox()
        QCheckBox_Image_Widget.setText("Show Graph")
        QCheckBox_Image_Widget.setCheckState(Qt.Checked)
        QCheckBox_SpecialCase_Widget = QCheckBox()
        QCheckBox_SpecialCase_Widget.setText("Special Case (0% <-> 100%)")
        QCheckBox_SpecialCase_Widget.setCheckState(Qt.Checked)
        QCheckBox_NormalCase_Widget = QCheckBox()
        QCheckBox_NormalCase_Widget.setText("Normal Case (50% <-> 100%)")
        QCheckBox_NormalCase_Widget.setCheckState(Qt.Checked)

        #Test checkbox
        QCheckBox_VoltageAccuracy_Widget = QCheckBox()
        QCheckBox_VoltageAccuracy_Widget.setText("Voltage Accuracy")
        QCheckBox_VoltageAccuracy_Widget.setCheckState(Qt.Checked)
        QCheckBox_VoltageLoadRegulation_Widget = QCheckBox()
        QCheckBox_VoltageLoadRegulation_Widget.setText("Voltage Load Regulation")
        QCheckBox_VoltageLoadRegulation_Widget.setCheckState(Qt.Checked)
        QCheckBox_TransientRecovery_Widget = QCheckBox()
        QCheckBox_TransientRecovery_Widget.setText("Transient Recovery")
        QCheckBox_TransientRecovery_Widget.setCheckState(Qt.Checked)

        layout1 = QFormLayout()
        self.OutputBox = QTextBrowser()
        self.OutputBox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.OutputBox.setMinimumHeight(50)  # Start small
        self.OutputBox.setMaximumHeight(200)
        self.OutputBox.append(my_result.getvalue())
        self.setLayout(layout1)

        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()
        Desp6 = QLabel()
        Desp7 = QLabel()
        PerformTest = QLabel()
        OscilloscopeSetting = QLabel()
        Desp8 = QLabel()

        Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)
        Desp6.setFont(desp_font)
        Desp7.setFont(desp_font)
        PerformTest.setFont(desp_font)
        OscilloscopeSetting.setFont(desp_font)
        #Desp8.setFont(desp_font)

        Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Voltage Sweep:")
        Desp4.setText("Current Sweep:")
        Desp5.setText("No. of Collection:")
        Desp6.setText("Rated Power of DUT Setting [W]:")
        Desp7.setText("Maximum Testing Current")
        PerformTest.setText("Perform Test:")
        OscilloscopeSetting.setText("Oscilloscope Setting:")
        #Desp8.setText("Testing Selection:")

        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        #Testing Selection
        QLabel_Testing_Selection = QLabel()
        QLabel_Testing_Selection.setFont(desp_font)
        QLabel_Testing_Selection.setText("Test:")

        #Find Instruments
        # Connections
        self.QLabel_PSU_VisaAddress = QLabel()
        self.QLabel_DMM_VisaAddress = QLabel()
        self.QLabel_OSC_VisaAddress = QLabel()
        self.QLabel_ELoad_VisaAddress = QLabel()
        self.QLabel_DUT = QLabel()
        QLabel_DMM_Instrument = QLabel()
        self.QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        self.QLabel_DMM_VisaAddress.setText("Visa Address (DMM):")
        self.QLabel_OSC_VisaAddress.setText("Visa Address (OSC):")
        self.QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        self.QLabel_DUT.setText("DUT:")

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddress = QComboBox()
        self.QLineEdit_OSC_VisaAddress = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        QComboBox_DMM_Instrument = QComboBox()
        self.QLineEdit_DUT = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel() #PSU
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        QLabel_set_PSU_Channel.setText("Set PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Set Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")
        QLabel_Programming_Error_Gain.setText("Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Readback Desired Specification (Offset):")

        QComboBox_Voltage_Res = QComboBox()  
        QComboBox_set_PSU_Channel = QComboBox()
        QComboBox_set_ELoad_Channel = QComboBox()
        QComboBox_set_Function = QComboBox()
        QComboBox_Voltage_Sense = QComboBox()
        self.QLineEdit_Programming_Error_Gain = QLineEdit()
        self.QLineEdit_Programming_Error_Offset = QLineEdit()
        self.QLineEdit_Readback_Error_Gain = QLineEdit()
        self.QLineEdit_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_OSC_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_DMM_VisaAddress.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])
        self.QLineEdit_DUT.addItems(["Others", "Excavator"])

        QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        QComboBox_DMM_Instrument.setEnabled(False)
        QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        QComboBox_set_Function.setEnabled(False)
        QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        QComboBox_set_PSU_Channel.setEnabled(True)
        QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        QComboBox_set_ELoad_Channel.setEnabled(True)
        QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])
        QComboBox_Voltage_Sense.setEnabled(True)

        #Power for test
        QLabel_Power = QLabel()
        QLabel_Power.setText("Power Set for Test(W):")
        self.QLineEdit_Power = QLineEdit()

        #Rated Power
        QLabel_power_rated = QLabel()
        QLabel_power_rated.setText("Rated Power of DUT (W):")
        self.QLineEdit_power_rated = QLineEdit()

        # Current Sweep
        QLabel_minCurrent = QLabel()
        QLabel_maxCurrent = QLabel()
        QLabel_current_step_size = QLabel()
        QLabel_current_rated = QLabel()
        QLabel_minCurrent.setText("Initial Current (A):")
        QLabel_maxCurrent.setText("Final Current (A):")
        QLabel_current_step_size.setText("Step Size:")
        QLabel_current_rated.setText("DUT Rated Current:")

        self.QLineEdit_minCurrent = QLineEdit()
        self.QLineEdit_maxCurrent = QLineEdit()
        self.QLineEdit_current_stepsize = QLineEdit()
        self.QLineEdit_current_rated = QLineEdit()

        # Voltage Sweep
        QLabel_minVoltage = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_step_size = QLabel()
        QLabel_voltage_rated = QLabel()
        QLabel_minVoltage.setText("Initial Voltage (V):")
        QLabel_maxVoltage.setText("Final Voltage (V):")
        QLabel_voltage_step_size.setText("Step Size:")
        QLabel_voltage_rated.setText("DUT Rated Voltage:")

        self.QLineEdit_minVoltage = QLineEdit()
        self.QLineEdit_maxVoltage = QLineEdit()
        self.QLineEdit_voltage_stepsize = QLineEdit()
        self.QLineEdit_voltage_rated = QLineEdit()
        self.QLineEdit_voltage_rated.setFixedSize(100, 40)  # Set fixed size

        # Oscilloscope Settings
        QLabel_OSC_Display_Channel = QLabel()
        QLabel_V_Settling_Band = QLabel()
        QLabel_T_Settling_Band = QLabel()
        QLabel_Probe_Setting = QLabel()
        QLabel_Acq_Type = QLabel()
        QLabel_OSC_Display_Channel.setText("Display Channel (OSC)")
        QLabel_V_Settling_Band.setText("Settling Band Voltage (V) / Error Band:")
        QLabel_T_Settling_Band.setText("Settling Band Time (s):")
        QLabel_Probe_Setting.setText("Probe Setting Ratio:")
        QLabel_Acq_Type.setText("Acquire Mode:")
        QLineEdit_OSC_Display_Channel = QLineEdit()
        QLineEdit_V_Settling_Band = QLineEdit()
        QLineEdit_T_Settling_Band = QLineEdit()
        QComboBox_Probe_Setting = QComboBox()
        QComboBox_Acq_Type = QComboBox()
        QComboBox_Probe_Setting.addItems(["X1", "X10", "X20", "X100"])
        QComboBox_Acq_Type.addItems(["NORMal", "PEAK", "AVERage", "HRESolution"])
        QLabel_Channel_CouplingMode = QLabel()
        QLabel_Trigger_Mode = QLabel()
        QLabel_Trigger_CouplingMode = QLabel()
        QLabel_Trigger_SweepMode = QLabel()
        QLabel_Trigger_SlopeMode = QLabel()
        QLabel_TimeScale = QLabel()
        QLabel_VerticalScale = QLabel()

        QLabel_Channel_CouplingMode.setText("Coupling Mode (Channel)")
        QLabel_Trigger_Mode.setText("Trigger Mode:")
        QLabel_Trigger_CouplingMode.setText("Coupling Mode (Trigger):")
        QLabel_Trigger_SweepMode.setText("Trigger Sweep Mode:")
        QLabel_Trigger_SlopeMode.setText("Trigger Slope Mode:")
        QLabel_TimeScale.setText("Time Scale:")
        QLabel_VerticalScale.setText("Vertical Scale:")

        QComboBox_Channel_CouplingMode = QComboBox()
        QComboBox_Trigger_Mode = QComboBox()
        QComboBox_Trigger_CouplingMode = QComboBox()
        QComboBox_Trigger_SweepMode = QComboBox()
        QComboBox_Trigger_SlopeMode = QComboBox()
        QLineEdit_TimeScale = QLineEdit()
        QLineEdit_VerticalScale = QLineEdit()

        QComboBox_Channel_CouplingMode.addItems(["AC", "DC"])
        QComboBox_Trigger_Mode.addItems(["EDGE", "IIC", "EBUR"])
        QComboBox_Trigger_CouplingMode.addItems(["AC", "DC"])
        QComboBox_Trigger_SweepMode.addItems(["NORMAL", "AUTO"])
        QComboBox_Trigger_SlopeMode.addItems(["ALT", "POS", "NEG", "EITH"])

        QComboBox_Channel_CouplingMode.setEnabled(True)
        QComboBox_Trigger_Mode.setEnabled(True)
        QComboBox_Trigger_CouplingMode.setEnabled(True)
        QComboBox_Trigger_SweepMode.setEnabled(True)
        QComboBox_Trigger_SlopeMode.setEnabled(True)
        QComboBox_Probe_Setting.setEnabled(True)
        QComboBox_Acq_Type.setEnabled(True)

        #Num of Loop/Delay Time
        QLabel_noofloop = QLabel()
        QLabel_noofloop.setText("No. of Data Collection:")
        QComboBox_noofloop = QComboBox()
        QComboBox_noofloop.addItems(["1","2","3","4","5","6","7","8","9","10"])
        QLabel_updatedelay = QLabel()
        QLabel_updatedelay.setText("Delay Time (second) :(Default=50ms)")
        QComboBox_updatedelay = QComboBox()
        QComboBox_updatedelay.addItems(["0.0","0.8","1.0","2.0","3.0", "4.0"])
            
        # Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QHBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
        #save_path_layout.addWidget(QLineEdit_Save_Path)  # QLineEdit for the path
        save_path_layout.addWidget(QCheckBox_Report_Widget)  # Checkbox for "Generate Excel Report"
        save_path_layout.addWidget(QCheckBox_Image_Widget)  # Checkbox for "Show Graph"

        # Create a horizontal layout for the "Save Path" and checkboxes
        testing_selection_layout = QHBoxLayout()
        testing_selection_layout.addWidget(QLabel_Testing_Selection)  # QLabel for "Save Path"
        testing_selection_layout.addWidget(QCheckBox_VoltageAccuracy_Widget)  # Checkbox for "Generate Excel Report"
        testing_selection_layout.addWidget(QCheckBox_VoltageLoadRegulation_Widget)  # Checkbox for "Show Graph"
        testing_selection_layout.addWidget(QCheckBox_TransientRecovery_Widget) 

        voltage_sweep_layout = QHBoxLayout()
        voltage_sweep_layout.addWidget(Desp3)  # QLabel for "Save Path"
        voltage_sweep_layout.addWidget(QLabel_voltage_rated)  # Checkbox for "Generate Excel Report"
        voltage_sweep_layout.addWidget(self.QLineEdit_voltage_rated )  # Checkbox for "Show Graph"

        current_sweep_layout = QHBoxLayout()
        current_sweep_layout.addWidget(Desp4)  # QLabel for "Save Path"
        current_sweep_layout.addWidget(QLabel_current_rated)  # Checkbox for "Generate Excel Report"
        current_sweep_layout.addWidget(self.QLineEdit_current_rated )  # Checkbox for "Show Graph"

        power_sweep_layout = QHBoxLayout()
        power_sweep_layout.addWidget(Desp6)  # QLabel for "Save Path"
        power_sweep_layout.addWidget(QLabel_power_rated)  # Checkbox for "Generate Excel Report"
        power_sweep_layout.addWidget(self.QLineEdit_power_rated)  # Checkbox for "Show Graph"

        voltage_inifin_layout = QHBoxLayout()
        voltage_inifin_layout.addWidget(QLabel_minVoltage)  # QLabel for "Save Path"
        voltage_inifin_layout.addWidget(self.QLineEdit_minVoltage)  # Checkbox for "Generate Excel Report"
        voltage_inifin_layout.addWidget(QLabel_maxVoltage)  # Checkbox for "Show Graph"
        voltage_inifin_layout.addWidget(self.QLineEdit_maxVoltage)  # Checkbox for "Show Graph"

        current_inifin_layout = QHBoxLayout()
        current_inifin_layout.addWidget(QLabel_minCurrent)  # QLabel for "Save Path"
        current_inifin_layout.addWidget(self.QLineEdit_minCurrent)  # Checkbox for "Generate Excel Report"
        current_inifin_layout.addWidget(QLabel_maxCurrent)  # Checkbox for "Show Graph"
        current_inifin_layout.addWidget(self.QLineEdit_maxCurrent)  # Checkbox for "Show Graph"

        programming_error_layout = QHBoxLayout()
        programming_error_layout.addWidget(QLabel_Programming_Error_Gain)  # QLabel for "Save Path"
        programming_error_layout.addWidget(self.QLineEdit_Programming_Error_Gain)  # Checkbox for "Generate Excel Report"
        programming_error_layout.addWidget(QLabel_Programming_Error_Offset)  # Checkbox for "Show Graph"
        programming_error_layout.addWidget(self.QLineEdit_Programming_Error_Offset)  # Checkbox for "Show Graph"

        readback_error_layout = QHBoxLayout()
        readback_error_layout.addWidget(QLabel_Readback_Error_Gain)  # QLabel for "Save Path"
        readback_error_layout.addWidget(self.QLineEdit_Readback_Error_Gain)  # Checkbox for "Generate Excel Report"
        readback_error_layout.addWidget(QLabel_Readback_Error_Offset)  # Checkbox for "Show Graph"
        readback_error_layout.addWidget(self.QLineEdit_Readback_Error_Offset)  # Checkbox for "Show Graph"

        performtest_layout = QHBoxLayout()
        performtest_layout.addWidget(Desp5)  # QLabel for "Save Path"
        performtest_layout.addWidget(QCheckBox_SpecialCase_Widget)  # Checkbox for "Generate Excel Report"
        performtest_layout.addWidget(QCheckBox_NormalCase_Widget)  # Checkbox for "Show Graph"
    

        layout1.setSpacing(5)  # Reduce spacing
        layout1.setContentsMargins(10, 10, 10, 10)  # Reduce margins
        # Add the combined layout to the main layout
        layout1.addRow(save_path_layout)

        # Rest of your layout remains unchanged
        layout1.addRow(self.OutputBox)
        #layout1.addRow(Desp0)
        layout1.addRow(QPushButton_Widget0)
        layout1.addRow(Desp1)
        layout1.addRow(QPushButton_Widget4)
        layout1.addRow(self.QLabel_DUT, self.QLineEdit_DUT)
        layout1.addRow(self.QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        layout1.addRow(self.QLabel_DMM_VisaAddress, self.QLineEdit_DMM_VisaAddress)
        layout1.addRow(self.QLabel_OSC_VisaAddress, self.QLineEdit_OSC_VisaAddress)
        layout1.addRow(self.QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        layout1.addRow(QLabel_DMM_Instrument, QComboBox_DMM_Instrument)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_set_PSU_Channel, QComboBox_set_PSU_Channel)
        layout1.addRow(QLabel_set_ELoad_Channel, QComboBox_set_ELoad_Channel)
        layout1.addRow(QLabel_set_Function, QComboBox_set_Function)
        layout1.addRow(QLabel_Voltage_Sense, QComboBox_Voltage_Sense)
        layout1.addRow(programming_error_layout)
        layout1.addRow(readback_error_layout)
        #layout1.addRow(QLabel_Programming_Error_Gain, QLineEdit_Programming_Error_Gain)
        #layout1.addRow(QLabel_Programming_Error_Offset, QLineEdit_Programming_Error_Offset)
        #layout1.addRow(QLabel_Readback_Error_Gain, QLineEdit_Readback_Error_Gain)
        #layout1.addRow(QLabel_Readback_Error_Offset, QLineEdit_Readback_Error_Offset)
        #layout1.addRow(Desp6)
        layout1.addRow(power_sweep_layout)
        layout1.addRow(QLabel_Power, self.QLineEdit_Power)
        layout1.addRow(voltage_sweep_layout)
        layout1.addRow(voltage_inifin_layout)
        #layout1.addRow(QLabel_minVoltage, QLineEdit_minVoltage)
        #layout1.addRow(QLabel_maxVoltage, QLineEdit_maxVoltage)
        layout1.addRow(QLabel_voltage_step_size, self.QLineEdit_voltage_stepsize)
        layout1.addRow(current_sweep_layout)
        layout1.addRow(current_inifin_layout)
        #layout1.addRow(QLabel_minCurrent, QLineEdit_minCurrent)
        #layout1.addRow(QLabel_maxCurrent, QLineEdit_maxCurrent)
        layout1.addRow(QLabel_current_step_size, self.QLineEdit_current_stepsize)
        #layout1.addRow(Desp5)

        layout1.addRow(OscilloscopeSetting)
        layout1.addRow(QLabel_OSC_Display_Channel, QLineEdit_OSC_Display_Channel)
        layout1.addRow(QLabel_V_Settling_Band, QLineEdit_V_Settling_Band)
        layout1.addRow(QLabel_T_Settling_Band, QLineEdit_T_Settling_Band)
        layout1.addRow(QLabel_Probe_Setting, QComboBox_Probe_Setting)
        layout1.addRow(QLabel_Acq_Type, QComboBox_Acq_Type)
        layout1.addRow(QLabel_Channel_CouplingMode, QComboBox_Channel_CouplingMode)
        layout1.addRow(QLabel_Trigger_CouplingMode, QComboBox_Trigger_CouplingMode)
        layout1.addRow(QLabel_Trigger_Mode, QComboBox_Trigger_Mode)
        layout1.addRow(QLabel_Trigger_SweepMode, QComboBox_Trigger_SweepMode)
        layout1.addRow(QLabel_Trigger_SlopeMode, QComboBox_Trigger_SlopeMode)
        layout1.addRow(QLabel_TimeScale, QLineEdit_TimeScale)
        layout1.addRow(QLabel_VerticalScale, QLineEdit_VerticalScale)
        layout1.addRow(performtest_layout)
        layout1.addRow(QLabel_noofloop, QComboBox_noofloop)
        layout1.addRow(QLabel_updatedelay, QComboBox_updatedelay)
        layout1.addRow(testing_selection_layout)
        layout1.addRow(QPushButton_Widget3)
        layout1.addRow(QPushButton_Widget2)
        layout1.addRow(QPushButton_Widget1)        

        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)

        #Set window and scroll area
        self.setLayout(layout1)
        scroll_area(self,layout1)

        QLineEdit_V_Settling_Band.textEdited.connect(self.V_Settling_Band_changed)
        QLineEdit_T_Settling_Band.textEdited.connect(self.T_Settling_Band_changed)
        QLineEdit_OSC_Display_Channel.textEdited.connect(self.OSC_Channel_changed)
        QComboBox_Channel_CouplingMode.currentTextChanged.connect(self.Channel_CouplingMode_changed )
        QComboBox_Trigger_CouplingMode.currentTextChanged.connect(
            self.Trigger_CouplingMode_changed
        )
        QComboBox_Trigger_Mode.currentTextChanged.connect(self.Trigger_Mode_changed)
        QComboBox_Trigger_SweepMode.currentTextChanged.connect(
            self.Trigger_SweepMode_changed
        )
        QComboBox_Trigger_SlopeMode.currentTextChanged.connect(
            self.Trigger_SlopeMode_changed
        )
        QComboBox_Probe_Setting.currentTextChanged.connect(
            self.Probe_Setting_changed
        )
        QComboBox_Acq_Type.currentTextChanged.connect(
            self.Acq_Type_changed
        )
        QLineEdit_TimeScale.textEdited.connect(self.TimeScale_changed)
        QLineEdit_VerticalScale.textEdited.connect(self.VerticalScale_changed)

        QCheckBox_SpecialCase_Widget.stateChanged.connect(self.checkbox_state_SpecialCase)
        QCheckBox_NormalCase_Widget.stateChanged.connect(self.checkbox_state_NormalCase)

        self.QLineEdit_DUT.currentTextChanged.connect(self.DUT_changed)
        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddress.currentTextChanged.connect(self.DMM_VisaAddress_changed)
        self.QLineEdit_OSC_VisaAddress.currentTextChanged.connect(self.OSC_VisaAddress_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)

        self.QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        self.QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        self.QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        self.QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)

        self.QLineEdit_Power.textEdited.connect(self.Power_changed)
        self.QLineEdit_power_rated.textEdited.connect(self.Power_Rating_changed)
        self.QLineEdit_current_rated.textEdited.connect(self.Current_Rating_changed)
        self.QLineEdit_voltage_rated.textEdited.connect(self.Voltage_Rating_changed)
        self.QLineEdit_minVoltage.textEdited.connect(self.minVoltage_changed)
        self.QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)
        self.QLineEdit_minCurrent.textEdited.connect(self.minCurrent_changed)
        self.QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        self.QLineEdit_voltage_stepsize.textEdited.connect(self.voltage_step_size_changed)
        self.QLineEdit_current_stepsize.textEdited.connect(self.current_step_size_changed)
        
        QComboBox_set_PSU_Channel.currentTextChanged.connect(self.set_PSU_Channel_changed)
        QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.ELoad_Channel_changed)
        QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)

        QComboBox_noofloop.currentTextChanged.connect(self.noofloop_changed)
        QComboBox_updatedelay.currentTextChanged.connect(self.updatedelay_changed)

        QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)
        self.QLineEdit_DUT.currentIndexChanged.connect(self.on_current_index_changed)

        QCheckBox_Report_Widget.stateChanged.connect(self.checkbox_state_Report)
        QCheckBox_Image_Widget.stateChanged.connect(self.checkbox_state_Image)
        QCheckBox_VoltageAccuracy_Widget.stateChanged.connect(self.checkbox_state_VoltageAccuracy)
        QCheckBox_VoltageLoadRegulation_Widget.stateChanged.connect(self.checkbox_state_VoltageLoadRegulation)
        QCheckBox_TransientRecovery_Widget.stateChanged.connect(self.checkbox_state_TransientRecovery)
  
        QPushButton_Widget0.clicked.connect(self.savepath)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget3.clicked.connect(self.estimateTime)
        QPushButton_Widget4.clicked.connect(self.doFind)

    def on_current_index_changed(self):
        selected_text = self.QLineEdit_DUT.currentText()
        if selected_text == "Excavator":
            self.QLineEdit_Programming_Error_Gain.setText(self.Programming_Error_Gain)
            self.QLineEdit_Programming_Error_Offset.setText(self.Programming_Error_Offset)
            self.QLineEdit_Readback_Error_Gain.setText(self.Readback_Error_Gain)
            self.QLineEdit_Readback_Error_Offset.setText(self.Readback_Error_Offset)
            self.QLineEdit_power_rated.setText(self.Power_Rating)
            self.QLineEdit_Power.setText(self.Power)
            self.QLineEdit_voltage_rated.setText(self.Voltage_Rating)
            self.QLineEdit_minVoltage.setText(self.minVoltage)
            self.QLineEdit_maxVoltage.setText(self.maxVoltage)
            self.QLineEdit_voltage_stepsize.setText(self.voltage_step_size)
            self.QLineEdit_current_rated.setText(self.Current_Rating)
            self.QLineEdit_minCurrent.setText(self.minCurrent)
            self.QLineEdit_maxCurrent.setText(self.maxCurrent)
            self.QLineEdit_current_stepsize.setText(self.current_step_size)
            # self.QLineEdit_OSC_Display_Channel.setText(self.OSC_Channel)
            # self.QLineEdit_V_Settling_Band.setText(self.V_Settling_Band)
            # self.QLineEdit_T_Settling_Band.setText(self.T_Settling_Band)
            # self.QComboBox_Probe_Setting.setCurrentText(self.Probe_Setting)
            # self.QComboBox_Acq_Type.setCurrentText(self.Acq_Type)
            # self.QComboBox_Channel_CouplingMode.setCurrentText(self.Channel_CouplingMode)
            # self.QComboBox_Trigger_Mode.setCurrentText(self.Trigger_Mode)
            # self.QComboBox_Trigger_CouplingMode.setCurrentText(self.Trigger_CouplingMode)
            # self.QComboBox_Trigger_SweepMode.setCurrentText(self.Trigger_SweepMode)
            # self.QComboBox_Trigger_SlopeMode.setCurrentText(self.Trigger_SlopeMode)
            # self.QLineEdit_TimeScale.setText(self.TimeScale)
            # self.QLineEdit_VerticalScale.setText(self.VerticalScale)
            # self.QComboBox_updatedelay.setCurrentText(self.updatedelay)
        
    def set_PSU_Channel_changed(self, s):
        self.PSU_Channel = s

    def ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s

    def Voltage_Rating_changed(self, value):
        self.Voltage_Rating = value

    def Current_Rating_changed(self, value):
        self.Current_Rating = value

    def Power_Rating_changed(self, value):
        self.Power_Rating = value
     
    def Power_changed(self, value):
        self.Power = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddress.clear()
            self.QLineEdit_OSC_VisaAddress.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            discovery = NewGetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
                
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_OSC_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])

        except:
            self.OutputBox.append("No Devices Found!!!")
        return 

  

    def checkbox_state_SpecialCase(self, s):
        self.checkbox_SpecialCase = s

    def checkbox_state_NormalCase(self, s):
        self.checkbox_NormalCase = s

    def updatedelay_changed(self, value):
        self.updatedelay = value
        self.OutputBox.append(str(self.updatedelay))

    def noofloop_changed(self, value):
        self.noofloop = value
        self.OutputBox.append(str(self.noofloop))

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def DUT_changed(self, s):
        self.DUT = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s    

    def DMM_VisaAddress_changed(self, s):
        self.DMM = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s
    
    def OSC_VisaAddress_changed(self, s):
        self.OSC = s

    def OSC_Channel_changed(self, s):
        self.OSC_Channel = s

    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s

    def minVoltage_changed(self, s):
        self.minVoltage = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def minCurrent_changed(self, s):
        self.minCurrent = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def voltage_step_size_changed(self, s):
        self.voltage_step_size = s

    def current_step_size_changed(self, s):
        self.current_step_size = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.VoltageRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setInputZ(self, value):
        AdvancedSettingsList[3] = value

    def checkbox_state_Report(self, s):
        self.checkbox_data_Report = s

    def checkbox_state_Image(self, s):
        self.checkbox_data_Image = s
    
    def checkbox_state_VoltageAccuracy(self, s):
        self.checkbox_test_VoltageAccuracy = s

    def checkbox_state_VoltageLoadRegulation(self, s):
        self.checkbox_test_VoltageLoadRegulation = s
    
    def checkbox_state_TransientRecovery(self, s):
        self.checkbox_test_TransientRecovery = s

    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value

    def openDialog(self):
        dlg = AdvancedSetting_Voltage()
        dlg.exec()
    
    def T_Settling_Band_changed(self, s):
        self.T_Settling_Band = s

    def V_Settling_Band_changed(self, s):
        self.V_Settling_Band = s

    def Channel_CouplingMode_changed(self, s):
        self.Channel_CouplingMode = s

    def Trigger_CouplingMode_changed(self, s):
        self.Trigger_CouplingMode = s

    def Trigger_Mode_changed(self, s):
        self.Trigger_Mode = s

    def Trigger_SweepMode_changed(self, s):
        self.Trigger_SweepMode = s

    def Trigger_SlopeMode_changed(self, s):
        self.Trigger_SlopeMode = s
    
    def Probe_Setting_changed(self, s):
        self.Probe_Setting = s
    
    def Acq_Type_changed(self, s):
        self.Acq_Type = s

    def TimeScale_changed(self, s):
        self.TimeScale = s

    def VerticalScale_changed(self, s):
        self.VerticalScale = s

    def estimateTime(self):
        
        self.OutputBox.append(str(self.updatedelay))
           
        self.currloop = ((float(self.maxCurrent) - float(self.minCurrent))/ float(self.current_step_size)) + 1
        self.voltloop = ((float(self.maxVoltage) - float(self.minVoltage))/ float(self.voltage_step_size)) + 1

        if self.updatedelay == 0.0:
            constant = 0
                
        else:
            constant = 1

        self.estimatetime = (self.currloop * self.voltloop *(0.2 + 0.8 + (float(self.updatedelay) * constant) )) * float(self.noofloop)
        self.OutputBox.append(str(self.estimatetime) + "seconds")
    


    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

         
    

    def executeTest(self):
        global globalvv

        for x in range (int(self.noofloop)):

            """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
            then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
            will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
            begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
            begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
            are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
            optionally export all the details into a CSV file or display a graph after the test is completed.

            """
            self.infoList = []
            self.dataList = []
            self.dataList2 = []

            dict = dictGenerator.input(
                savedir=self.savelocation,
                V_Rating=self.Voltage_Rating,
                I_Rating=self.Current_Rating,
                P_Rating=self.Power_Rating,
                power=self.Power,
                estimatetime=self.estimatetime,
                updatedelay=self.updatedelay,
                readbackvoltage=self.readbackvoltage,
                readbackcurrent=self.readbackcurrent,
                noofloop=self.noofloop,
                Instrument=self.DMM_Instrument,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,
                unit=self.unit,
                minCurrent=self.minCurrent,
                maxCurrent=self.maxCurrent,
                current_step_size=self.current_step_size,
                minVoltage=self.minVoltage,
                maxVoltage=self.maxVoltage,
                voltage_step_size=self.voltage_step_size,
                PSU=self.PSU,
                DMM=self.DMM,
                OSC=self.OSC,
                ELoad=self.ELoad,
                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],

                OSC_Channel=self.OSC_Channel,
                Channel_CouplingMode=self.Channel_CouplingMode,
                Trigger_Mode=self.Trigger_Mode,
                Trigger_CouplingMode=self.Trigger_CouplingMode,
                Trigger_SweepMode=self.Trigger_SweepMode,
                Trigger_SlopeMode=self.Trigger_SlopeMode,
                Probe_Setting=self.Probe_Setting,
                Acq_Type=self.Acq_Type,
                TimeScale=self.TimeScale,
                VerticalScale=self.VerticalScale,
                
                V_Settling_Band=self.V_Settling_Band,
                T_Settling_Band=self.T_Settling_Band,
            )

            #Function: Check if any of the parameters are empty
            if not self.check_missing_params(dict):
                return

            if self.checkbox_test_TransientRecovery != 2 and self.checkbox_test_VoltageAccuracy != 2 and self.checkbox_test_VoltageLoadRegulation != 2:
                QMessageBox.warning(
                    self,
                    "Test is not selected !!!",
                    "Please select the required tests before proceeding."
                )
                break

            if x == 0:
                QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
                                    )
                """ globalvv = dict["estimatetime"]
                loading_thread = threading.Thread(target=self.tkinter_loading_screen, args=(globalvv,))
                loading_thread.start()
                """
                
                                      
            for i in [dict]:
                if i == "":
                    QMessageBox.warning(
                        self, "Error", "One of the parameters are not filled in"
                    )
                    break

            else:
                A = VisaResourceManager()
                flag, args = A.openRM(self.ELoad, self.PSU, self.DMM)

                if flag == 0:
                    string = ""
                    for item in args:
                        string = string + item

                    QMessageBox.warning(self, "VISA IO ERROR", string)
                    exit()

                if self.checkbox_test_VoltageAccuracy == 2:
                    if self.DMM_Instrument == "Keysight":
                        try:(
                            infoList,
                            dataList,
                            dataList2
                            ) = NewVoltageMeasurement.Execute_Voltage_Accuracy(self, dict)
                        
                        
                        except Exception as e:
                            QMessageBox.warning(self, "Error", str(e))
                            exit()


                    if x == (int(self.noofloop) - 1):   
                        self.OutputBox.append(my_result.getvalue())
                        self.OutputBox.append("Measurement is complete !")


                    if self.checkbox_data_Report == 2:
                        instrumentData(self.PSU, self.DMM, self.ELoad)
                        datatoCSV_Accuracy(infoList, dataList, dataList2)
                        datatoGraph(infoList, dataList,dataList2)
                        datatoGraph.scatterCompareVoltage(self, float(self.Programming_Error_Gain), float(self.Programming_Error_Offset), float(self.Readback_Error_Gain), float(self.Readback_Error_Offset), str(self.unit), float(self.Voltage_Rating))

                        A = xlreport(save_directory=self.savelocation, file_name=str(self.unit))
                        A.run()
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        df.to_csv("C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/config.csv")

                    if x == (int(self.noofloop) - 1):  
                        if self.checkbox_data_Image == 2:
                            self.image_dialog = image_Window()
                            self.image_dialog.setModal(True)
                            self.image_dialog.show()

                if self.checkbox_test_VoltageLoadRegulation == 2:
                    if self.DMM_Instrument == "Keysight":
                        try:
                            LoadRegulation.executeCV_LoadRegulationA(self, dict)

                        except Exception as e:
                            QMessageBox.warning(self, "Error", str(e))
                            exit()

                
                    self.OutputBox.append(my_result.getvalue())
                    self.OutputBox.append("Measurement is complete !")
                
                if self.checkbox_test_TransientRecovery == 2:
                    try:
                        if self.checkbox_SpecialCase == 2:
                            RiseFallTime.executeA(self, dict)
                        
                        if self.checkbox_NormalCase == 2:
                            RiseFallTime.executeB(self, dict)

                    except Exception as e:
                        print_console_safe(e)
                        QMessageBox.warning(self, "Error", str(e))
                        exit()

                    self.OutputBox.append(my_result.getvalue())
                    self.OutputBox.append("Measurement is complete !")                   
   
class BundleMeasurementCurrentandPowerDialog(QDialog):
    """Class for configuring the voltage measurement DUT Tests Dialog.
    A widget is declared for each parameter that can be customized by the user. These widgets can come in
    the form of QLineEdit, or QComboBox where user can select their preferred parameters. When the widgets
    detect changes, a signal will be transmitted to a designated slot which is a method in this class
    (e.g. [paramter_name]_changed). The parameter values will then be updated. At runtime execution of the
    DUT Test, the program will compile all the parameters into a dictionary which will be passed as an argument
    into the test methods and execute the DUT Tests accordingly.

    For more details regarding the arguments, see DUT_Test_Scripts/Dolphin/DUT_Test.py.


    """

    def __init__(self):
        """Method where Widgets, Signals and Slots are defined in the GUI for Voltage Measurement"""
        super().__init__()
        # Initialize default CSV path
        self.DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/data.csv"  # or use a dynamic path if needed
        self.IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/Chart.png"
        self.ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/error.csv"
        self.image_dialog = None
        self.POWER_DATA_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerdata.csv"  # or use a dynamic path if needed
        self.POWER_IMAGE_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/images/powerChart.png"
        self.POWER_ERROR_CSV_PATH = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powererror.csv"

        # Default Values
        #self.PSU_VisaAddress = "USB0::0x2A8D::0xCC04::MY00000037::0::INSTR"
        #self.DMM_VisaAddress = "USB0::0x2A8D::0x1601::MY60094787::0::INSTR"
        #self.ELoad_VisaAddress = "USB0::0x2A8D::0x3902::MY60260005::0::INSTR"
        self.rshunt = "0.05"
        self.Power_Rating = "6000"
        self.Current_Rating = "24"
        self.Voltage_Rating = "800"
        self.Power = "2000"
        self.powerfin = self.Power
        self.powerini = "0"
        self.power_step_size = "60"
        self.currloop = "1"
        self.voltloop = "1"
        self.estimatetime = "0"
        self.readbackvoltage = "0"
        self.readbackcurrent = "0"
        self.savelocation = "C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Test Data/File Export Testing"
        self.noofloop = "1"
        self.updatedelay = "0.0"
        self.unit = "CURRENT"
        self.Programming_Error_Offset = "0.001"
        self.Programming_Error_Gain = "0.001"
        self.Readback_Error_Offset = "0.001"
        self.Readback_Error_Gain = "0.001"

        self.Power_Programming_Error_Offset = "0.005"
        self.Power_Programming_Error_Gain = "0.005"
        self.Power_Readback_Error_Offset = "0.005"
        self.Power_Readback_Error_Gain = "0.005"

        self.minCurrent = "1"
        self.maxCurrent = "1"
        self.current_step_size = "1"
        self.minVoltage = "1"
        self.maxVoltage = "10"
        self.voltage_step_size = "1"

        self.PSU = "USB0::0x2A8D::0xDA04::CN24350083::0::INSTR"
        self.DMM = "USB0::0x2A8D::0x0201::MY57702128::0::INSTR"
        self.DMM2 = "USB0::0x2A8D::0x0201::MY54701197::0::INSTR"
        self.ELoad = "USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"
        self.ELoad_Channel = "1"
        self.PSU_Channel = "1"
        self.DMM_Instrument = "Keysight"

        self.setFunction = "Current" #Set Eload in CC Mode
        self.VoltageRes = "DEF"
        self.VoltageSense = "INT"

        self.checkbox_data_Report = 2
        self.checkbox_data_Image = 2
        self.checkbox_test_CurrentAccuracy = 2
        self.checkbox_test_CurrentLoadRegulation = 2
        self.checkbox_test_PowerAccuracy = 2

        self.Range = "Auto"
        self.Aperture = "0.2"
        self.AutoZero = "ON"
        self.inputZ = "ON"
        self.UpTime = "50"
        self.DownTime = "50"
        
        # Ensure DATA_CSV_PATH is defined before use
        if not hasattr(self, 'DATA_CSV_PATH') or not self.DATA_CSV_PATH:
            QMessageBox.warning(self, "Error", "CSV path is not set.")

        if not hasattr(self, 'POWER_DATA_CSV_PATH') or not self.POWER_DATA_CSV_PATH:
            QMessageBox.warning(self, "Error", "POWER CSV path is not set.")

        self.setWindowTitle("Bundle Measurement Current, Load Regulation and Power")

        self.image_window = None
        self.setWindowFlags(Qt.Window)

        # Allow resizing
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setMinimumSize(800, 600)
        self.image_window = None
        self.setStyleSheet("font-size: 18px;")

        # create find button 
        QPushButton_Widget0 = QPushButton()
        QPushButton_Widget0.setText("Save Path")
        QPushButton_Widget1 = QPushButton()
        QPushButton_Widget1.setText("Execute Test")
        QPushButton_Widget2 = QPushButton()
        QPushButton_Widget2.setText("Advanced Settings")
        QPushButton_Widget3 = QPushButton()
        QPushButton_Widget3.setText("Estimate Data Collection Time")
        QPushButton_Widget4 = QPushButton()
        QPushButton_Widget4.setText("Find Instruments")
        QCheckBox_Report_Widget = QCheckBox()
        QCheckBox_Report_Widget.setText("Generate Excel Report")
        QCheckBox_Report_Widget.setCheckState(Qt.Checked)
        QCheckBox_Image_Widget = QCheckBox()
        QCheckBox_Image_Widget.setText("Show Graph")
        QCheckBox_Image_Widget.setCheckState(Qt.Checked)

          #Test checkbox
        QCheckBox_CurrentAccuracy_Widget = QCheckBox()
        QCheckBox_CurrentAccuracy_Widget.setText("Current Accuracy")
        QCheckBox_CurrentAccuracy_Widget.setCheckState(Qt.Checked)
        QCheckBox_CurrentLoadRegulation_Widget = QCheckBox()
        QCheckBox_CurrentLoadRegulation_Widget.setText("Current Load Regulation")
        QCheckBox_CurrentLoadRegulation_Widget.setCheckState(Qt.Checked)
        QCheckBox_PowerAccuracy_Widget = QCheckBox()
        QCheckBox_PowerAccuracy_Widget.setText("Power Accuracy")
        QCheckBox_PowerAccuracy_Widget.setCheckState(Qt.Checked)

        layout1 = QFormLayout()
        self.OutputBox = QTextBrowser()

        self.OutputBox.append(my_result.getvalue())

        
        Desp0 = QLabel()
        Desp1 = QLabel()
        Desp2 = QLabel()
        Desp3 = QLabel()
        Desp4 = QLabel()
        Desp5 = QLabel()
        Desp6 = QLabel()
        Desp7 = QLabel()
        Desp8 = QLabel()



        #Desp0.setFont (desp_font)
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        Desp3.setFont(desp_font)
        Desp4.setFont(desp_font)
        Desp5.setFont(desp_font)
        Desp6.setFont(desp_font)
        Desp7.setFont(desp_font)
        #Desp8.setFont(desp_font)


        #Desp0.setText("Save Path:")
        Desp1.setText("Connections:")
        Desp2.setText("General Settings:")
        Desp3.setText("Voltage Sweep:")
        Desp4.setText("Current Sweep:")
        Desp5.setText("No. of Collection:")
        Desp6.setText("Rated Power of DUT Setting [W]:")
        Desp7.setText("Maximum Testing Current")
        #Desp8.setText("Testing Selection:")



        #Save Path
        QLabel_Save_Path = QLabel()
        QLabel_Save_Path.setFont(desp_font)
        QLabel_Save_Path.setText("Drive Location/Output Wndow:")

        #Testing Selection
        QLabel_Testing_Selection = QLabel()
        QLabel_Testing_Selection.setFont(desp_font)
        QLabel_Testing_Selection.setText("Test:")


        #Find Instruments
        # Connections
        self.QLabel_PSU_VisaAddress = QLabel()
        self.QLabel_DMM_VisaAddressforVoltage = QLabel()
        self.QLabel_DMM_VisaAddressforCurrent = QLabel()
        self.QLabel_ELoad_VisaAddress = QLabel()
        QLabel_DMM_Instrument = QLabel()
        self.QLabel_PSU_VisaAddress.setText("Visa Address (PSU):")
        self.QLabel_DMM_VisaAddressforVoltage.setText("Visa Address (DMM for Measure Voltage):")
        self.QLabel_DMM_VisaAddressforCurrent.setText("Visa Address (DMM for Measure Current Shunt):")
        self.QLabel_ELoad_VisaAddress.setText("Visa Address (ELoad):")
        QLabel_DMM_Instrument.setText("Instrument Type (DMM):")
        #QLineEdit_PSU_VisaAddress = QLineEdit()
        #QLineEdit_DMM_VisaAddress = QLineEdit()
        #QLineEdit_ELoad_VisaAddress = QLineEdit()

        self.QLineEdit_PSU_VisaAddress = QComboBox()
        self.QLineEdit_DMM_VisaAddressforVoltage = QComboBox()
        self.QLineEdit_DMM_VisaAddressforCurrent = QComboBox()
        self.QLineEdit_ELoad_VisaAddress = QComboBox()
        QComboBox_DMM_Instrument = QComboBox()

        # General Settings
        QLabel_Voltage_Res = QLabel()
        #QLabel_ELoad_Display_Channel = QLabel()
        #QLabel_PSU_Display_Channel = QLabel()
        QLabel_set_PSU_Channel = QLabel()
        QLabel_set_ELoad_Channel = QLabel()
        QLabel_set_Function = QLabel()
        QLabel_Voltage_Sense = QLabel()
    
        QLabel_Programming_Error_Gain = QLabel()
        QLabel_Programming_Error_Offset = QLabel()
        QLabel_Readback_Error_Gain = QLabel()
        QLabel_Readback_Error_Offset = QLabel()

        QLabel_Power_Programming_Error_Gain = QLabel()
        QLabel_Power_Programming_Error_Offset = QLabel()
        QLabel_Power_Readback_Error_Gain = QLabel()
        QLabel_Power_Readback_Error_Offset = QLabel()

        QLabel_Voltage_Res.setText("Voltage Resolution (DMM):")
        #QLabel_ELoad_Display_Channel.setText("Display Channel (Eload):")
        #QLabel_PSU_Display_Channel.setText("Display Channel (PSU):")

        QLabel_set_PSU_Channel.setText("PSU Channel:")
        QLabel_set_ELoad_Channel.setText("Eload Channel:")
        QLabel_set_Function.setText("Mode(Eload):")
        QLabel_Voltage_Sense.setText("Voltage Sense:")

        QLabel_Programming_Error_Gain.setText("Current_Programming Desired Specification (Gain):")
        QLabel_Programming_Error_Offset.setText("Current_Programming Desired Specification (Offset):")
        QLabel_Readback_Error_Gain.setText("Current_Readback Desired Specification (Gain):")
        QLabel_Readback_Error_Offset.setText("Current_Readback Desired Specification (Offset):")

        QLabel_Power_Programming_Error_Gain.setText("Power_Programming Desired Specification (Gain):")
        QLabel_Power_Programming_Error_Offset.setText("Power_Programming Desired Specification (Offset):")
        QLabel_Power_Readback_Error_Gain.setText("Power_Readback Desired Specification (Gain):")
        QLabel_Power_Readback_Error_Offset.setText("Power_Readback Desired Specification (Offset):")

        QComboBox_Voltage_Res = QComboBox()
        QComboBox_set_PSU_Channel = QComboBox()
        QComboBox_set_ELoad_Channel = QComboBox()
        QComboBox_set_Function = QComboBox()
        QComboBox_Voltage_Sense = QComboBox()

        QLineEdit_Programming_Error_Gain = QLineEdit()
        QLineEdit_Programming_Error_Offset = QLineEdit()
        QLineEdit_Readback_Error_Gain = QLineEdit()
        QLineEdit_Readback_Error_Offset = QLineEdit()

        QLineEdit_Power_Programming_Error_Gain = QLineEdit()
        QLineEdit_Power_Programming_Error_Offset = QLineEdit()
        QLineEdit_Power_Readback_Error_Gain = QLineEdit()
        QLineEdit_Power_Readback_Error_Offset = QLineEdit()

        self.QLineEdit_PSU_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350097::0::INSTR"])
        self.QLineEdit_DMM_VisaAddressforVoltage.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_DMM_VisaAddressforCurrent.addItems(["USB0::0x2A8D::0x1601::MY60094787::0::INSTR"])
        self.QLineEdit_ELoad_VisaAddress.addItems(["USB0::0x2A8D::0xDA04::CN24350105::0::INSTR"])

        QComboBox_DMM_Instrument.addItems(["Keysight", "Keithley"])
        QComboBox_DMM_Instrument.setEnabled(False)
        QComboBox_Voltage_Res.addItems(["SLOW", "MEDIUM", "FAST"])
        QComboBox_set_Function.addItems(
            [
                "Current Priority",
                "Voltage Priority",
                "Resistance Priority",
            ]
        )
        QComboBox_set_PSU_Channel.addItems(["1", "2", "3", "4"])
        QComboBox_set_PSU_Channel.setEnabled(True)

        QComboBox_set_ELoad_Channel.addItems(["1", "2"])
        QComboBox_set_ELoad_Channel.setEnabled(True)

        QComboBox_set_Function.setEnabled(False)
        QComboBox_Voltage_Sense.addItems(["2 Wire", "4 Wire"])

        #Power for test
        QLabel_Power = QLabel()
        QLabel_Power.setText("Power Set for Test(W):")
        QLineEdit_Power = QLineEdit()

        #Rated Power
        QLabel_power_rated = QLabel()
        QLabel_power_rated.setText("Rated Power of DUT (W):")
        QLineEdit_power_rated = QLineEdit()

        QLabel_PowerINI = QLabel()
        QLabel_PowerINI.setText("Initial Power (W):")
        QLineEdit_PowerINI = QLineEdit()

        QLabel_PowerFIN = QLabel()
        QLabel_PowerFIN.setText("Final Power (W):")
        QLineEdit_PowerFIN = QLineEdit()

        QLabel_power_step_size = QLabel()
        QLabel_power_step_size.setText("Step Size:")
        QLineEdit_power_step_size = QLineEdit()

        # Current Sweep
        QLabel_minCurrent = QLabel()
        QLabel_maxCurrent = QLabel()
        QLabel_current_step_size = QLabel()
        QLabel_current_rated = QLabel()
        QLabel_minCurrent.setText("Initial Current (A):")
        QLabel_maxCurrent.setText("Final Current (A):")
        QLabel_current_step_size.setText("Step Size:")
        QLabel_current_rated.setText("DUT Rated Current:")

        QLineEdit_minCurrent = QLineEdit()
        QLineEdit_maxCurrent = QLineEdit()
        QLineEdit_current_stepsize = QLineEdit()
        QLineEdit_current_rated = QLineEdit()

        # Voltage Sweep
        QLabel_minVoltage = QLabel()
        QLabel_maxVoltage = QLabel()
        QLabel_voltage_step_size = QLabel()
        QLabel_voltage_rated = QLabel()
        QLabel_minVoltage.setText("Initial Voltage (V):")
        QLabel_maxVoltage.setText("Final Voltage (V):")
        QLabel_voltage_step_size.setText("Step Size:")
        QLabel_voltage_rated.setText("DUT Rated Voltage:")

        QLineEdit_minVoltage = QLineEdit()
        QLineEdit_maxVoltage = QLineEdit()
        QLineEdit_voltage_stepsize = QLineEdit()
        QLineEdit_voltag_rated = QLineEdit()
        QLineEdit_voltag_rated.setFixedSize(100, 40)  # Set fixed size

        QLabel_rshunt = QLabel()
        QLabel_rshunt.setText("Shunt Resistance Value (ohm):")
        QLineEdit_rshunt = QLineEdit()

        #Loop
        QLabel_noofloop = QLabel()
        QLabel_noofloop.setText("No. of Data Collection:")
        QComboBox_noofloop = QComboBox()
        QComboBox_noofloop.addItems(["1","2","3","4","5","6","7","8","9","10"])


        QLabel_updatedelay = QLabel()
        QLabel_updatedelay.setText("Delay Time (second) :(Default=50ms)")
        QComboBox_updatedelay = QComboBox()
        QComboBox_updatedelay.addItems(["0.0","0.8","1.0","2.0","3.0", "4.0"])
        
        # Create a horizontal layout for the "Save Path" and checkboxes
        save_path_layout = QHBoxLayout()
        save_path_layout.addWidget(QLabel_Save_Path)  # QLabel for "Save Path"
        #save_path_layout.addWidget(QLineEdit_Save_Path)  # QLineEdit for the path
        save_path_layout.addWidget(QCheckBox_Report_Widget)  # Checkbox for "Generate Excel Report"
        save_path_layout.addWidget(QCheckBox_Image_Widget)  # Checkbox for "Show Graph"

        # Create a horizontal layout for the "Save Path" and checkboxes
        testing_selection_layout = QHBoxLayout()
        testing_selection_layout.addWidget(QLabel_Testing_Selection)  # QLabel for "Save Path"
        testing_selection_layout.addWidget(QCheckBox_CurrentAccuracy_Widget)  # Checkbox for "Generate Excel Report"
        testing_selection_layout.addWidget(QCheckBox_CurrentLoadRegulation_Widget)  # Checkbox for "Show Graph"
        testing_selection_layout.addWidget(QCheckBox_PowerAccuracy_Widget) 

        voltage_sweep_layout = QHBoxLayout()
        voltage_sweep_layout.addWidget(Desp3)  # QLabel for "Save Path"
        voltage_sweep_layout.addWidget(QLabel_voltage_rated)  # Checkbox for "Generate Excel Report"
        voltage_sweep_layout.addWidget( QLineEdit_voltag_rated )  # Checkbox for "Show Graph"

        current_sweep_layout = QHBoxLayout()
        current_sweep_layout.addWidget(Desp4)  # QLabel for "Save Path"
        current_sweep_layout.addWidget(QLabel_current_rated)  # Checkbox for "Generate Excel Report"
        current_sweep_layout.addWidget( QLineEdit_current_rated )  # Checkbox for "Show Graph"

        power_sweep_layout = QHBoxLayout()
        power_sweep_layout.addWidget(Desp6)  # QLabel for "Save Path"
        power_sweep_layout.addWidget(QLabel_power_rated)  # Checkbox for "Generate Excel Report"
        power_sweep_layout.addWidget( QLineEdit_power_rated)  # Checkbox for "Show Graph"

        voltage_inifin_layout = QHBoxLayout()
        voltage_inifin_layout.addWidget(QLabel_minVoltage)  # QLabel for "Save Path"
        voltage_inifin_layout.addWidget(QLineEdit_minVoltage)  # Checkbox for "Generate Excel Report"
        voltage_inifin_layout.addWidget(QLabel_maxVoltage)  # Checkbox for "Show Graph"
        voltage_inifin_layout.addWidget(QLineEdit_maxVoltage)  # Checkbox for "Show Graph"

        current_inifin_layout = QHBoxLayout()
        current_inifin_layout.addWidget(QLabel_minCurrent)  # QLabel for "Save Path"
        current_inifin_layout.addWidget(QLineEdit_minCurrent)  # Checkbox for "Generate Excel Report"
        current_inifin_layout.addWidget(QLabel_maxCurrent)  # Checkbox for "Show Graph"
        current_inifin_layout.addWidget(QLineEdit_maxCurrent)  # Checkbox for "Show Graph"

        programming_error_layout = QHBoxLayout()
        programming_error_layout.addWidget(QLabel_Programming_Error_Gain)  # QLabel for "Save Path"
        programming_error_layout.addWidget(QLineEdit_Programming_Error_Gain)  # Checkbox for "Generate Excel Report"
        programming_error_layout.addWidget(QLabel_Programming_Error_Offset)  # Checkbox for "Show Graph"
        programming_error_layout.addWidget(QLineEdit_Programming_Error_Offset)  # Checkbox for "Show Graph"

        readback_error_layout = QHBoxLayout()
        readback_error_layout.addWidget(QLabel_Readback_Error_Gain)  # QLabel for "Save Path"
        readback_error_layout.addWidget(QLineEdit_Readback_Error_Gain)  # Checkbox for "Generate Excel Report"
        readback_error_layout.addWidget(QLabel_Readback_Error_Offset)  # Checkbox for "Show Graph"
        readback_error_layout.addWidget(QLineEdit_Readback_Error_Offset)  # Checkbox for "Show Graph"

        power_programming_error_layout = QHBoxLayout()
        power_programming_error_layout.addWidget(QLabel_Power_Programming_Error_Gain)  # QLabel for "Save Path"
        power_programming_error_layout.addWidget(QLineEdit_Power_Programming_Error_Gain)  # Checkbox for "Generate Excel Report"
        power_programming_error_layout.addWidget(QLabel_Power_Programming_Error_Offset)  # Checkbox for "Show Graph"
        power_programming_error_layout.addWidget(QLineEdit_Power_Programming_Error_Offset)  # Checkbox for "Show Graph"

        power_readback_error_layout = QHBoxLayout()
        power_readback_error_layout.addWidget(QLabel_Power_Readback_Error_Gain)  # QLabel for "Save Path"
        power_readback_error_layout.addWidget(QLineEdit_Power_Readback_Error_Gain)  # Checkbox for "Generate Excel Report"
        power_readback_error_layout.addWidget(QLabel_Power_Readback_Error_Offset)  # Checkbox for "Show Graph"
        power_readback_error_layout.addWidget(QLineEdit_Power_Readback_Error_Offset)  # Checkbox for "Show Graph"

        # Add the combined layout to the main layout
        layout1.addRow(save_path_layout)

        # Rest of your layout remains unchanged
        layout1.addRow(self.OutputBox)
        #layout1.addRow(Desp0)
        layout1.addRow(QPushButton_Widget0)
        layout1.addRow(Desp1)
        layout1.addRow(QPushButton_Widget4)
        layout1.addRow(self.QLabel_PSU_VisaAddress, self.QLineEdit_PSU_VisaAddress)
        layout1.addRow(self.QLabel_DMM_VisaAddressforVoltage, self.QLineEdit_DMM_VisaAddressforVoltage)
        layout1.addRow(self.QLabel_DMM_VisaAddressforCurrent, self.QLineEdit_DMM_VisaAddressforCurrent)
        layout1.addRow(self.QLabel_ELoad_VisaAddress, self.QLineEdit_ELoad_VisaAddress)
        layout1.addRow(QLabel_DMM_Instrument, QComboBox_DMM_Instrument)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_set_PSU_Channel, QComboBox_set_PSU_Channel)
        layout1.addRow(QLabel_set_ELoad_Channel, QComboBox_set_ELoad_Channel)
        layout1.addRow(QLabel_set_Function, QComboBox_set_Function)
        layout1.addRow(QLabel_Voltage_Sense, QComboBox_Voltage_Sense)
        layout1.addRow(programming_error_layout)
        layout1.addRow(readback_error_layout)
        layout1.addRow(power_programming_error_layout)
        layout1.addRow(power_readback_error_layout)
        layout1.addRow(Desp6)
        layout1.addRow(power_sweep_layout)
        layout1.addRow(QLabel_Power, QLineEdit_Power)
        layout1.addRow(QLabel_PowerINI, QLineEdit_PowerINI)
        layout1.addRow(QLabel_PowerFIN, QLineEdit_PowerFIN)
        layout1.addRow(QLabel_power_step_size, QLineEdit_power_step_size)
        layout1.addRow(voltage_sweep_layout)
        layout1.addRow(voltage_inifin_layout)
        #layout1.addRow(QLabel_minVoltage, QLineEdit_minVoltage)
        #layout1.addRow(QLabel_maxVoltage, QLineEdit_maxVoltage)
        layout1.addRow(QLabel_voltage_step_size, QLineEdit_voltage_stepsize)
        layout1.addRow(Desp4)
        layout1.addRow(current_sweep_layout)
        layout1.addRow(current_inifin_layout)
        #layout1.addRow(QLabel_minCurrent, QLineEdit_minCurrent)
        #layout1.addRow(QLabel_maxCurrent, QLineEdit_maxCurrent)
        layout1.addRow(QLabel_current_step_size, QLineEdit_current_stepsize)
        layout1.addRow(Desp5)
        layout1.addRow(QLabel_rshunt, QLineEdit_rshunt)
        layout1.addRow(QLabel_noofloop, QComboBox_noofloop)
        layout1.addRow(QLabel_updatedelay, QComboBox_updatedelay)
        layout1.addRow(testing_selection_layout)
        layout1.addRow(QPushButton_Widget3)
        layout1.addRow(QPushButton_Widget2)
        layout1.addRow(QPushButton_Widget1)        

        AdvancedSettingsList.append(self.Range)
        AdvancedSettingsList.append(self.Aperture)
        AdvancedSettingsList.append(self.AutoZero)
        AdvancedSettingsList.append(self.inputZ)
        AdvancedSettingsList.append(self.UpTime)
        AdvancedSettingsList.append(self.DownTime)

        #Set window and scroll area
        self.setLayout(layout1)
        scroll_area(self,layout1)

        self.QLineEdit_PSU_VisaAddress.currentTextChanged.connect(self.PSU_VisaAddress_changed)
        self.QLineEdit_DMM_VisaAddressforVoltage.currentTextChanged.connect(self.DMM_VisaAddressforVoltage_changed)
        self.QLineEdit_DMM_VisaAddressforCurrent.currentTextChanged.connect(self.DMM_VisaAddressforCurrent_changed)
        self.QLineEdit_ELoad_VisaAddress.currentTextChanged.connect(self.ELoad_VisaAddress_changed)
        #QLineEdit_ELoad_Display_Channel.textEdited.connect(self.ELoad_Channel_changed)
        #QLineEdit_PSU_Display_Channel.textEdited.connect(self.PSU_Channel_changed)
        QLineEdit_Programming_Error_Gain.textEdited.connect(self.Programming_Error_Gain_changed)
        QLineEdit_Programming_Error_Offset.textEdited.connect(self.Programming_Error_Offset_changed)
        QLineEdit_Readback_Error_Gain.textEdited.connect(self.Readback_Error_Gain_changed)
        QLineEdit_Readback_Error_Offset.textEdited.connect(self.Readback_Error_Offset_changed)

        QLineEdit_Power_Programming_Error_Gain.textEdited.connect(self.Power_Programming_Error_Gain_changed)
        QLineEdit_Power_Programming_Error_Offset.textEdited.connect(self.Power_Programming_Error_Offset_changed)
        QLineEdit_Power_Readback_Error_Gain.textEdited.connect(self.Power_Readback_Error_Gain_changed)
        QLineEdit_Power_Readback_Error_Offset.textEdited.connect(self.Power_Readback_Error_Offset_changed)

        QLineEdit_Power.textEdited.connect(self.Power_changed)
        QLineEdit_PowerINI.textEdited.connect(self.PowerINI_changed)
        QLineEdit_PowerFIN.textEdited.connect(self.PowerFIN_changed)
        QLineEdit_power_rated.textEdited.connect(self.Power_Rating_changed)
        QLineEdit_power_step_size.textEdited.connect(self.power_step_size_changed)
        QLineEdit_rshunt.textEdited.connect(self.rshunt_changed)

        QLineEdit_minVoltage.textEdited.connect(self.minVoltage_changed)
        QLineEdit_maxVoltage.textEdited.connect(self.maxVoltage_changed)
        QLineEdit_minCurrent.textEdited.connect(self.minCurrent_changed)
        QLineEdit_maxCurrent.textEdited.connect(self.maxCurrent_changed)
        QLineEdit_voltag_rated.textEdited.connect(self.Volatge_Rating_changed)
        QLineEdit_current_rated.textEdited.connect(self.Current_Rating_changed)
        QLineEdit_voltage_stepsize.textEdited.connect(self.voltage_step_size_changed)
        QLineEdit_current_stepsize.textEdited.connect(self.current_step_size_changed)

        QComboBox_set_PSU_Channel.currentTextChanged.connect(self.set_PSU_Channel_changed)
        QComboBox_set_ELoad_Channel.currentTextChanged.connect(self.set_ELoad_Channel_changed)
        QComboBox_set_Function.currentTextChanged.connect(self.set_Function_changed)
        QComboBox_Voltage_Res.currentTextChanged.connect(self.set_VoltageRes_changed)
        QComboBox_Voltage_Sense.currentTextChanged.connect(self.set_VoltageSense_changed)

        QComboBox_noofloop.currentTextChanged.connect(self.noofloop_changed)
        QComboBox_updatedelay.currentTextChanged.connect(self.updatedelay_changed)

        QComboBox_DMM_Instrument.currentTextChanged.connect(self.DMM_Instrument_changed)

        QCheckBox_Report_Widget.stateChanged.connect(self.checkbox_state_Report)
        QCheckBox_Image_Widget.stateChanged.connect(self.checkbox_state_Image)
        QCheckBox_CurrentAccuracy_Widget.stateChanged.connect(self.checkbox_state_CurrentAccuracy)
        QCheckBox_CurrentLoadRegulation_Widget.stateChanged.connect(self.checkbox_state_CurrentLoadRegulation)
        QCheckBox_PowerAccuracy_Widget.stateChanged.connect(self.checkbox_state_PowerAccuracy)
    

        QPushButton_Widget0.clicked.connect(self.savepath)
        QPushButton_Widget1.clicked.connect(self.executeTest)
        QPushButton_Widget2.clicked.connect(self.openDialog)
        QPushButton_Widget3.clicked.connect(self.estimateTime)
        QPushButton_Widget4.clicked.connect(self.doFind)
    
    def Current_Rating_changed(self, value):
        self.Current_Rating = value
    
    def Volatge_Rating_changed(self, value):
        self.Voltage_Rating = value

    def set_PSU_Channel_changed(self, s):
        self.PSU_Channel = s
    
    def set_ELoad_Channel_changed(self, s):
        self.ELoad_Channel = s

    def rshunt_changed(self, value):
        self.rshunt = value

    def Power_Rating_changed(self, value):
        self.Power_Rating = value
     
    def Power_changed(self, value):
        self.Power = value
    
    def PowerINI_changed(self, value):
        self.powerini= value
    
    def PowerFIN_changed(self, value):
        self.powerfin = value
    
    def power_step_size_changed(self, value):
        self.power_step_size = value

    def doFind(self):
        try:
            self.QLineEdit_PSU_VisaAddress.clear()
            self.QLineEdit_DMM_VisaAddressforVoltage.clear()
            self.QLineEdit_DMM_VisaAddressforCurrent.clear()
            self.QLineEdit_ELoad_VisaAddress.clear()
            
            discovery = GetVisaSCPIResources()
            self.visaIdList = discovery.addresses
            self.nameList = discovery.identities
            
            for i in range(len(self.nameList)):
                self.OutputBox.append(str(self.nameList[i]) + str(self.visaIdList[i]))
                self.QLineEdit_PSU_VisaAddress.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddressforVoltage.addItems([str(self.visaIdList[i])])
                self.QLineEdit_DMM_VisaAddressforCurrent.addItems([str(self.visaIdList[i])])
                self.QLineEdit_ELoad_VisaAddress.addItems([str(self.visaIdList[i])])
                
        except:
            self.OutputBox.append("No Devices Found!!!")
        return   



    def updatedelay_changed(self, value):
        self.updatedelay = value
        self.OutputBox.append(str(self.updatedelay))

    def noofloop_changed(self, value):
        self.noofloop = value
        self.OutputBox.append(str(self.noofloop))

    def DMM_Instrument_changed(self, s):
        self.DMM_Instrument = s

    def PSU_VisaAddress_changed(self, s):
        self.PSU = s    

    def DMM_VisaAddressforVoltage_changed(self, s):
        self.DMM = s
    
    def DMM_VisaAddressforCurrent_changed(self, s):
        self.DMM2 = s

    def ELoad_VisaAddress_changed(self, s):
        self.ELoad = s

    def Programming_Error_Gain_changed(self, s):
        self.Programming_Error_Gain = s

    def Programming_Error_Offset_changed(self, s):
        self.Programming_Error_Offset = s

    def Readback_Error_Gain_changed(self, s):
        self.Readback_Error_Gain = s

    def Readback_Error_Offset_changed(self, s):
        self.Readback_Error_Offset = s

    def Power_Programming_Error_Gain_changed(self, s):
        self.Power_Programming_Error_Gain = s

    def Power_Programming_Error_Offset_changed(self, s):
        self.Power_Programming_Error_Offset = s

    def Power_Readback_Error_Gain_changed(self, s):
        self.Power_Readback_Error_Gain = s

    def Power_Readback_Error_Offset_changed(self, s):
        self.Power_Readback_Error_Offset = s

    def minVoltage_changed(self, s):
        self.minVoltage = s

    def maxVoltage_changed(self, s):
        self.maxVoltage = s

    def minCurrent_changed(self, s):
        self.minCurrent = s

    def maxCurrent_changed(self, s):
        self.maxCurrent = s

    def voltage_step_size_changed(self, s):
        self.voltage_step_size = s

    def current_step_size_changed(self, s):
        self.current_step_size = s

    def set_Function_changed(self, s):
        if s == "Voltage Priority":
            self.setFunction = "Voltage"

        elif s == "Current Priority":
            self.setFunction = "Current"

        elif s == "Resistance Priority":
            self.setFunction = "Resistance"

    def set_VoltageRes_changed(self, s):
        self.VoltageRes = s

    def set_VoltageSense_changed(self, s):
        if s == "2 Wire":
            self.VoltageSense = "INT"
        elif s == "4 Wire":
            self.VoltageSense = "EXT"

    def setRange(self, value):
        AdvancedSettingsList[0] = value

    def setAperture(self, value):
        AdvancedSettingsList[1] = value

    def setAutoZero(self, value):
        AdvancedSettingsList[2] = value

    def setInputZ(self, value):
        AdvancedSettingsList[3] = value

    def checkbox_state_Report(self, s):
        self.checkbox_data_Report = s

    def checkbox_state_Image(self, s):
        self.checkbox_data_Image = s
    
    def checkbox_state_CurrentAccuracy(self, s):
        self.checkbox_test_CurrentAccuracy = s

    def checkbox_state_CurrentLoadRegulation(self, s):
        self.checkbox_test_CurrentLoadRegulation = s
    
    def checkbox_state_PowerAccuracy(self, s):
        self.checkbox_test_PowerAccuracy = s

    def setUpTime(self, value):
        AdvancedSettingsList[4] = value

    def setDownTime(self, value):
        AdvancedSettingsList[5] = value

    def openDialog(self):
        dlg = AdvancedSetting_Voltage()
        dlg.exec()

    def estimateTime(self):
        
        self.OutputBox.append(str(self.updatedelay))
           
        self.currloop = ((float(self.maxCurrent) - float(self.minCurrent))/ float(self.current_step_size)) + 1
        self.voltloop = ((float(self.maxVoltage) - float(self.minVoltage))/ float(self.voltage_step_size)) + 1

        if self.updatedelay == 0.0:
            constant = 0
                
        else:
            constant = 1

        self.estimatetime = (self.currloop * self.voltloop *(0.2 + 0.8 + (float(self.updatedelay) * constant) )) * float(self.noofloop)
        self.OutputBox.append(str(self.estimatetime) + "seconds")
    



    def savepath(self):  
        # Create a Tkinter root window
        root = Tk()
        root.withdraw()  # Hide the root window

        # Open a directory dialog and return the selected directory path
        directory = filedialog.askdirectory()
        self.savelocation = directory
        self.OutputBox.append(str(self.savelocation))

         
    

    def executeTest(self):
        global globalvv

        for x in range (int(self.noofloop)):

            """The method begins by compiling all the parameters in a dictionary for ease of storage and calling,
            then the parameters are looped through to check if any of them are empty or return NULL, a warning dialogue
            will appear if the statement is true, and the users have to troubleshoot the issue. After so, the tests will
            begin right after another warning dialogue prompting the user that the tests will begin very soon. When test
            begins, the VISA_Addresses of the Instruments are passed through the VISA Resource Manager to make sure there
            are connected. Then the actual DUT Tests will commence. Depending on the users selection, the method can
            optionally export all the details into a CSV file or display a graph after the test is completed.

            """
            self.infoList = []
            self.dataList = []
            self.dataList2 = []

            dict = dictGenerator.input(
                rshunt = self.rshunt,
                savedir=self.savelocation,
                V_Rating=self.Voltage_Rating,
                I_Rating=self.Current_Rating,
                P_Rating=self.Power_Rating,

                power=self.Power,
                powerini=self.powerini,
                powerfin=self.powerfin,
                power_step_size = self.power_step_size,

                estimatetime=self.estimatetime,
                updatedelay=self.updatedelay,
                readbackvoltage=self.readbackvoltage,
                readbackcurrent=self.readbackcurrent,
                noofloop=self.noofloop,
                Instrument=self.DMM_Instrument,
                Programming_Error_Gain=self.Programming_Error_Gain,
                Programming_Error_Offset=self.Programming_Error_Offset,
                Readback_Error_Gain=self.Readback_Error_Gain,
                Readback_Error_Offset=self.Readback_Error_Offset,

                PowerProgramming_Error_Gain=self.Power_Programming_Error_Gain,
                PowerProgramming_Error_Offset=self.Power_Programming_Error_Offset,
                PowerReadback_Error_Gain=self.Power_Readback_Error_Gain,
                PowerReadback_Error_Offset=self.Power_Readback_Error_Offset,

                unit=self.unit,
                minCurrent=self.minCurrent,
                maxCurrent=self.maxCurrent,
                current_step_size=self.current_step_size,
                minVoltage=self.minVoltage,
                maxVoltage=self.maxVoltage,
                voltage_step_size=self.voltage_step_size,
                

                PSU=self.PSU,
                DMM=self.DMM,
                DMM2=self.DMM2,
                ELoad=self.ELoad,
                ELoad_Channel=self.ELoad_Channel,
                PSU_Channel=self.PSU_Channel,
                VoltageSense=self.VoltageSense,
                VoltageRes=self.VoltageRes,
                setFunction=self.setFunction,
                Range=AdvancedSettingsList[0],
                Aperture=AdvancedSettingsList[1],
                AutoZero=AdvancedSettingsList[2],
                InputZ=AdvancedSettingsList[3],
                UpTime=AdvancedSettingsList[4],
                DownTime=AdvancedSettingsList[5],
            )

            #Function: Check if any of the parameters are empty
            if not self.check_missing_params(dict):
                return

            #Function: Check if any test selected
            if self.checkbox_test_CurrentAccuracy != 2 and self.checkbox_test_CurrentLoadRegulation != 2 and self.checkbox_test_PowerAccuracy != 2:
                QMessageBox.warning(
                    self,
                    "Test is not selected !!!",
                    "Please select the required tests before proceeding."
                )
                break
            
            if x == 0:
                QMessageBox.warning(
                self,
                "In Progress",
                "Measurement will start now , please do not close the main window until test is completed",
                                    )
                """globalvv = dict["estimatetime"]
                loading_thread = threading.Thread(target=self.tkinter_loading_screen, args=(globalvv,))
                loading_thread.start()"""
                
                
                                      
            for i in [dict]:
                if i == "":
                    QMessageBox.warning(
                        self, "Error", "One of the parameters are not filled in"
                    )
                    break

            else:
                A = VisaResourceManager()
                flag, args = A.openRM(self.ELoad, self.PSU, self.DMM, self.DMM2)

                if flag == 0:
                    string = ""
                    for item in args:
                        string = string + item

                    QMessageBox.warning(self, "VISA IO ERROR", string)
                    break

                if self.checkbox_test_CurrentAccuracy == 2:
                    if self.DMM_Instrument == "Keysight":
                        try:
                        # Clear existing values before assigning new ones

                        # Execute measurement and assign new values
                            infoList, dataList, dataList2 = NewCurrentMeasurement.executeCurrentMeasurementA(self, dict)
                        except Exception as e:
                            QMessageBox.warning(self, "Error", str(e))
                            exit()


                    if self.checkbox_data_Report == 2:
                        instrumentData(self.PSU, self.DMM, self.ELoad)
                        datatoCSV_Accuracy2(infoList, dataList, dataList2)
                        datatoGraph2(infoList, dataList,dataList2)
                        datatoGraph2.scatterCompareCurrent2(self, float(self.Programming_Error_Gain), float(self.Programming_Error_Offset), float(self.Readback_Error_Gain), float(self.Readback_Error_Offset), str(self.unit), float(self.Current_Rating))
                        

                        # Create DataFrame from dictionary
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        
                        # Rename the index to a custom name, e.g., 'Index'
                        df = df.rename_axis('Test Condition')

                        # Save to CSV with the renamed index
                        df.to_csv("C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/config.csv")

                        # Create report
                        A = xlreport(save_directory=self.savelocation, file_name=str(self.unit))
                        A.run()
                        

                    if x == (int(self.noofloop) - 1):  
                        if self.checkbox_data_Image == 2:
                            self.image_dialog = image_Window()
                            self.image_dialog.setModal(True)
                            self.image_dialog.show()

                if self.checkbox_test_CurrentLoadRegulation == 2:
                    if self.DMM_Instrument == "Keysight":
                        try:
                            LoadRegulation.executeCC_LoadRegulationA(self, dict)

                            """ A = xlreport(save_directory=self.savelocation, file_name=str(self.unit))
                            A.run()
                            df = pd.DataFrame.from_dict(dict, orient="index")
                            df.to_csv("csv/config.csv")"""

                        except Exception as e:
                            QMessageBox.warning(self, "Error", str(e))
                            exit()

                
                    self.OutputBox.append(my_result.getvalue())
                    self.OutputBox.append("Measurement is complete !")
                
                if self.checkbox_test_PowerAccuracy == 2:
                    try:
                        if self.checkbox_test_CurrentAccuracy == 2:
                            # Clear existing values before assigning new ones
                            infoList.clear()
                            dataList.clear()
                            dataList2.clear()
                            infoList, dataList, dataList2 = PowerMeasurement.executePowerMeasurementA(self, dict)  # Power CC
                        else:
                            # Execute measurement and assign new values
                            infoList, dataList, dataList2 = PowerMeasurement.executePowerMeasurementA(self, dict)  # Power CC

                    except Exception as e:
                        QMessageBox.warning(self, "Error", str(e))
                        exit()

                    if self.checkbox_data_Report == 2:
                        powerinstrumentData(self.PSU, self.DMM, self.DMM2, self.ELoad)
                        datatoCSV_PowerAccuracy(infoList, dataList, dataList2)
                        datatoGraph3(infoList, dataList,dataList2)
                        self.setFunction="Current"
                        datatoGraph3.scatterComparePower(self, float(self.Power_Programming_Error_Gain), float(self.Power_Programming_Error_Offset), float(self.Power_Readback_Error_Gain), float(self.Power_Readback_Error_Offset), str(self.setFunction), float(self.Power_Rating) )
                        A = xlreportpower(save_directory=self.savelocation, file_name="Power_CC")
                        A.run()
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        df.to_csv("C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerconfig.csv")
                    
        
                    try:
                        # Clear existing values before assigning new ones
                        infoList.clear()
                        dataList.clear()
                        dataList2.clear()

                        # Execute measurement and assign new values
                        infoList, dataList, dataList2 = PowerMeasurement.executePowerMeasurementB(self, dict)  # Power CC

                    except Exception as e:
                        QMessageBox.warning(self, "Error", str(e))
                        exit()

                    if self.checkbox_data_Report == 2:   
                        powerinstrumentData(self.PSU, self.DMM, self.DMM2, self.ELoad)
                        datatoCSV_PowerAccuracy(infoList, dataList, dataList2)
                        datatoGraph3(infoList, dataList,dataList2)
                        self.setFunction="Voltage"
                        datatoGraph3.scatterComparePower(self, float(self.Power_Programming_Error_Gain), float(self.Power_Programming_Error_Offset), float(self.Power_Readback_Error_Gain), float(self.Power_Readback_Error_Offset), str(self.setFunction), float(self.Power_Rating) )
                        A = xlreportpower(save_directory=self.savelocation, file_name="Power_CV")
                        A.run()
                        df = pd.DataFrame.from_dict(dict, orient="index")
                        df.to_csv("C:/PyVisa - Copy  - Excavator - Copy/PyVisa/Executable/GUI/powerconfig.csv")
        
                    if x == (int(self.noofloop) - 1):   
                        self.OutputBox.append(my_result.getvalue())
                        self.OutputBox.append("Measurement is complete !")

                    if x == (int(self.noofloop) - 1):  
                        if self.checkbox_data_Image == 2:
                            self.image_dialog = image_Window2()
                            self.image_dialog.setModal(True)
                            self.image_dialog.show()
                        
class VoltageCalibrationDialog(QDialog):
    """Class for configuring the voltage calibration DUT Tests Dialog."""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PSU Voltage Calibration GUI")
        self.resize(800, 520)

        # Form inputs
        form = QFormLayout()
        self.psu_input = QLineEdit("TCPIP0::141.183.191.52::inst0::INSTR")
        form.addRow("PSU Address:", self.psu_input)

        self.dmm_input = QLineEdit("USB0::0x2A8D::0x0201::MY57700532::0::INSTR")
        form.addRow("DMM Address:", self.dmm_input)

        self.pw_input = QLineEdit("PP8000A")
        form.addRow("Calibration Password:", self.pw_input)

        self.channel_input = QSpinBox()
        self.channel_input.setRange(1, 16)
        self.channel_input.setValue(1)
        form.addRow("Channel:", self.channel_input)

        self.points_input = QLineEdit("P1,P2")
        form.addRow("Calibration Points (csv):", self.points_input)

        # Buttons
        self.start_btn = QPushButton("Start Calibration")
        self.start_btn.clicked.connect(self.on_start)
        self.stop_btn = QPushButton("Stop")
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.on_stop)

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.stop_btn)

        # Log
        self.log = QTextEdit()
        self.log.setReadOnly(True)

        # Layout
        layout = QVBoxLayout()
        layout.addLayout(form)
        layout.addLayout(btn_layout)
        layout.addWidget(QLabel("Log:"))
        layout.addWidget(self.log)
        self.setLayout(layout)

        self.worker = None

    def append_log(self, text):
        self.log.append(text)
        self.log.moveCursor(self.log.textCursor().End)

    def on_start(self):
        psu_addr = self.psu_input.text().strip()
        dmm_addr = self.dmm_input.text().strip()
        pw = self.pw_input.text().strip()
        ch = self.channel_input.value()
        points = [p.strip() for p in self.points_input.text().split(',') if p.strip()]

        if not psu_addr or not dmm_addr:
            QMessageBox.warning(self, "Missing", "Please provide both PSU and DMM VISA addresses.")
            return

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.log.clear()
        self.append_log("Starting calibration thread...")

        self.worker = CalWorker(psu_addr, dmm_addr, pw, ch, points)
        self.worker.log.connect(self.append_log)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)

        '''# Create crash-safe CSV
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  #Shamman changes
        self.live_csv_path = os.path.join(
            self.params.savelocation,
            f"live_voltage_data_{timestamp}.csv"
        )

        self.csv_file = open(self.live_csv_path, "w", newline="")
        self.csv_writer = csv.writer(self.csv_file)

        # Write header
        self.csv_writer.writerow([
            "Index",
            "Timestamp",
            "Programming_V",
            "Readback_V"
        ])
        self.csv_file.flush()

        self.worker.start()'''

    def on_stop(self):
        if self.worker:
            self.append_log("Stop requested...")
            self.worker.stop()
            self.stop_btn.setEnabled(False)

    def on_finished(self):
        self.append_log("Calibration finished")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.worker = None

    def on_error(self, msg):
        QMessageBox.critical(self, "Error", f"Calibration error:\n{msg}")

class PeakPowerTestDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PSU Voltage Calibration GUI")
        self.resize(800, 520)

class CalWorker(QThread):
    log = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, psu_addr, dmm_addr, password, channel, cal_points):
        super().__init__()
        self.psu_addr = psu_addr
        self.dmm_addr = dmm_addr
        self.password = password
        self.channel = int(channel)
        self.cal_points = cal_points
        self._stop_requested = False

    def stop(self):
        self._stop_requested = True

    def scpi(self, inst, cmd, delay=0.2):
        inst.write(cmd)
        time.sleep(delay)

    def check_error(self, psu):
        try:
            err = psu.query("SYST:ERR?")
            if not err.startswith("0"):
                self.log.emit(f"⚠ PSU ERROR: {err}")
            return err
        except Exception as e:
            self.log.emit(f"Error querying SYST:ERR?: {e}")
            return None

    """def run(self):
        try:
            self.log.emit("Opening VISA resources...")
            rm = create_resource_manager()
            psu = None
            dmm = None
            try:
                psu = rm.open_resource(self.psu_addr)
                psu.timeout = 60000
                dmm = rm.open_resource(self.dmm_addr)
                dmm.timeout = 10000
            except Exception as e:
                raise RuntimeError(f"Failed to open instruments: {e}")

            # Step 1: CAL:STAT ON
            self.scpi(psu, 'OUTPUT:STATe ON, (@1)')  # Ensure output is ON"')
            time.sleep(2)
            self.log.emit('Step 1: CAL:STAT ON, "PP8000A"')
            self.scpi(psu, 'CAL:STAT ON,"PP8000A"')
            time.sleep(2)
            self.check_error(psu)
            if self._stop_requested:
                raise RuntimeError("Stopped by user")

            # Step 2: CAL:VOLT 60, (@1)
            self.log.emit(f"Step 2: CAL:VOLT 60, (@{self.channel})")
            self.scpi(psu, f"CAL:VOLT 60,(@{self.channel})")
            time.sleep(2)
            self.check_error(psu)
            if self._stop_requested:
                raise RuntimeError("Stopped by user")

            # Configure DMM
            self.log.emit("Configuring DMM...")
            dmm.write(":CONFigure:VOLTage:DC AUTO,MAXimum")
            dmm.write("TRIG:SOUR IMM")
            dmm.write("SAMP:COUN 1")
            time.sleep(2)

            # Steps 3-6: Loop through cal points
            for point in self.cal_points:
                if self._stop_requested:
                    raise RuntimeError("Stopped by user")

                self.log.emit(f"Step 3/5: CAL:LEV {point}")
                self.scpi(psu, f"CAL:LEV {point}")
                time.sleep(2)
                self.check_error(psu)
                time.sleep(30)

                self.log.emit("Measuring DMM...")
                val = dmm.query("MEAS:VOLT:DC?").strip()
                self.log.emit(f"DMM reading = {val}")

                self.log.emit(f"Step 4/6: CAL:DATA {val}")
                self.scpi(psu, f"CAL:DATA {val}")
                time.sleep(2)
                self.check_error(psu)
                time.sleep(2)

            # Step 7: CAL:SAVE
            self.log.emit("Step 7: CAL:SAVE")
            self.scpi(psu, "CAL:SAVE")
            time.sleep(2)
            self.check_error(psu)
            time.sleep(2.0)

            # Step 8: CAL:STAT OFF
            self.log.emit("Step 8: CAL:STAT OFF")
            self.scpi(psu, "CAL:STAT OFF")
            time.sleep(2)
            self.check_error(psu)

            self.log.emit("=== CALIBRATION COMPLETE ✅ ===")

        except Exception as e:
            tb = traceback.format_exc()
            self.log.emit(f"ERROR: {e}\n{tb}")
            self.error.emit(str(e))
        finally:
            try:
                if 'psu' in locals() and psu is not None:
                    psu.close()
                if 'dmm' in locals() and dmm is not None:
                    dmm.close()
            except Exception:
                pass
            self.finished.emit()"""
    def run2(self):
        dict = dictGenerator.input(
            PSU=self.psu_addr,
            DMM=self.dmm_addr,
            password=self.password,
            channel=self.channel,
            cal_points=self.cal_points,
        )
        try:
            HornbillVoltageCalibration.Execute_Voltage_Calibration(self, dict)
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e)) 
        
        self.finished.emit()
########-----------------------MultiTreading Features--------------------------------#####################


########----------------------- Advanced Setting for Standalone Code ----------------#####################
class AdvancedSetting_Voltage(QDialog):
    """This class is to configure the Advanced Settings when conducting voltage measurements,
    It prompts a secondary dialogue for users to customize more advanced parametes such as
    aperture, range, AutoZero, input impedance etc.
    """

    def __init__(self):
        """Method defining the signals, slots and widgets for Advaced Settings of Voltage Measurements"""
        super().__init__()

        self.setWindowTitle("Advanced Window (Voltage)")

        # Create a font with the desired size
        desp_font = QFont("Arial", 20)  # Set font to Arial with size 12
        input_font = QFont("Arial", 15)  # Set font for input fields (smaller size)

        QPushButton_Widget = QPushButton()
        QPushButton_Widget.setText("Confirm")
        layout1 = QFormLayout()

        Desp1 = QLabel("DMM Settings:")
        Desp2 = QLabel("PSU Settings:")

        # Apply font size to the labels
        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)

        QLabel_Range = QLabel("DC Voltage Range")
        QLabel_Aperture = QLabel("NPLC / PLC")
        QLabel_AutoZero = QLabel("Auto Zero Function")
        QLabel_InputZ = QLabel("Input Impedance")
        QLabel_UpTime = QLabel("Programming Settling Time (UP) (in ms)")
        QLabel_DownTime = QLabel("Programming Setting Time (Down) (in ms)")

        # Apply font to the other labels
        QLabel_Range.setFont(desp_font)
        QLabel_Aperture.setFont(desp_font)
        QLabel_AutoZero.setFont(desp_font)
        QLabel_InputZ.setFont(desp_font)
        QLabel_UpTime.setFont(desp_font)
        QLabel_DownTime.setFont(desp_font)

        self.QComboBox_Range = QComboBox()
        self.QComboBox_Aperture = QComboBox()
        self.QComboBox_AutoZero = QComboBox()
        self.QComboBox_InputZ = QComboBox()
        self.QLineEdit_UpTime = QLineEdit()
        self.QLineEdit_DownTime = QLineEdit()

        self.QComboBox_Range.addItems(["Auto", "100mV", "1V", "10V", "100V", "1kV"])
        # Time Value: "0.001", "0.002", "0.006",
        self.QComboBox_Aperture.addItems(["0.02", "0.06", "0.2", "1", "10", "100"])
        self.QComboBox_AutoZero.addItems(["ON", "OFF"])
        self.QComboBox_InputZ.addItems(["10M", "Auto"])

        # Apply font size to combo boxes
        self.QComboBox_Range.setFont(input_font)
        self.QComboBox_Aperture.setFont(input_font)
        self.QComboBox_AutoZero.setFont(input_font)
        self.QComboBox_InputZ.setFont(input_font)

        # Apply font size to line edits
        self.QLineEdit_UpTime.setFont(input_font)
        self.QLineEdit_DownTime.setFont(input_font)

        layout1.addRow(Desp1)
        layout1.addRow(QLabel_Range, self.QComboBox_Range)
        layout1.addRow(QLabel_Aperture, self.QComboBox_Aperture)
        layout1.addRow(QLabel_AutoZero, self.QComboBox_AutoZero)
        layout1.addRow(QLabel_InputZ, self.QComboBox_InputZ)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_UpTime, self.QLineEdit_UpTime)
        layout1.addRow(QLabel_DownTime, self.QLineEdit_DownTime)

        # Adding the button at the bottom
        layout1.addRow(QPushButton_Widget)

        self.setLayout(layout1)

        self.QComboBox_Range.setCurrentText(self.Range)
        self.QComboBox_Aperture.setCurrentText(self.Aperture)
        self.QComboBox_AutoZero.setCurrentText(self.AutoZero)
        self.QComboBox_InputZ.setCurrentText(self.inputZ)
        #Uptime and Downtime Not Setting Yet
        self.QLineEdit_UpTime.setText(self.UpTime)
        self.QLineEdit_DownTime.setText(self.DownTime)

        # Connect the button to close the window
        QPushButton_Widget.clicked.connect(self.close)

        # Connect the signals for changes in the fields (if necessary)
        self.QComboBox_Range.currentTextChanged.connect(self.RangeChanged)
        self.QComboBox_Aperture.currentTextChanged.connect(self.ApertureChanged)
        self.QComboBox_AutoZero.currentTextChanged.connect(self.AutoZeroChanged)
        self.QComboBox_InputZ.currentTextChanged.connect(self.InputZChanged)
        self.QLineEdit_UpTime.textEdited.connect(self.UpTimeChanged)
        self.QLineEdit_DownTime.textEdited.connect(self.DownTimeChanged)


    def RangeChanged(self, s):
        self.Range = s

    def ApertureChanged(self, s):
        self.Aperture = s

    def AutoZeroChanged(self, s):
        self.AutoZero = s

    def InputZChanged(self, s):
        self.inputZ = s

    def UpTimeChanged(self, s):
        self.UpTime = s

    def DownTimeChanged(self, s):
        self.DownTime = s

class AdvancedSetting_Current(QDialog):
    """This class is to configure the Advanced Settings when conducting current measurements,
    It prompts a secondary dialogue for users to customize more advanced parametes such as
    aperture, range, AutoZero, input impedance etc.
    """

    def __init__(self):
        """Method defining the signals, slots and widgets for Advaced Settings of Voltage Measurements"""
        super().__init__()
        self.setWindowTitle("Advanced Window (Current)")
        QPushButton_Widget = QPushButton()

        QPushButton_Widget.setText("Confirm")
        layout1 = QFormLayout()

        Desp1 = QLabel()
        Desp1.setText("DMM Settings:")
        Desp2 = QLabel()
        Desp2.setText("PSU Settings:")

        Desp1.setFont(desp_font)
        Desp2.setFont(desp_font)
        QLabel_Range = QLabel()
        QLabel_Aperture = QLabel()
        QLabel_AutoZero = QLabel()
        QLabel_Terminal = QLabel()
        QLabel_UpTime = QLabel()
        QLabel_DownTime = QLabel()

        QLabel_Range.setText("DC Voltage Range")
        QLabel_Aperture.setText("NPLC / PLC")
        QLabel_AutoZero.setText("Auto Zero Function")
        QLabel_Terminal.setText("Current Terminal:")
        QLabel_UpTime.setText("Programming Settling Time (UP) (in ms)")
        QLabel_DownTime.setText("Programming Setting Time (Down) (in ms)")

        QComboBox_Range = QComboBox()
        QComboBox_Aperture = QComboBox()
        QComboBox_AutoZero = QComboBox()
        QComboBox_Terminal = QComboBox()
        QLineEdit_UpTime = QLineEdit()
        QLineEdit_DownTime = QLineEdit()

        QComboBox_Range.addItems(["Auto", "0.001", "0.01", "0.1", "1", "3"])
        QComboBox_Aperture.addItems(
            ["0.001", "0.002", "0.006", "0.02", "0.06", "0.2", "1", "10", "100"]
        )
        QComboBox_AutoZero.addItems(["ON", "OFF"])
        QComboBox_Terminal.addItems(["3A", "10A"])

        layout1.addRow(Desp1)
        layout1.addRow(QLabel_Range, QComboBox_Range)
        layout1.addRow(QLabel_Aperture, QComboBox_Aperture)
        layout1.addRow(QLabel_AutoZero, QComboBox_AutoZero)
        layout1.addRow(QLabel_Terminal, QComboBox_Terminal)
        layout1.addRow(Desp2)
        layout1.addRow(QLabel_UpTime, QLineEdit_UpTime)
        layout1.addRow(QLabel_DownTime, QLineEdit_DownTime)
        layout1.addRow(QPushButton_Widget)
        self.setLayout(layout1)

        QPushButton_Widget.clicked.connect(self.close)
        QComboBox_Range.currentTextChanged.connect(self.RangeChanged)
        QComboBox_Aperture.currentTextChanged.connect(self.ApertureChanged)
        QComboBox_AutoZero.currentTextChanged.connect(self.AutoZeroChanged)
        QComboBox_Terminal.currentTextChanged.connect(self.TerminalChanged)
        QLineEdit_UpTime.textEdited.connect(self.UpTimeChanged)
        QLineEdit_DownTime.textEdited.connect(self.DownTimeChanged)

        self.QComboBox_Range = QComboBox_Range

    def RangeChanged(self, s):
        self.Range = s
        CurrentMeasurementDialog.setRange(self, self.Range)
        CC_LoadRegulationDialog.setRange(self, self.Range)

    def ApertureChanged(self, s):
        self.Aperture = s
        CurrentMeasurementDialog.setAperture(self, self.Aperture)
        CC_LoadRegulationDialog.setAperture(self, self.Aperture)

    def AutoZeroChanged(self, s):
        self.AutoZero = s
        CurrentMeasurementDialog.setAutoZero(self, self.AutoZero)
        CC_LoadRegulationDialog.setAutoZero(self, self.AutoZero)

    def TerminalChanged(self, s):
        self.Terminal = s
        CurrentMeasurementDialog.setTerminal(self, self.Terminal)
        CC_LoadRegulationDialog.setTerminal(self, self.Terminal)
        if self.Terminal == "10A":
            self.QComboBox_Range.setEnabled(False)

        else:
            self.QComboBox_Range.setEnabled(True)

    def UpTimeChanged(self, s):
        self.UpTime = s
        CurrentMeasurementDialog.setUpTime(self, self.UpTime)
        CC_LoadRegulationDialog.setUpTime(self, self.UpTime)

    def DownTimeChanged(self, s):
        self.DownTime = s
        CurrentMeasurementDialog.setDownTime(self, self.DownTime)
        CC_LoadRegulationDialog.setDownTime(self, self.DownTime)


#########----------------------- Image Graph Display ---------------------------------####################
"""class image_Window(QDialog):
    Class to display graph of DUT Test results

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Image")
        self.im = QPixmap(IMAGE_PATH)
        self.im2 = QPixmap(IMAGE_PATH_2)
        
        # Check if the image loaded successfully
        if self.im.isNull():
            QMessageBox.warning(self, "Error", "Failed to load image.")
            self.close()  # Close the window
            return
        if self.im2.isNull():
            QMessageBox.warning(self, "Error", "Failed to load image.")
            self.close()  # Close the window
            return
        
        self.label = QLabel()
        self.label.setPixmap(self.im)

        self.label2 = QLabel()
        self.label2.setPixmap(self.im2)
        
        self.grid = QGridLayout()
        self.grid.addWidget(self.label, 1, 1)
        self.grid.addWidget(self.label2, 1, 2)

        self.setLayout(self.grid)

        self.setWindowFlags(Qt.Window)
        self.setModal(False)  # Set the dialog to be non-modal
        self.show()
        # Ensure standard window behavior
        self.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint)

    def close_window(self):
        self.close()"""

from ui.all_test_dialog import (
    ACSourceSetting,
    AllTestMeasurement,
    ComboBoxWheelFilter,
    DiscoveryResult,
    ParameterSnapshot,
    Parameters,
    ScanSelectedVisaResources,
    TestCancelled,
    TestState,
    TestWorker,
    VoltageAccuracyPlotWindow,
    image_Window,
    image_Window2,
)
 

if __name__ == "__main__":
    original_stdout = sys.stdout
    my_result.seek(0)
    my_result.truncate(0)
    sys.stdout = my_result

    # # Create the application
    app = QApplication(sys.argv)

    win = MainWindow()
    win.setWindowFlags(Qt.Window | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
    win.show()  # Start the window in full screen mode

    sys.stdout = original_stdout
    sys.exit(app.exec())
