""" This module contains all of the data processing and visualization tools and functions

    The module mainly uses maltplotlib to plot graphs and pandas to process the data. 

"""
import os
import datetime
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from path import *
from SCPI_Library.IEEEStandard import IDN
from SCPI_Library.Keysight import System
from path import csv_folder, IMAGE_DIR, IMAGE_PATH, IMAGE_PATH_2

#------------------Instrument Data Collection---------------------
class instrumentData(object):
    """Stores instrument IDN and SCPI version, safely handling missing or invalid addresses."""

    def __init__(self, *args):
        instrumentIDN = []
        instrumentVersion = []

        for x in args:
            try:
                if x in (None, "None"):
                    instrumentIDN.append("N/A")
                    instrumentVersion.append("N/A")
                else:
                    instrumentIDN.append(IDN(x).query())
                    instrumentVersion.append(System(x).version())
            except Exception as e:
                instrumentIDN.append(f"Error: {e}")
                instrumentVersion.append(f"Error: {e}")

        df1 = pd.DataFrame(instrumentIDN, columns=["Instruments Used"])
        df2 = pd.DataFrame(instrumentVersion, columns=["SCPI Version"])
        instrument = pd.concat([df1, df2], axis=1)

        instrument.to_csv(INSTRUMENT_DATA_PATH, index=False)
        
class powerinstrumentData(object):
            """This class stores and facilitates the collection of Instrument Data to be placed in Excel Report

            Attributes:
                *args: arguments should contain strings of VISA Addresses of instruments used.
                instrumentIDN: List containing the Identification Name of the Instruments
                instrumentVersion: List containing the SCPI Version of the Instruments

            """

            def __init__(self, *args):
                instrumentIDN = []
                instrumentVersion = []

                for x in args:
                    instrumentIDN.append(IDN(x).query())
                    instrumentVersion.append(System(x).version())

                df1 = pd.DataFrame(instrumentIDN, columns=["Instruments Used: "])
                df2 = pd.DataFrame(instrumentVersion, columns=["SCPI Version"])

                ## Shift the data in `df1` or `df2` as needed
                #df1 = df1.shift(1)  # Shifts the entire `df1` down by one row
                #df2 = df2.shift(1)  # Shifts the entire `df2` down by one row

                # Concatenate the shifted dataframes
                instrument = pd.concat([df1, df2], axis=1)

                # Save to CSV
                instrument.to_csv(POWER_INSTRUMENT_DATA_PATH, index=False)

#------------------Voltage/Current Accuracy Data to Excel---------
#Voltage Measurements
class datatoCSV_Accuracy:
                """Preprocess voltage/current accuracy test data and export to CSV.

                Attributes:
                    infoList: List with program-collected data.
                    dataList: List with DUT-collected measurements.
                """

                def __init__(self, infoList, dataList, dataList2):
                    
                    """Initialize data processing and generate CSV file.

                    Args:
                        infoList: List with input program data.
                        dataList: List with measured DUT DMM data.
                        dataList2: List with measured DUT Readback
                    """
                                        

                    # Convert list columns to Series
                    #Voltage
                    Vset = pd.Series(self.column(infoList, 0))
                    #Current
                    Iset = pd.Series(self.column(infoList, 1))
                    #Progess
                    Key = pd.Series(self.column(infoList, 2))

                    #DMM reading
                    Vmeasured = pd.Series(self.column(dataList, 0))
                    Imeasured = pd.Series(self.column(dataList, 1))

                    # Calculate errors and handle division by zero
                    Vabsolute_error = Vmeasured - Vset
                    Iabsolute_error = Imeasured - Iset

                    #mV
                    ProgrammingV_error = Vabsolute_error 
                    #A
                    ProgrammingI_error = Iabsolute_error 

                    Vpercent_error = Vabsolute_error #programming percentage error
                    
                    Ipercent_error = Iabsolute_error

                    #PSU Readback Measurement
                    Vreadback = pd.Series(self.column(dataList2, 0))
                    Ireadback = pd.Series(self.column(dataList2, 1))



                    Vreadback_error = (Vreadback - Vmeasured) 
                    Ireadback_error = (Ireadback - Imeasured) 

                    Vreadback_percent_error = Vreadback_error



                    # Convert Series to DataFrames
                    columns = {
                        #Programming Accuracy
                        "PSU Readback Voltage": Vreadback, "PSU Readback Current" : Ireadback,
                        "PSU Voltage Set": Vset, "Load Current Set": Iset, "DMM Voltage Measured": Vmeasured,
                        "key": Key, "Programming/Voltage Absolute Error (V)": ProgrammingV_error,"Relative/Voltage Percentage Error (%)": Vpercent_error,   
                        "PSU Readback Voltage Error (V)": Vreadback_error, "PSU Readback Voltage Percentage Error (%)": Vreadback_percent_error,
                      
    
                        
                    }
                    CSV1 = pd.DataFrame(columns)

                    # Save to CSV
                    CSV1.to_csv(DATA_CSV_PATH, index=False)

                @staticmethod
                def column(matrix, i):
                    """Convert row data to a single column.

                    Args:
                        matrix: 2D data matrix.
                        i: Column index to extract.
                    """
                    return [row[i] if len(row) > i else None for row in matrix]

#Current Measurements
class datatoCSV_Accuracy2:
                """Preprocess voltage/current accuracy test data and export to CSV.

                Attributes:
                    infoList: List with program-collected data.
                    dataList: List with DUT-collected measurements.
                """
                
                def __init__(self, infoList, dataList, dataList2):
                    
                    """Initialize data processing and generate CSV file.

                    Args:
                        infoList: List with input program data.
                        dataList: List with measured DUT DMM data.
                        dataList2: List with measured DUT Readback
                    """
                    if os.path.exists(DATA_CSV_PATH):
                        os.remove(DATA_CSV_PATH)  # Delete if exists
                                        

                    # Convert list columns to Series
                    #Voltage
                    Vset = pd.Series(self.column(infoList, 0))
                    #Current
                    Iset = pd.Series(self.column(infoList, 1))
                    #Progess
                    Key = pd.Series(self.column(infoList, 2))

                    #DMM reading
                    Vmeasured = pd.Series(self.column(dataList, 0))
                    Imeasured = pd.Series(self.column(dataList, 1))

                    # Calculate errors and handle division by zero
                    Iabsolute_error = Imeasured - Iset
                    #A
                    ProgrammingI_error = Iabsolute_error 
                    
                    Ipercent_error = (Iabsolute_error / Iset.replace(0, float('nan'))) * 100
                    
                    #PSU Readback Measurement
                    Vreadback = pd.Series(self.column(dataList2, 0))
                    Ireadback = pd.Series(self.column(dataList2, 1))

                    Ireadback_error = (Ireadback - Imeasured) 



                    # Convert Series to DataFrames
                    columns = {
                    
                        "PSU Readback Voltage": Vreadback,"PSU Readback Current" : Ireadback,
                        "Load Voltage Set": Vset, "PSU Current Set": Iset,"DMM Current Measured": Imeasured, 
                        "key": Key, "Programming/Current Absolute Error (A)": ProgrammingI_error, "Relative/Current Percentage Error (%)": Ipercent_error, 
                        "PSU Readback Current Error (A)": Ireadback_error, "Relative/Current Percentage Error (%)": Ipercent_error

                    }
                    CSV1 = pd.DataFrame(columns)

                    # Save to CSV
                    CSV1.to_csv(DATA_CSV_PATH, index=False)

                @staticmethod
                def column(matrix, i):
                    """Convert row data to a single column.

                    Args:
                        matrix: 2D data matrix.
                        i: Column index to extract.
                    """
                    return [row[i] if len(row) > i else None for row in matrix]

#Power Measurements
class datatoCSV_PowerAccuracy:
                """Preprocess voltage/current accuracy test data and export to CSV.

                Attributes:
                    infoList: List with program-collected data.
                    dataList: List with DUT-collected measurements.
                """
           
                def __init__(self, infoList, dataList, dataList2):
                    
                    """Initialize data processing and generate CSV file.

                    Args:
                        infoList: List with input program data.
                        dataList: List with measured DUT DMM data.
                        dataList2: List with measured DUT Readback
                    """
                                        

                    # Convert list columns to Series
                    SpecError =1

                    #Progess
                    Key = pd.Series(self.column(infoList, 0) )
                    Pset = pd.Series(self.column(infoList, 1))

                    #DUT Readback
                    Preadback = pd.Series(self.column(dataList, 0))

                    #DMM reading
                    Vmeasured = pd.Series(self.column(dataList2, 0))
                    Imeasured = pd.Series(self.column(dataList2, 1))
                    Pmeasured = pd.Series(self.column(dataList2, 2))

                    # Calculate errors and handle division by zero
                    PowerProgramming_error = Pmeasured - Pset
                    ProgrammingV_error = PowerProgramming_error

                    Percent_error = (ProgrammingV_error / Pset.replace(0, float('nan'))) * 100 #SI unit in percentage
                    
                    OverallError = ProgrammingV_error / SpecError


                    Preadback_error = (Preadback - Pmeasured)
           

                    # Convert Series to DataFrames
                    columns = {
                        #Programming Accuracy
                        "key": Key, "PSU Set Power (W)": Pset, "DMM Voltage Measured (V)": Vmeasured, "DMM Current Measured (A)": Imeasured, "DMM Power Calculated (W)": Pmeasured,
                        "Programming Error (W)": ProgrammingV_error,"Percentage Error (%)": Percent_error,

                        #Readback Accuracy
                        "PSU Readback Power (W)": Preadback, 
                        "Readback Error (W)": Preadback_error
                    }
                    CSV1 = pd.DataFrame(columns)

                    # Save to CSV
                    CSV1.to_csv(POWER_DATA_CSV_PATH, index=False)

                @staticmethod
                def column(matrix, i):
                    """Convert row data to a single column.

                    Args:
                        matrix: 2D data matrix.
                        i: Column index to extract.
                    """
                    return [row[i] if len(row) > i else None for row in matrix]

#------------------Graph Plotting---------------------------------
class datatoGraph(datatoCSV_Accuracy):
                """Child class of datatoCSV_Accuracy to plot error boundaries for Voltage/Current accuracy testing"""

                def __init__(self, infoList, dataList, dataList2):
                    super().__init__(infoList, dataList, dataList2)
                    self.data = pd.read_csv(DATA_CSV_PATH)
                    
                    # Ensure that the images directory exists
                    if not os.path.exists(IMAGE_DIR):
                        os.makedirs(IMAGE_DIR)
                    
                def scatterCompareVoltage(self, param1, param2, param3, param4, unit, Vrated):
                    """Function to compare and scatter plot Voltage error boundaries for each current group."""
                    ungrouped_df = pd.read_csv(DATA_CSV_PATH)
                    grouped_df = ungrouped_df.groupby("key")
                    
                    # Initialize Series for each column that will be saved to CSV
                    upper_error_limitC = pd.Series(dtype="float64")
                    lower_error_limitC = pd.Series(dtype="float64")
                    conditionC = pd.Series(dtype="object")

                    upper_erro_percent_limitC = pd.Series(dtype="float64")
                    lower_erro_percent_limitC = pd.Series(dtype="float64")
                    condition_percentC = pd.Series(dtype="object")

                    upper_error_limitC2 = pd.Series(dtype="float64")
                    lower_error_limitC2 = pd.Series(dtype="float64")
                    conditionC2 = pd.Series(dtype="object")

                    upper_erro_percent_limitC2 = pd.Series(dtype="float64")
                    lower_erro_percent_limitC2 = pd.Series(dtype="float64")
                    condition2_percentC = pd.Series(dtype="object")

                    all_Vset = pd.Series(dtype="float64")
                    all_upper_error_limit = pd.Series(dtype="float64")
                    all_lower_error_limit = pd.Series(dtype="float64")
                    all_upper_error_limit2 = pd.Series(dtype="float64")
                    all_lower_error_limit2 = pd.Series(dtype="float64")

                    all_upper_percent_error_limit = pd.Series(dtype="float64")
                    all_lower_percent_error_limit = pd.Series(dtype="float64")
                    all_upper_percent_error_limit2 = pd.Series(dtype="float64")
                    all_lower_percent_error_limit2 = pd.Series(dtype="float64")

            
                                        # Create subplots
                    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))
                    fig_percent, (ax3, ax4) = plt.subplots(1, 2, figsize=(12, 6))

                    for key, group in grouped_df:
                        Vset = group["PSU Voltage Set"]
                        Iset = group["Load Current Set"]
                        ProgrammingV_error = group["Programming/Voltage Absolute Error (V)"]
                        Vreadback_error = group["PSU Readback Voltage Error (V)"]
                        ProgrammingV_percent_error = group["Relative/Voltage Percentage Error (%)"]
                        ReadbackV_percent_error = group["PSU Readback Voltage Percentage Error (%)"]

                        # Calculate boundaries and pass/fail condition
                        upper_error_limit = ((param1 * Vset) + param2)
                        lower_error_limit = -upper_error_limit
                        upper_error_limit2 = ((param3 * Vset) + param4)
                        lower_error_limit2 = -upper_error_limit2

                        #Calculate Percentage Error
                        ProgrammingV_percent_error = (ProgrammingV_error / upper_error_limit) * 100 
                        ReadbackV_percent_error    = (Vreadback_error / upper_error_limit) * 100

                        # Append to lists so they are saved later
                        self.ProgrammingV_percent_error_list.append(ProgrammingV_percent_error)
                        self.ReadbackV_percent_error_list.append(ReadbackV_percent_error)

                         # percentage limits (always ±100)
                        upper_erro_percent_limit = (upper_error_limit/upper_error_limit) * 100
                        lower_erro_percent_limit = (lower_error_limit/upper_error_limit)* 100
                        upper_erro_percent_limit2 = (upper_error_limit/upper_error_limit)* 100
                        lower_erro_percent_limit2 = (lower_error_limit/upper_error_limit)* 100

                        all_Vset = pd.concat([all_Vset, Vset])
                        all_upper_error_limit = pd.concat([all_upper_error_limit, upper_error_limit])
                        all_lower_error_limit = pd.concat([all_lower_error_limit, lower_error_limit])
                        all_upper_error_limit2 = pd.concat([all_upper_error_limit2, upper_error_limit2])
                        all_lower_error_limit2 = pd.concat([all_lower_error_limit2, lower_error_limit2])

                        all_upper_percent_error_limit = pd.concat([all_upper_percent_error_limit, upper_erro_percent_limit])
                        all_lower_percent_error_limit = pd.concat([all_lower_percent_error_limit, lower_erro_percent_limit])
                        all_upper_percent_error_limit2 = pd.concat([all_upper_percent_error_limit2, upper_erro_percent_limit2])
                        all_lower_percent_error_limit2 = pd.concat([all_lower_percent_error_limit2, lower_erro_percent_limit2])
                        

                        condition1 = ProgrammingV_error > upper_error_limit
                        condition2 = ProgrammingV_error < lower_error_limit
                        condition3 = Vreadback_error > upper_error_limit2
                        condition4 = Vreadback_error < lower_error_limit2

                        condition5 = ProgrammingV_percent_error > upper_erro_percent_limit
                        condition6 = ProgrammingV_percent_error < lower_erro_percent_limit
                        condition7 = ReadbackV_percent_error > upper_erro_percent_limit
                        condition8 = ReadbackV_percent_error < lower_erro_percent_limit


                        boolList = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition1, condition2)]
                        boolList2 = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition3, condition4)]

                        boolList_percent = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition5, condition6)]
                        boolList2_percent = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition7, condition8)]

                        condition_series = pd.Series(boolList)
                        condition_series2 = pd.Series(boolList2)

                        condition_series_percent = pd.Series(boolList_percent)
                        condition_series2_percent = pd.Series(boolList2_percent)

                        # Scatter plot setup based on pass/fail condition
                        color_condition = np.where(condition_series == "PASS", "black", "red")
                        size_condition = np.where(condition_series == "PASS", 6, 12)
                        alpha_condition = np.where(condition_series == "PASS", 0.6, 1)

                        # Plot for Programming Voltage Error
                        ax1.scatter(Vset, ProgrammingV_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax1.plot(Vset, ProgrammingV_error, label=f"Current = {Iset.iloc[0]}", linewidth=0.8)

                        # Plot for Percentage Error Boundaries
                        ax3.scatter(Vset, ProgrammingV_percent_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax3.plot(Vset, ProgrammingV_percent_error, label=f"Current = {Iset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """"for i, txt in enumerate(ProgrammingV_error):
                            ax1.annotate(f'{txt:.2f}', (Vset.iloc[i], ProgrammingV_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Plot for Readback Voltage Error
                        color_condition2 = np.where(condition_series2 == "PASS", "black", "red")
                        size_condition2 = np.where(condition_series2 == "PASS", 6, 12)
                        alpha_condition2 = np.where(condition_series2 == "PASS", 0.6, 1)

                        ax2.scatter(Vset, Vreadback_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)                     
                        ax2.plot(Vset, Vreadback_error, label=f"Current = {Iset.iloc[0]}", linewidth=0.8)

                        # Plot for Readback Percentage Error Boundaries
                        ax4.scatter(Vset, ReadbackV_percent_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)
                        ax4.plot(Vset, ReadbackV_percent_error, label=f"Current = {Iset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """for i, txt in enumerate(Vreadback_error):
                            ax2.annotate(f'{txt:.2f}', (Vset.iloc[i], Vreadback_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Collect results for CSV output
                        upper_error_limitC = pd.concat([upper_error_limitC, upper_error_limit])
                        lower_error_limitC = pd.concat([lower_error_limitC, lower_error_limit])
                        conditionC = pd.concat([conditionC, condition_series])

                        upper_error_limitC2 = pd.concat([upper_error_limitC2, upper_error_limit2])
                        lower_error_limitC2 = pd.concat([lower_error_limitC2, lower_error_limit2])
                        conditionC2 = pd.concat([conditionC2, condition_series2])

                        # Collect percentage results for CSV output
                        upper_erro_percent_limitC = pd.concat([upper_erro_percent_limitC, upper_erro_percent_limit])
                        lower_erro_percent_limitC = pd.concat([lower_erro_percent_limitC, lower_erro_percent_limit])
                        condition_percentC = pd.concat([condition_percentC, condition_series_percent])  

                        upper_erro_percent_limitC2 = pd.concat([upper_erro_percent_limitC2, upper_erro_percent_limit2])
                        lower_erro_percent_limitC2 = pd.concat([lower_erro_percent_limitC2, lower_erro_percent_limit2])
                        condition2_percentC = pd.concat([condition2_percentC, condition_series2_percent])
                    
            
                    ax1.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax2.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax3.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax4.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)

                    # Plot error boundaries and save plot
                    ax1.plot(all_Vset, all_upper_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax1.plot(all_Vset, all_lower_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax1.legend(loc="lower left", fontsize=6)
                    ax1.set_title(unit)
                    ax1.set_xlabel("Voltage (V)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax1.set_ylabel("Programming/Voltage Absolute Error (V)")

                    ax2.plot(all_Vset, all_upper_error_limit2, label="Upper Bound", color="red", linewidth=1)
                    ax2.plot(all_Vset, all_lower_error_limit2, label="Lower Bound", color="red", linewidth=1)
                    ax2.legend(loc="lower left", fontsize=6)
                    ax2.set_title(unit)
                    ax2.set_xlabel("Voltage (V)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax2.set_ylabel("PSU Readback Voltage Error (V)")
                    

                    ax3.plot(all_Vset, all_upper_percent_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax3.plot(all_Vset, all_lower_percent_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax3.legend(loc="lower left", fontsize=6)
                    ax3.set_title(f"{unit} Percentage Error")
                    ax3.set_xlabel("Voltage (V)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax3.set_ylabel("Relative/Voltage Percentage Error (%)")

                    ax4.plot(all_Vset, all_upper_percent_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax4.plot(all_Vset, all_lower_percent_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax4.legend(loc="lower left", fontsize=6)
                    ax4.set_title(f"{unit} Readback Percentage Error")
                    ax4.set_xlabel("Voltage (V)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax4.set_ylabel("PSU Readback Voltage Percentage Error (%)")
                    


                    # Save the error boundaries and conditions to CSV
                    
                    conditionFF = conditionC.reset_index(drop=True).to_frame(name="Programming Condition")
                    upper_error_limitF = upper_error_limitC.reset_index(drop=True).to_frame(name="Programming Upper Error Boundary (V)")
                    lower_error_limitF = lower_error_limitC.reset_index(drop=True).to_frame(name="Programming Lower Error Boundary (V)")

                    conditionFF2 = conditionC2.reset_index(drop=True).to_frame(name="Readback Condition")
                    upper_error_limitF2 = upper_error_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Error Boundary (V)")
                    lower_error_limitF2 = lower_error_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Error Boundary (V)")

                    conditionFF_percent = condition_percentC.reset_index(drop=True).to_frame(name="Programming Percentage Condition")
                    upper_error_limitF_percent = upper_erro_percent_limitC.reset_index(drop=True).to_frame(name="Programming Upper Percentage Error Boundary (%)")
                    lower_error_limitF_percent = lower_erro_percent_limitC.reset_index(drop=True).to_frame(name="Programming Lower Percentage Error Boundary (%)")

                    conditionFF2_percent = condition2_percentC.reset_index(drop=True).to_frame(name="Readback Percentage Condition")
                    upper_error_limitF2_percent = upper_erro_percent_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Percentage Error Boundary (%)")
                    lower_error_limitF2_percent = lower_erro_percent_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Percentage Error Boundary (%)")

                    # Drop the 'key' column from ungrouped_df
                    ungrouped_df.drop(columns=["key"], inplace=True)
                    # After the loop
                    ungrouped_df["Relative/Voltage Percentage Error (%)"] = pd.concat(self.ProgrammingV_percent_error_list).reset_index(drop=True)
                    ungrouped_df["PSU Readback Voltage Percentage Error (%)"] = pd.concat(self.ReadbackV_percent_error_list).reset_index(drop=True)


                    # Combine all DataFrames into a single DataFrame
                    combined_df = pd.concat([ungrouped_df, upper_error_limitF, lower_error_limitF, conditionFF, upper_error_limitF2, lower_error_limitF2, conditionFF2, upper_error_limitF_percent, lower_error_limitF_percent, conditionFF_percent, upper_error_limitF2_percent, lower_error_limitF2_percent, conditionFF2_percent], axis=1)
                    #combined_df2 = pd.concat([combined_df, upper_error_limitF_percent, lower_error_limitF_percent, conditionFF_percent, upper_error_limitF2_percent, lower_error_limitF2_percent, conditionFF2_percent], axis=1)

                    # Save the combined DataFrame to a CSV file
                    combined_df.to_csv(ERROR_CSV_PATH, index=False)
                    #combined_df2.to_csv(ERROR_CSV_PATH_PERCENT, index=False)
                    fig.savefig(IMAGE_PATH)
                    plt.close(fig)
                    fig_percent.savefig(IMAGE_PATH_2)
                    plt.close(fig_percent)
                
                def scatterCompareVoltage_Current_Change(self, param1, param2, param3, param4, unit, Vrated):
                    """Function to compare and scatter plot Voltage error boundaries for each current group."""
                    ungrouped_df = pd.read_csv(DATA_CSV_PATH)
                    grouped_df = ungrouped_df.groupby("key")
                    
                    # Initialize Series for each column that will be saved to CSV
                    upper_error_limitC = pd.Series(dtype="float64")
                    lower_error_limitC = pd.Series(dtype="float64")
                    conditionC = pd.Series(dtype="object")

                    upper_erro_percent_limitC = pd.Series(dtype="float64")
                    lower_erro_percent_limitC = pd.Series(dtype="float64")
                    condition_percentC = pd.Series(dtype="object")

                    upper_error_limitC2 = pd.Series(dtype="float64")
                    lower_error_limitC2 = pd.Series(dtype="float64")
                    conditionC2 = pd.Series(dtype="object")

                    upper_erro_percent_limitC2 = pd.Series(dtype="float64")
                    lower_erro_percent_limitC2 = pd.Series(dtype="float64")
                    condition2_percentC = pd.Series(dtype="object")

                    all_Iset = pd.Series(dtype="float64")
                    all_upper_error_limit = pd.Series(dtype="float64")
                    all_lower_error_limit = pd.Series(dtype="float64")
                    all_upper_error_limit2 = pd.Series(dtype="float64")
                    all_lower_error_limit2 = pd.Series(dtype="float64")

                    all_upper_percent_error_limit = pd.Series(dtype="float64")
                    all_lower_percent_error_limit = pd.Series(dtype="float64")
                    all_upper_percent_error_limit2 = pd.Series(dtype="float64")
                    all_lower_percent_error_limit2 = pd.Series(dtype="float64")

            
                                        # Create subplots
                    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))
                    fig_percent, (ax3, ax4) = plt.subplots(1, 2, figsize=(12, 6))

                    for key, group in grouped_df:
                        Vset = group["PSU Voltage Set"]
                        Iset = group["Load Current Set"]
                        ProgrammingV_error = group["Programming/Voltage Absolute Error (V)"]
                        Vreadback_error = group["PSU Readback Voltage Error (V)"]
                        ProgrammingV_percent_error = group["Relative/Voltage Percentage Error (%)"]
                        ReadbackV_percent_error = group["PSU Readback Voltage Percentage Error (%)"]

                        # Calculate boundaries and pass/fail condition
                        upper_error_limit = ((param1 * Vset) + param2)
                        lower_error_limit = -upper_error_limit
                        upper_error_limit2 = ((param3 * Vset) + param4)
                        lower_error_limit2 = -upper_error_limit2

                        #Calculate Percentage Error
                        ProgrammingV_percent_error = (ProgrammingV_percent_error / upper_error_limit)
                        ReadbackV_percent_error    = (Vreadback_error / upper_error_limit)

                         # percentage limits (always ±100)
                        upper_erro_percent_limit = upper_error_limit/upper_error_limit
                        lower_erro_percent_limit = lower_error_limit/upper_error_limit
                        upper_erro_percent_limit2 = upper_error_limit/upper_error_limit
                        lower_erro_percent_limit2 = lower_error_limit/upper_error_limit

                        all_Iset = pd.concat([all_Iset, Iset])
                        all_upper_error_limit = pd.concat([all_upper_error_limit, upper_error_limit])
                        all_lower_error_limit = pd.concat([all_lower_error_limit, lower_error_limit])
                        all_upper_error_limit2 = pd.concat([all_upper_error_limit2, upper_error_limit2])
                        all_lower_error_limit2 = pd.concat([all_lower_error_limit2, lower_error_limit2])

                        all_upper_percent_error_limit = pd.concat([all_upper_percent_error_limit, upper_erro_percent_limit])
                        all_lower_percent_error_limit = pd.concat([all_lower_percent_error_limit, lower_erro_percent_limit])
                        all_upper_percent_error_limit2 = pd.concat([all_upper_percent_error_limit2, upper_erro_percent_limit2])
                        all_lower_percent_error_limit2 = pd.concat([all_lower_percent_error_limit2, lower_erro_percent_limit2])
                        

                        condition1 = ProgrammingV_error > upper_error_limit
                        condition2 = ProgrammingV_error < lower_error_limit
                        condition3 = Vreadback_error > upper_error_limit2
                        condition4 = Vreadback_error < lower_error_limit2

                        condition5 = ProgrammingV_percent_error > upper_erro_percent_limit
                        condition6 = ProgrammingV_percent_error < lower_erro_percent_limit
                        condition7 = ReadbackV_percent_error > upper_erro_percent_limit
                        condition8 = ReadbackV_percent_error < lower_erro_percent_limit


                        boolList = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition1, condition2)]
                        boolList2 = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition3, condition4)]

                        boolList_percent = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition5, condition6)]
                        boolList2_percent = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition7, condition8)]

                        condition_series = pd.Series(boolList)
                        condition_series2 = pd.Series(boolList2)

                        condition_series_percent = pd.Series(boolList_percent)
                        condition_series2_percent = pd.Series(boolList2_percent)

                        # Scatter plot setup based on pass/fail condition
                        color_condition = np.where(condition_series == "PASS", "black", "red")
                        size_condition = np.where(condition_series == "PASS", 6, 12)
                        alpha_condition = np.where(condition_series == "PASS", 0.6, 1)

                        # Plot for Programming Voltage Error
                        ax1.scatter(Iset, ProgrammingV_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax1.plot(Iset, ProgrammingV_error, label=f"Voltage = {Vset.iloc[0]}", linewidth=0.8)

                        # Plot for Percentage Error Boundaries
                        ax3.scatter(Iset, ProgrammingV_percent_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax3.plot(Iset, ProgrammingV_percent_error, label=f"Voltage = {Vset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """"for i, txt in enumerate(ProgrammingV_error):
                            ax1.annotate(f'{txt:.2f}', (Vset.iloc[i], ProgrammingV_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Plot for Readback Voltage Error
                        color_condition2 = np.where(condition_series2 == "PASS", "black", "red")
                        size_condition2 = np.where(condition_series2 == "PASS", 6, 12)
                        alpha_condition2 = np.where(condition_series2 == "PASS", 0.6, 1)

                        ax2.scatter(Iset, Vreadback_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)                     
                        ax2.plot(Iset, Vreadback_error, label=f"Voltage = {Vset.iloc[0]}", linewidth=0.8)

                        # Plot for Readback Percentage Error Boundaries
                        ax4.scatter(Iset, ReadbackV_percent_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)
                        ax4.plot(Iset, ReadbackV_percent_error, label=f"Voltage = {Vset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """for i, txt in enumerate(Vreadback_error):
                            ax2.annotate(f'{txt:.2f}', (Vset.iloc[i], Vreadback_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Collect results for CSV output
                        upper_error_limitC = pd.concat([upper_error_limitC, upper_error_limit])
                        lower_error_limitC = pd.concat([lower_error_limitC, lower_error_limit])
                        conditionC = pd.concat([conditionC, condition_series])

                        upper_error_limitC2 = pd.concat([upper_error_limitC2, upper_error_limit2])
                        lower_error_limitC2 = pd.concat([lower_error_limitC2, lower_error_limit2])
                        conditionC2 = pd.concat([conditionC2, condition_series2])

                        # Collect percentage results for CSV output
                        upper_erro_percent_limitC = pd.concat([upper_erro_percent_limitC, upper_erro_percent_limitC])
                        lower_erro_percent_limitC = pd.concat([lower_erro_percent_limitC, lower_erro_percent_limitC])
                        condition_percentC = pd.concat([condition_percentC, condition_series_percent])  

                        upper_erro_percent_limitC2 = pd.concat([upper_erro_percent_limitC2, upper_erro_percent_limitC])
                        lower_erro_percent_limitC2 = pd.concat([lower_erro_percent_limitC2, lower_erro_percent_limitC])
                        condition2_percentC = pd.concat([condition2_percentC, condition_series2_percent])
                    
            
                    ax1.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax2.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax3.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax4.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)

                    # Plot error boundaries and save plot
                    ax1.plot(all_Iset, all_upper_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax1.plot(all_Iset, all_lower_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax1.legend(loc="lower left", fontsize=6)
                    ax1.set_title(unit)
                    ax1.set_xlabel("Current (A)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax1.set_ylabel("Programming/Voltage Absolute Error (V)")

                    ax2.plot(all_Iset, all_upper_error_limit2, label="Upper Bound", color="red", linewidth=1)
                    ax2.plot(all_Iset, all_lower_error_limit2, label="Lower Bound", color="red", linewidth=1)
                    ax2.legend(loc="lower left", fontsize=6)
                    ax2.set_title(unit)
                    ax2.set_xlabel("Current" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax2.set_ylabel("PSU Readback Voltage Error (V)")
                    

                    ax3.plot(all_Iset, all_upper_percent_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax3.plot(all_Iset, all_lower_percent_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax3.legend(loc="lower left", fontsize=6)
                    ax3.set_title(f"{unit} Percentage Error")
                    ax3.set_xlabel("Current (A)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax3.set_ylabel("Relative/Voltage Percentage Error (%)")

                    ax4.plot(all_Iset, all_upper_percent_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax4.plot(all_Iset, all_lower_percent_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax4.legend(loc="lower left", fontsize=6)
                    ax4.set_title(f"{unit} Readback Percentage Error")
                    ax4.set_xlabel("Current (A)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax4.set_ylabel("PSU Readback Voltage Percentage Error (%)")
                    


                    # Save the error boundaries and conditions to CSV
                    conditionFF = conditionC.reset_index(drop=True).to_frame(name="Programming Condition")
                    upper_error_limitF = upper_error_limitC.reset_index(drop=True).to_frame(name="Programming Upper Error Boundary (V)")
                    lower_error_limitF = lower_error_limitC.reset_index(drop=True).to_frame(name="Programming Lower Error Boundary (V)")

                    conditionFF2 = conditionC2.reset_index(drop=True).to_frame(name="Readback Condition")
                    upper_error_limitF2 = upper_error_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Error Boundary (V)")
                    lower_error_limitF2 = lower_error_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Error Boundary (V)")

                    conditionFF_percent = condition_percentC.reset_index(drop=True).to_frame(name="Programming Percentage Condition")
                    upper_error_limitF_percent = upper_erro_percent_limitC.reset_index(drop=True).to_frame(name="Programming Upper Percentage Error Boundary (%)")
                    lower_error_limitF_percent = lower_erro_percent_limitC.reset_index(drop=True).to_frame(name="Programming Lower Percentage Error Boundary (%)")

                    conditionFF2_percent = condition2_percentC.reset_index(drop=True).to_frame(name="Readback Percentage Condition")
                    upper_error_limitF2_percent = upper_erro_percent_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Percentage Error Boundary (%)")
                    lower_error_limitF2_percent = lower_erro_percent_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Percentage Error Boundary (%)")

                    # Drop the 'key' column from ungrouped_df
                    ungrouped_df.drop(columns=["key"], inplace=True)

                    # Combine all DataFrames into a single DataFrame
                    combined_df = pd.concat([ungrouped_df, upper_error_limitF, lower_error_limitF, conditionFF, upper_error_limitF2, lower_error_limitF2, conditionFF2], axis=1)
                    combined_df2 = pd.concat([combined_df, upper_error_limitF_percent, lower_error_limitF_percent, conditionFF_percent, upper_error_limitF2_percent, lower_error_limitF2_percent, conditionFF2_percent], axis=1)

                    # Save the combined DataFrame to a CSV file
                    combined_df.to_csv(ERROR_CSV_PATH, index=False)
                    combined_df2.to_csv(ERROR_CSV_PATH_PERCENT, index=False)
                    fig.savefig(IMAGE_PATH)
                    plt.close(fig)
                    fig_percent.savefig(IMAGE_PATH_2)
                    plt.close(fig_percent)


                    

class datatoGraph2(datatoCSV_Accuracy2):
                """Child class of datatoCSV_Accuracy to plot error boundaries for Voltage/Current accuracy testing"""

                def __init__(self, infoList, dataList, dataList2):
                    super().__init__(infoList, dataList, dataList2)
                    self.data = pd.read_csv(DATA_CSV_PATH)
                    
                    # Ensure that the images directory exists
                    if not os.path.exists(IMAGE_DIR):
                        os.makedirs(IMAGE_DIR)


                def scatterCompareCurrent2(self, param1, param2, param3, param4, unit, Irated):
                    """Function to compare and scatter plot Voltage error boundaries for each current group."""
                    ungrouped_df = pd.read_csv(DATA_CSV_PATH)
                    grouped_df = ungrouped_df.groupby("key")

                    # Initialize Series for each column that will be saved to CSV
                    upper_error_limitC = pd.Series(dtype="float64")
                    lower_error_limitC = pd.Series(dtype="float64")
                    conditionC = pd.Series(dtype="object")

                    upper_erro_percent_limitC = pd.Series(dtype="float64")
                    lower_erro_percent_limitC = pd.Series(dtype="float64")
                    condition_percentC = pd.Series(dtype="object")

                    upper_error_limitC2 = pd.Series(dtype="float64")
                    lower_error_limitC2 = pd.Series(dtype="float64")
                    conditionC2 = pd.Series(dtype="object")

                    upper_erro_percent_limitC2 = pd.Series(dtype="float64")
                    lower_erro_percent_limitC2 = pd.Series(dtype="float64")
                    condition2_percentC = pd.Series(dtype="object")

                    # Initialize Series for columns to save to CSV
                    upper_error_limitC = pd.Series(dtype="float64")
                    lower_error_limitC = pd.Series(dtype="float64")
                    conditionC = pd.Series(dtype="object")

                    upper_error_limitC2 = pd.Series(dtype="float64")
                    lower_error_limitC2 = pd.Series(dtype="float64")
                    conditionC2 = pd.Series(dtype="object")

                    all_Iset = pd.Series(dtype="float64")
                    all_upper_error_limit = pd.Series(dtype="float64")
                    all_lower_error_limit = pd.Series(dtype="float64")
                    all_upper_error_limit2 = pd.Series(dtype="float64")
                    all_lower_error_limit2 = pd.Series(dtype="float64")

                    all_upper_percent_error_limit = pd.Series(dtype="float64")
                    all_lower_percent_error_limit = pd.Series(dtype="float64")
                    all_upper_percent_error_limit2 = pd.Series(dtype="float64")
                    all_lower_percent_error_limit2 = pd.Series(dtype="float64")

                    # Create subplots
                    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))
                    fig_percent, (ax3, ax4) = plt.subplots(1, 2, figsize=(12, 6))
    
                    for key, group in grouped_df:
                        Vset = group["Load Voltage Set"]
                        Iset = group["PSU Current Set"]
                        ProgrammingI_error = group["Programming/Current Absolute Error (A)"]
                        Ireadback_error = group["PSU Readback Current Error (A)"]
                        ProgrammingI_percent_error = group["Relative/Current Percentage Error (%)"]
                        ReadbackI_percent_error = group["PSU Readback Current Percentage Error (%)"]

                        # Calculate boundaries and pass/fail condition
                        upper_error_limit = ((param1 * Iset) + (param2))
                        lower_error_limit = -upper_error_limit
                        upper_error_limit2 = ((param3 * Iset) + (param4)) 
                        lower_error_limit2 = -upper_error_limit2

                        #Calculate Percentage Error
                        ProgrammingI_percent_error = (ProgrammingI_error / upper_error_limit) * 100 
                        ReadbackI_percent_error    = (Ireadback_error / upper_error_limit) * 100

                        # Append to lists so they are saved later
                        self.ProgrammingI_percent_error_list.append(ProgrammingI_percent_error)
                        self.ReadbackI_percent_error_list.append(ReadbackI_percent_error)

                         # percentage limits (always ±100)
                        upper_erro_percent_limit = (upper_error_limit/upper_error_limit) * 100
                        lower_erro_percent_limit = (lower_error_limit/upper_error_limit)* 100
                        upper_erro_percent_limit2 = (upper_error_limit/upper_error_limit)* 100
                        lower_erro_percent_limit2 = (lower_error_limit/upper_error_limit)* 100

                        all_Iset = pd.concat([all_Iset, Iset])
                        all_upper_error_limit = pd.concat([all_upper_error_limit, upper_error_limit])
                        all_lower_error_limit = pd.concat([all_lower_error_limit, lower_error_limit])
                        all_upper_error_limit2 = pd.concat([all_upper_error_limit2, upper_error_limit2])
                        all_lower_error_limit2 = pd.concat([all_lower_error_limit2, lower_error_limit2])

                        all_upper_percent_error_limit = pd.concat([all_upper_percent_error_limit, upper_erro_percent_limit])
                        all_lower_percent_error_limit = pd.concat([all_lower_percent_error_limit, lower_erro_percent_limit])
                        all_upper_percent_error_limit2 = pd.concat([all_upper_percent_error_limit2, upper_erro_percent_limit2])
                        all_lower_percent_error_limit2 = pd.concat([all_lower_percent_error_limit2, lower_erro_percent_limit2])

                        condition1 = ProgrammingI_error > upper_error_limit
                        condition2 = ProgrammingI_error < lower_error_limit
                        condition3 = Ireadback_error > upper_error_limit2
                        condition4 = Ireadback_error < lower_error_limit2

                        condition5 = ProgrammingI_percent_error > upper_erro_percent_limit
                        condition6 = ProgrammingI_percent_error < lower_erro_percent_limit
                        condition7 = ReadbackI_percent_error > upper_erro_percent_limit
                        condition8 = ReadbackI_percent_error < lower_erro_percent_limit

                        boolList = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition1, condition2)]
                        boolList2 = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition3, condition4)]

                        boolList_percent = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition5, condition6)]
                        boolList2_percent = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition7, condition8)]

                        condition_series = pd.Series(boolList)
                        condition_series2 = pd.Series(boolList2)

                        condition_series_percent = pd.Series(boolList_percent)
                        condition_series2_percent = pd.Series(boolList2_percent)

                        # Scatter plot setup based on pass/fail condition
                        color_condition = np.where(condition_series == "PASS", "black", "red")
                        size_condition = np.where(condition_series == "PASS", 6, 12)
                        alpha_condition = np.where(condition_series == "PASS", 0.6, 1)

                        # Plot for Programming Voltage Error
                        ax1.scatter(Iset, ProgrammingI_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax1.plot(Iset, ProgrammingI_error, label=f"Voltage = {Vset.iloc[0]}", linewidth=0.8)

                        # Plot for Percentage Error Boundaries
                        ax3.scatter(Vset, ProgrammingI_percent_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax3.plot(Vset, ProgrammingI_percent_error, label=f"Current = {Iset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """"for i, txt in enumerate(ProgrammingV_error):
                            ax1.annotate(f'{txt:.2f}', (Vset.iloc[i], ProgrammingV_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Plot for Readback Voltage Error
                        color_condition2 = np.where(condition_series2 == "PASS", "black", "red")
                        size_condition2 = np.where(condition_series2 == "PASS", 6, 12)
                        alpha_condition2 = np.where(condition_series2 == "PASS", 0.6, 1)

                        ax2.scatter(Iset, Ireadback_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)
                        ax2.plot(Iset, Ireadback_error, label=f"Voltage = {Vset.iloc[0]}", linewidth=0.8)
                        
                         # Plot for Readback Percentage Error Boundaries
                        ax4.scatter(Vset, ReadbackI_percent_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)
                        ax4.plot(Vset, ReadbackI_percent_error, label=f"Current = {Iset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """for i, txt in enumerate(Vreadback_error):
                            ax2.annotate(f'{txt:.2f}', (Vset.iloc[i], Vreadback_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Collect results for CSV output
                        upper_error_limitC = pd.concat([upper_error_limitC, upper_error_limit])
                        lower_error_limitC = pd.concat([lower_error_limitC, lower_error_limit])
                        conditionC = pd.concat([conditionC, condition_series])

                        upper_error_limitC2 = pd.concat([upper_error_limitC2, upper_error_limit2])
                        lower_error_limitC2 = pd.concat([lower_error_limitC2, lower_error_limit2])
                        conditionC2 = pd.concat([conditionC2, condition_series2])

                        # Collect percentage results for CSV output
                        upper_erro_percent_limitC = pd.concat([upper_erro_percent_limitC, upper_erro_percent_limit])
                        lower_erro_percent_limitC = pd.concat([lower_erro_percent_limitC, lower_erro_percent_limit])
                        condition_percentC = pd.concat([condition_percentC, condition_series_percent])  

                        upper_erro_percent_limitC2 = pd.concat([upper_erro_percent_limitC2, upper_erro_percent_limit2])
                        lower_erro_percent_limitC2 = pd.concat([lower_erro_percent_limitC2, lower_erro_percent_limit2])
                        condition2_percentC = pd.concat([condition2_percentC, condition_series2_percent])
                    
                    ax1.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax2.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax3.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                    ax4.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)

                    # Plot error boundaries and save plot
                    ax1.plot(all_Iset, all_upper_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax1.plot(all_Iset, all_lower_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax1.legend(loc="lower left", fontsize=6)
                    ax1.set_title(unit)
                    ax1.set_xlabel("Current (A)")
                    ax1.set_ylabel("Programming/Current Absolute Error (A)")

                    ax2.plot(all_Iset, all_upper_error_limit2, label="Upper Bound", color="red", linewidth=1)
                    ax2.plot(all_Iset, all_lower_error_limit2, label="Lower Bound", color="red", linewidth=1)
                    ax2.legend(loc="lower left", fontsize=6)
                    ax2.set_title(unit)
                    ax2.set_xlabel("Current (A)")
                    ax2.set_ylabel("PSU Readback Current Error (A)")

                    ax3.plot(all_Iset, all_upper_percent_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax3.plot(all_Iset, all_lower_percent_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax3.legend(loc="lower left", fontsize=6)
                    ax3.set_title(f"{unit} Percentage Error")
                    ax3.set_xlabel("Voltage (V)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax3.set_ylabel("Relative/Current Percentage Error (%)")

                    ax4.plot(all_Iset, all_upper_percent_error_limit, label="Upper Bound", color="red", linewidth=1)
                    ax4.plot(all_Iset, all_lower_percent_error_limit, label="Lower Bound", color="red", linewidth=1)
                    ax4.legend(loc="lower left", fontsize=6)
                    ax4.set_title(f"{unit} Readback Percentage Error")
                    ax4.set_xlabel("Voltage (V)" if unit.upper() == "VOLTAGE" else "Current (A)")
                    ax4.set_ylabel("PSU Readback Current Percentage Error (%)")

                    # Save the error boundaries and conditions to CSV
                    conditionFF = conditionC.reset_index(drop=True).to_frame(name="Programming Condition")
                    upper_error_limitF = upper_error_limitC.reset_index(drop=True).to_frame(name="Programming Upper Error Boundary (A)")
                    lower_error_limitF = lower_error_limitC.reset_index(drop=True).to_frame(name="Programming Lower Error Boundary (A)")

                    conditionFF2 = conditionC2.reset_index(drop=True).to_frame(name="Readback Condition")
                    upper_error_limitF2 = upper_error_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Error Boundary (A)")
                    lower_error_limitF2 = lower_error_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Error Boundary (A)")

                    conditionFF_percent = condition_percentC.reset_index(drop=True).to_frame(name="Programming Percentage Condition")
                    upper_error_limitF_percent = upper_erro_percent_limitC.reset_index(drop=True).to_frame(name="Programming Upper Percentage Error Boundary (%)")
                    lower_error_limitF_percent = lower_erro_percent_limitC.reset_index(drop=True).to_frame(name="Programming Lower Percentage Error Boundary (%)")

                    conditionFF2_percent = condition2_percentC.reset_index(drop=True).to_frame(name="Readback Percentage Condition")
                    upper_error_limitF2_percent = upper_erro_percent_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Percentage Error Boundary (%)")
                    lower_error_limitF2_percent = lower_erro_percent_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Percentage Error Boundary (%)")
                    
                    # Drop the 'key' column from ungrouped_df
                    ungrouped_df.drop(columns=["key"], inplace=True)
                    # After the loop
                    ungrouped_df["Relative/Current Percentage Error (%)"] = pd.concat(self.ProgrammingI_percent_error_list).reset_index(drop=True)
                    ungrouped_df["PSU Readback Current Percentage Error (%)"] = pd.concat(self.ReadbackI_percent_error_list).reset_index(drop=True)

                    # Combine all DataFrames into a single DataFrame
                    combined_df = pd.concat([ungrouped_df, upper_error_limitF, lower_error_limitF, conditionFF, upper_error_limitF2, lower_error_limitF2, conditionFF2, upper_error_limitF_percent, lower_error_limitF_percent, conditionFF_percent, upper_error_limitF2_percent, lower_error_limitF2_percent, conditionFF2_percent], axis=1)

                    # Save the combined DataFrame to a CSV file
                    combined_df.to_csv(ERROR_CSV_PATH, index=False)

                    fig.savefig(IMAGE_PATH)
                    plt.close(fig)
                    fig_percent.savefig(IMAGE_PATH_2)
                    plt.close(fig_percent)

class datatoGraph3(datatoCSV_PowerAccuracy):
                """Child class of datatoCSV_Accuracy to plot error boundaries for Voltage/Current accuracy testing"""

                def __init__(self, infoList, dataList, dataList2):
                    super().__init__(infoList, dataList, dataList2)
                    self.data = pd.read_csv(POWER_DATA_CSV_PATH)
                    
                    # Ensure that the images directory exists
                    if not os.path.exists(IMAGE_DIR):
                        os.makedirs(IMAGE_DIR)

                def scatterComparePower(self, param1, param2, param3, param4, cccvmode, Prated):
                    """Function to compare and scatter plot Voltage error boundaries for each current group."""
                    ungrouped_df = pd.read_csv(POWER_DATA_CSV_PATH)
                    grouped_df = ungrouped_df.groupby("key")
                    
                    # Initialize Series for each column that will be saved to CSV
                    upper_error_limitC = pd.Series(dtype="float64")
                    lower_error_limitC = pd.Series(dtype="float64")
                    conditionC = pd.Series(dtype="object")
                    upper_error_limitC2 = pd.Series(dtype="float64")
                    lower_error_limitC2 = pd.Series(dtype="float64")
                    conditionC2 = pd.Series(dtype="object")

                    all_Pset = pd.Series(dtype="float64")
                    all_upper_error_limit = pd.Series(dtype="float64")
                    all_lower_error_limit = pd.Series(dtype="float64")
                    all_upper_error_limit2 = pd.Series(dtype="float64")
                    all_lower_error_limit2 = pd.Series(dtype="float64")
                    
            
                                        # Create subplots
                    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))

                    for key, group in grouped_df:
                        Pset = group["PSU Set Power (W)"]
                        Programming_error = group["Programming Error (W)"]
                        Preadback_error = group["Readback Error (W)"]

                        # Calculate boundaries and pass/fail condition
                        upper_error_limit = ((param1 * Pset) + param2 * Prated)
                        lower_error_limit = -upper_error_limit
                        upper_error_limit2 = ((param3 * Pset) + param4 * Prated)
                        lower_error_limit2 = -upper_error_limit2

                        all_Pset = pd.concat([all_Pset, Pset])
                        all_upper_error_limit = pd.concat([all_upper_error_limit, upper_error_limit])
                        all_lower_error_limit = pd.concat([all_lower_error_limit, lower_error_limit])
                        all_upper_error_limit2 = pd.concat([all_upper_error_limit2, upper_error_limit2])
                        all_lower_error_limit2 = pd.concat([all_lower_error_limit2, lower_error_limit2])

                        condition1 = Programming_error > upper_error_limit
                        condition2 = Programming_error < lower_error_limit
                        condition3 = Preadback_error > upper_error_limit2
                        condition4 = Preadback_error < lower_error_limit2

                        boolList = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition1, condition2)]
                        boolList2 = ["FAIL" if cond1 or cond2 else "PASS" for cond1, cond2 in zip(condition3, condition4)]

                        condition_series = pd.Series(boolList)
                        condition_series2 = pd.Series(boolList2)

                        # Scatter plot setup based on pass/fail condition
                        color_condition = np.where(condition_series == "PASS", "black", "red")
                        size_condition = np.where(condition_series == "PASS", 6, 12)
                        alpha_condition = np.where(condition_series == "PASS", 0.6, 1)

                        # Plot for Programming Voltage Error
                        ax1.scatter(Pset, Programming_error, color=color_condition, s=size_condition, alpha=alpha_condition)
                        ax1.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                        #ax1.plot(Pset, Programming_error, label=f"Power = {Pset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """"for i, txt in enumerate(ProgrammingV_error):
                            ax1.annotate(f'{txt:.2f}', (Vset.iloc[i], ProgrammingV_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Plot for Readback Voltage Error
                        color_condition2 = np.where(condition_series2 == "PASS", "black", "red")
                        size_condition2 = np.where(condition_series2 == "PASS", 6, 12)
                        alpha_condition2 = np.where(condition_series2 == "PASS", 0.6, 1)

                        ax2.scatter(Pset, Preadback_error, color=color_condition2, s=size_condition2, alpha=alpha_condition2)
                        ax2.axhline(y=0, color="grey", linestyle="--", linewidth=0.8)
                        #ax2.plot(Pset, Preadback_error, label=f"Power = {Pset.iloc[0]}", linewidth=0.8)

                        # Labeling data points
                        """for i, txt in enumerate(Vreadback_error):
                            ax2.annotate(f'{txt:.2f}', (Vset.iloc[i], Vreadback_error.iloc[i]), textcoords="offset points", xytext=(0, 10), ha='center')"""

                        # Collect results for CSV output
                        upper_error_limitC = pd.concat([upper_error_limitC, upper_error_limit])
                        lower_error_limitC = pd.concat([lower_error_limitC, lower_error_limit])
                        conditionC = pd.concat([conditionC, condition_series])

                        upper_error_limitC2 = pd.concat([upper_error_limitC2, upper_error_limit2])
                        lower_error_limitC2 = pd.concat([lower_error_limitC2, lower_error_limit2])
                        conditionC2 = pd.concat([conditionC2, condition_series2])

                    # Plot error boundaries and save plot
                    if cccvmode == "Current":
                        ax1.plot(all_Pset, all_upper_error_limit, label="Upper Bound", color="red", linewidth=1)
                        ax1.plot(all_Pset, all_lower_error_limit, label="Lower Bound", color="red", linewidth=1)
                        ax1.legend(loc="lower left", fontsize=6)

                        ax1.set_title(cccvmode)
                        ax1.set_xlabel(" Power (W)")
                        ax1.set_ylabel("Programming Error (W) _ CC")

                        ax2.plot(all_Pset, all_upper_error_limit2, label="Upper Bound", color="red", linewidth=1)
                        ax2.plot(all_Pset, all_lower_error_limit2, label="Lower Bound", color="red", linewidth=1)
                        ax2.legend(loc="lower left", fontsize=6)

                        ax2.set_title(cccvmode)
                        ax2.set_xlabel("Power (W)")
                        ax2.set_ylabel("Readback Error (W)_CC")
                    
                    elif cccvmode == "Voltage":
                        ax1.plot(all_Pset, all_upper_error_limit, label="Upper Bound", color="red", linewidth=1)
                        ax1.plot(all_Pset, all_lower_error_limit, label="Lower Bound", color="red", linewidth=1)
                        ax1.legend(loc="lower left", fontsize=6)

                        ax1.set_title(cccvmode)
                        ax1.set_xlabel(" Power (W)")
                        ax1.set_ylabel("Programming Error (W) _ CV")

                        ax2.plot(all_Pset, all_upper_error_limit2, label="Upper Bound", color="red", linewidth=1)
                        ax2.plot(all_Pset, all_lower_error_limit2, label="Lower Bound", color="red", linewidth=1)
                        ax2.legend(loc="lower left", fontsize=6)

                        ax2.set_title(cccvmode)
                        ax2.set_xlabel("Power (W)")
                        ax2.set_ylabel("Readback Error (W)_CV")

                    # Save the error boundaries and conditions to CSV
                    conditionFF = conditionC.reset_index(drop=True).to_frame(name="Programming Condition")
                    upper_error_limitF = upper_error_limitC.reset_index(drop=True).to_frame(name="Programming Upper Error Boundary (%)")
                    lower_error_limitF = lower_error_limitC.reset_index(drop=True).to_frame(name="Programming Lower Error Boundary (%)")

                    conditionFF2 = conditionC2.reset_index(drop=True).to_frame(name="Readback Condition")
                    upper_error_limitF2 = upper_error_limitC2.reset_index(drop=True).to_frame(name="Readback Upper Error Boundary (%)")
                    lower_error_limitF2 = lower_error_limitC2.reset_index(drop=True).to_frame(name="Readback Lower Error Boundary (%)")

                    # Drop the 'key' column from ungrouped_df
                    ungrouped_df.drop(columns=["key"], inplace=True)

                    # Combine all DataFrames into a single DataFrame
                    combined_df = pd.concat([ungrouped_df, upper_error_limitF, lower_error_limitF, conditionFF, upper_error_limitF2, lower_error_limitF2, conditionFF2], axis=1)

                    # Save the combined DataFrame to a CSV file
                    combined_df.to_csv(POWER_ERROR_CSV_PATH, index=False)

                    #plt.tight_layout()
                    #plt.show()
                    plt.savefig(POWER_IMAGE_PATH)
                    #plt.close(fig)


#--------------------Load/Line Regulation-------------------------
class datatoCSV_LoadRegulation:
    def __init__(self, Load_Reg_data, params):

        #Find Parameters
        DUT_Test_Model = params.get("PSU")
        Test = params.get("unit")
        save_location = params.get("savedir")

        self.Regulation_Data = Load_Reg_data[0]
        self.Model_Num = DUT_Test_Model.split("::")[-3]  # Gets the part before the last colon

        # Save path
        self.save_path = str(save_location)
        self.current_time = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

        if Test == "CURRENT":
            self.CC_Mode()
        elif Test == "VOLTAGE":
            self.CV_Mode()
        else:
            pass

    def CV_Mode(self):

        data = [
            ["Load Regulation Voltage (CV)"],
            ["Channel",  self.Regulation_Data[0]],
            ["Voltage Regulation Configuration for High Voltage Low Current",  self.Regulation_Data[1], "V",  self.Regulation_Data[2], "A"],
            ["V_No_Load_High Voltage Low Current",  self.Regulation_Data[3]], 
            ["V_Full_lLoad_High Voltage Low Current",  self.Regulation_Data[4]],
            ["Desired Voltage Regulation for High Voltage (CV) (V)",  self.Regulation_Data[5]],
            ["Calculated Voltage Regulation for High Voltage (V)",  self.Regulation_Data[6]],
            ["",""],
            ["Voltage Regulation Configuration for Low Voltage High Current",  self.Regulation_Data[7], "V",  self.Regulation_Data[8], "A"],
            ["V_No_Load_Low Voltage High Current",  self.Regulation_Data[9]],
            ["V_Full_Load_Low Voltage High Current",  self.Regulation_Data[10]],
            ["Desired Voltage Regulation for Low Voltage (CV) (V)",  self.Regulation_Data[11]],
            ["Calculated Voltage Regulation for Low Voltage (V)",  self.Regulation_Data[12]],
        ]

        # Save data to CSV file
        csv_file = os.path.join(self.save_path, f"Voltage_Regulation_{self.Model_Num}_{self.current_time}.csv")
        """with open(csv_file, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)
        print(f"✅ CV_Load Regulation CSV file saved to: {csv_file}")"""

        # Generate Excel sheet from CSV file
        excel_file = os.path.join(self.save_path, f"Voltage_Regulation_{self.current_time}.xlsx")
        df = pd.DataFrame(data)
        df.to_excel(excel_file, index=False, header=False)
        print(f"✅ CV_Load Regulation Excel file saved to: {excel_file}")

    def CC_Mode(self):
        data = [
            ["Load Regulation Current (CC)"],
            ["Channel",  self.Regulation_Data[0]],
            ["Current Regulation Configuration for High Current", self.Regulation_Data[1], "V", self.Regulation_Data[2], "A"],
            ["I_NL_High Current", self.Regulation_Data[3]], 
            ["I_FL_High Current", self.Regulation_Data[4]],
            ["Desired Load Regulation High Current (CC): (A)", self.Regulation_Data[5]],
            ["Calculated Load Regulation High Current (CC): (A)", self.Regulation_Data[6]],
            [""],
            ["Current Regulation Configuration for Low Current ", self.Regulation_Data[7], "V", self.Regulation_Data[8], "A"],
            ["I_NL_Low Current ", self.Regulation_Data[9]],
            ["I_FL_Low Current", self.Regulation_Data[10]],
            ["Desired Load Regulation Low Current (CC): (A)", self.Regulation_Data[11]],
            ["Calculated Load Regulation Low Current (CC): (A)", self.Regulation_Data[12]]
        ]

        # Save data to CSV file
        csv_file = self.save_path + f"/Current_Regulation_{self.Model_Num}_{self.current_time}.csv"
        """with open(csv_file, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)
        print(f"✅ CV_Load Regulation CSV file saved to: {csv_file}")"""

        # Generate Excel sheet from CSV file
        excel_file = os.path.join(self.save_path, f"Current_Regulation_{self.current_time}.xlsx")
        df = pd.DataFrame(data)
        df.to_excel(excel_file, index=False, header=False)
        print(f"✅ CV_Load Regulation Excel file saved to: {excel_file}")

class datatoCSV_Line_Regulation:
    #def __init__(self):
    def __init__(self, LineRegulationdata, params):

        #Find Parameters
        DUT_Test_Model = params.get("PSU")
        CV_CC_Test = params.get("unit")
        save_location = params.get("savedir")
        #DUT_Test_Model = "testfile"
        #CV_CC_Test = "VOLTAGE"
        #save_location = "C:\PyVisa - Copy  - Excavator - Copy\PyVisa\Manual Test - ZT\Excel"

        self.LineRegulationData = LineRegulationdata
        #self.LineRegulationData = [[1],[1,2,3,4,5,6,7,8,"PASS"],[1,2,3,4,5,6,7,8,9],[1,2,3,4,5,6,7,8,9]]
        self.channel = self.LineRegulationData[0]

        self.Model_Num = DUT_Test_Model.split("::")[-3]  # Gets the part before the last colon
        self.channel_data = f"Channel = {self.channel}"
        self.Data = self.LineRegulationData[1:]  #Skip first list

        # Save path
        self.save_path = str(save_location)
        self.current_time = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

        if CV_CC_Test == "CURRENT":
            self.CC_Mode()
        elif CV_CC_Test  == "VOLTAGE":
            self.CV_Mode()
        else:
            pass

    def CV_Mode(self):

        data_header = [
                "AC Line Voltage (Nominal)",
                "DUT High Voltage/ Low Voltage (V)",
                "ELoad Current Setting (A)",
                "V_LowLine (V)",
                "V_Nominal (V)",
                "V_HighLine (V)",  
                "Desired Voltage Regulation for High/Low Voltage (V)", 
                "Calculated Voltage Regulation for High/Low Voltage (V)", 
                "Condition"
                ]
        
        df = pd.DataFrame(self.Data, columns=data_header)
        excel_file = f"{self.save_path}/VoltLineRegulation_{self.Model_Num}_{self.current_time}.xlsx"
        #excel_file = f"{self.save_path}/VoltLineRegulation_TEST(NOTVALID)_{self.current_time}.xlsx"
        
        #Export to excel file
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            start_row = 2
            data_start_row = 3
            df.to_excel(writer, index=False, startrow=start_row)
            
            worksheet = writer.sheets['Sheet1']
            worksheet.cell(row=1, column=1).value = f"Channel: {self.channel}"

            # Define fills
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            # Find the "Condition" column index
            condition_col_idx = data_header.index('Condition') + 1  # +1 because openpyxl is 1-indexed

            # Apply conditional coloring
            for row in range(data_start_row , data_start_row  + len(df)):
                cell = worksheet.cell(row=row+1, column=condition_col_idx)
                if cell.value == "PASS":
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
        print(f"✅ CV_Line Regulation Excel file saved to: {excel_file}")

    def CC_Mode(self):
      
        data_header = [
                "AC Line Voltage (Nominal)",
                "DUT High Voltage/ Low Voltage (V)",
                "ELoad Current Setting (A)",
                "I_LowLine (A)",
                "I_Nominal (A)",
                "I_HighLine (A)",  
                "Desired Current Regulation for High/Low Voltage (A)", 
                "Calculated Current Regulation for High/Low Voltage (A)", 
                "Condition"
                ]
        
        df = pd.DataFrame(self.Data, columns=data_header)
        excel_file = f"{self.save_path}/CurrLineRegulation_{self.Model_Num}_{self.current_time}.xlsx"
        #excel_file = f"{self.save_path}/CurrLineRegulation_TEST(NOTVALID)_{self.current_time}.xlsx"
        
        #Export to excel file
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            start_row = 2
            data_start_row = 3
            df.to_excel(writer, index=False, startrow=start_row)
            
            worksheet = writer.sheets['Sheet1']
            worksheet.cell(row=1, column=1).value = f"Channel: {self.channel}"

            # Define fills
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            # Find the "Condition" column index
            condition_col_idx = data_header.index('Condition') + 1  # +1 because openpyxl is 1-indexed

            # Apply conditional coloring
            for row in range(data_start_row , data_start_row  + len(df)):
                cell = worksheet.cell(row=row+1, column=condition_col_idx)
                if cell.value == "PASS":
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
        print(f"✅ CC_Line Regulation Excel file saved to: {excel_file}")


#--------------------OVP/OCP Test-------------------------
class datatoCSV_OVP_Accuracy:
    def __init__(self, OVP_data, params):
        
        #Find DUT model
        DUT_Test_Model = params.get("PSU")
        save_location = params.get("savedir")
        Model_Num = DUT_Test_Model.split("::")[-3]  # Gets the part before the last colon

        # True enable all parameters to be export
        include_all_params = False
        selected_keys = ["V_Rating","I_Rating","P_Rating","OVP_Level","updatedelay","selected_DUT","PSU"]  # Used if include_all_params is False

        # === Filtered parameters ===
        if include_all_params:
            filtered_params = params
        else:
            filtered_params = {k: params[k] for k in selected_keys if k in params}
            
        # Data to be saved
        data_header = [
              "PSU Channel","Triggered OVP Level", "Voltage Set", "Readback Voltage", 
              "Lower Error Boundary", "Calculated OVP Error", "Upper Error Boundary", "Condition"]

        # Convert dict_info to DataFrame (str parameter info)
        df_info = pd.DataFrame(list(filtered_params.items()), columns=["Parameter", "Value"])
        df_info["Value"] = df_info["Value"].astype(str) 
        current_time = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        time_row = pd.DataFrame([["Time", current_time]], columns=["Parameter", "Value"])
        df_info = pd.concat([df_info, time_row], ignore_index=True)

        #Data obtained from OVP Test
        df_OVP = pd.DataFrame(OVP_data, columns=data_header)

        # Side-by-side Arrangement on Excel 
        if len(df_OVP) < len(df_info):
            pad_rows = len(df_info) - len(df_OVP)
            df_OVP = pd.concat([df_OVP, pd.DataFrame([[""] * len(df_OVP.columns)] * pad_rows, columns=df_OVP.columns)], ignore_index=True)

        empty_column = pd.DataFrame([""] * len(df_info), columns=[""])
        df_combined = pd.concat([df_info,empty_column, df_OVP], axis = 1)

        # === Save paths ===
        self.save_path = str(save_location)
        excel_file = f"{self.save_path}/OVP_{Model_Num}_{current_time}.xlsx"
        csv_file = f"{self.save_path}/OVP_{Model_Num}_{current_time}.csv"

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df_combined.to_excel(writer, index=False)

            worksheet = writer.sheets['Sheet1']
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            condition_col_idx = data_header.index('Condition') + 1 + len(df_info.columns) +1
             # +1 because openpyxl is 1-indexed

            for row in range(2 , 2  + len(df_OVP)):
                cell = worksheet.cell(row=row, column=condition_col_idx)
                if cell.value == "PASS":
                    cell.fill = green_fill
                elif cell.value == "FAIL":
                    cell.fill = red_fill
                else:
                     pass

        # === Save files ===
        #df_combined.to_excel(excel_file, index=False)
        df_combined.to_csv(csv_file, index=False)
        print(f"✅ OVP Test: Excel file saved to: {excel_file}")
        print(f"✅ OVP Test: CSV file saved to:   {csv_file}")

class datatoCSV_OCP_Test:
    def __init__(self, params):
        
        #Find DUT model
        DUT_Test_Model = params.get("PSU")
        self.save_path = str(params.get("savedir"))
        self.Model_Num = DUT_Test_Model.split("::")[-3]  # Gets the part before the last colon

        # True enable all parameters to be export
        include_all_params = False
        selected_keys = ["V_Rating","I_Rating","P_Rating","OCP_Level","updatedelay","selected_DUT","PSU"]  # Used if include_all_params is False

        # === Filtered parameters ===
        if include_all_params:
            filtered_params = params
        else:
            filtered_params = {k: params[k] for k in selected_keys if k in params}
        
        # Convert dict_info to DataFrame (str parameter info)
        self.df_info = pd.DataFrame(list(filtered_params.items()), columns=["Parameter", "Value"])
        self.df_info["Value"] = self.df_info["Value"].astype(str) 
        self.current_time = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        time_row = pd.DataFrame([["Time", self.current_time]], columns=["Parameter", "Value"])
        self.df_info = pd.concat([self.df_info, time_row], ignore_index=True)

    def AccuracyTest(self, OCP_data):

        # Data to be saved
        data_header = [
            "PSU Channel","Triggered OVP Level", "Current Set", "Readback Current", 
            "Calculated OCP Error"]
        df_OCP = pd.DataFrame(OCP_data, columns=data_header)
  
        # === Save paths for Accuracy Test ===
        excel_file = f"{self.save_path}/OCPAccuracy_{self.Model_Num}_{self.current_time}.xlsx"
        csv_file = f"{self.save_path}/OCPAccuracy_{self.Model_Num}_{self.current_time}.csv"

        #Excel file
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            # Write df_info on the left
            self.df_info.to_excel(writer, sheet_name="Test Results", startrow=0, startcol=0, index=False)

            # Write df_OCP on the right
            df_OCP.to_excel(writer, sheet_name="Test Results", startrow=0, \
                            startcol=len(self.df_info.columns) + 1, index=False)

        print(f"✅ OCP Test: Excel file saved to: {excel_file}")

        #CSV File
        empty_col = pd.DataFrame([""] * len(self.df_info), columns=[""])
        combined_df = pd.concat([self.df_info, empty_col,df_OCP], axis=1)
        combined_df.to_csv(csv_file, index=False)

        print(f"✅ OCP Test: CSV file saved to: {csv_file}")

    def ActivationTime(self,Activation_data):

        time_header = [
            "PSU Channel","OCP Delay Set (s)","PSU Voltage (V)","OCP Level (A)", \
            "Measured Activation Time (t/s)", "Time Limit", "Condition"
        ]
        df_activation = pd.DataFrame(Activation_data, columns=time_header)
        
         # === Save paths for Activation Time Test ===
        excel_file = f"{self.save_path}/OCPActivationTime_{self.Model_Num}_{self.current_time}.xlsx"
        csv_file = f"{self.save_path}/OCPActivationTime_{self.Model_Num}_{self.current_time}.csv"

        #Excel file
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            # Write df_info on the left
            self.df_info.to_excel(writer, sheet_name="Test Results", startrow=0, startcol=0, index=False)

            # Write df_OCP on the right
            df_activation.to_excel(writer, sheet_name="Test Results", startrow=0, \
                            startcol=len(self.df_info.columns) + 1, index=False)

        print(f"✅ OCP Activation: Excel file saved to: {excel_file}")

        #CSV File
        empty_col = pd.DataFrame([""] * len(self.df_info), columns=[""])
        combined_df = pd.concat([self.df_info, empty_col,df_activation], axis=1)
        combined_df.to_csv(csv_file, index=False)

        print(f"✅ OCP Activation: CSV file saved to: {csv_file}")

class datatoCSV_Programming_Response:
    #def __init__(self):
    def __init__(self, ProgrammingResponsendata, CurrentTime, params):

        #Find Parameters
        DUT_Test_Model = params.get("PSU")
        save_location = params.get("savedir")

        self.ProgrammingResponseData = ProgrammingResponsendata

        self.Model_Num = DUT_Test_Model.split("::")[-3]  # Gets the part before the last colon

        # Save path
        self.save_path = str(save_location)
        self.current_time = CurrentTime

        data_header = [
                "Channel",
                "Test Points",
                "Responses",
                "Initial Voltage (Vo)",
                "Final Voltage (Vf)",
                "ELoad Max Current (A)",
                "Results (t/s)",
                "Limit (t/s)",
                "Condition"
                ]
        
        df = pd.DataFrame(self.ProgrammingResponseData, columns=data_header)
        excel_file = f"{self.save_path}/ProgrammingResponse_{self.Model_Num}_{self.current_time}.xlsx"

        #Export to excel file
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            start_row = 1
            data_start_row = 2
            df.to_excel(writer, index=False, startrow=start_row)
            
            """
            # Define fills
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            # Find the "Condition" column index
            condition_col_idx = data_header.index('Condition') + 1  # +1 because openpyxl is 1-indexed

            # Apply conditional coloring
            for row in range(data_start_row , data_start_row  + len(df)):
                cell = worksheet.cell(row=row+1, column=condition_col_idx)
                if cell.value == "PASS":
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            """
        print(f"✅ Programming Responses Excel file saved to: {excel_file}")

        # Load the existing workbook
        image_folder = os.path.join(self.save_path, f"ProgrammingSpeed_{self.current_time}")

        wb = load_workbook(excel_file)

        # Get or create the sheet where you want to insert the images
        sheet = wb.active  # or wb["YourSheetName"]

        image_files = sorted(f for f in os.listdir(image_folder) if f.lower().endswith(".png"))

        # Starting row (e.g. row 20) and row spacing
        start_row = 20
        row_spacing = 20

        # Loop through all image files
        for i, image_name in enumerate(image_files):
            image_path = os.path.join(image_folder, image_name)

            try:
                img = ExcelImage(image_path)
                cell_location = f"A{start_row + i * row_spacing}"
                sheet.add_image(img, cell_location)
            except Exception as e:
                print(f"Failed to insert {image_name}: {e}")

        # Save workbook
        output_excel_file = os.path.join(image_folder, "updated_template.xlsx")
        wb.save(output_excel_file)

        print(f"All images inserted and Excel file saved at: {output_excel_file}")

